"""
Telegram-бот для онлайн-магазина цифровых товаров и доната
----------------------------------------------------------
Стек: Python 3.11+, aiogram 3, aiosqlite (SQLite-хранилище)

УСТАНОВКА:
    pip install aiogram==3.13.1 aiosqlite

ЗАПУСК:
    python bot.py

Все настройки и тексты находятся в блоке CONFIG ниже —
смело редактируйте их под себя.
"""

import asyncio
import io
import logging
import re
import secrets
from datetime import datetime, timedelta, timezone

MSK_TZ = timezone(timedelta(hours=3))


def _fmt_msk(iso_str: str) -> str:
    """Преобразует ISO-время (UTC) в строку в МСК."""
    if not iso_str:
        return ""
    s = iso_str.replace("T", " ")
    try:
        dt = datetime.fromisoformat(iso_str)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        dt_msk = dt.astimezone(MSK_TZ)
        return dt_msk.strftime("%Y-%m-%d %H:%M МСК")
    except Exception:
        return s
from html import escape

import os

import asyncpg
from aiogram import BaseMiddleware, Bot, Dispatcher, F
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.base import StorageKey
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    ErrorEvent,
    FSInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)

# =====================================================================
# CONFIG  --  ОТРЕДАКТИРУЙТЕ ЭТИ ЗНАЧЕНИЯ
# =====================================================================

# --- Основное ---
BOT_TOKEN = os.environ["BOT_TOKEN"]

# Telegram ID модератора (узнайте у @userinfobot).
# Сюда будут приходить уведомления о новых заказах и сообщения от клиентов.
# Если оставить 0 — пересылка будет отключена.
MODERATOR_CHAT_ID = 1727614596

# Юзернеймы (везде, где упоминается @username)
SUPPORT_USERNAME = "@cheeze0729"  # Поддержка
OTHER_APPS_USERNAME = "@cheeze0729"  # Донат в других приложениях
BRAWL_PROMO_USERNAME = "@cheeze0729"  # Акции по Brawl Stars

# Стартовое изображение (URL или file_id Telegram).
# Можно заменить на свою картинку.
WELCOME_IMAGE_URL = "https://i.postimg.cc/SxHKrmbd/IMG-3992.png"

# Реквизиты для пополнения баланса (карта/СБП).
# Эти реквизиты будут показаны пользователю после ввода суммы.
# Поддерживается HTML-разметка: <code>...</code> делает текст кликабельным
# для копирования в Telegram.
CARD_DETAILS = (
    "💳 <b>Карта Т-Банк:</b>\n"
    "<code>2200 7017 3823 7798</code>\n\n"
    "👤 <b>Получатель:</b> Марсель С.\n\n"
    "ℹ️ При невозможности оплаты по карте обратитесь в поддержку для оплаты по СБП."
)

# Ссылка на отзывы (канал, чат или страница с отзывами клиентов)
REVIEWS_URL = "https://t.me/cheezereviews"
# Username канала для публикации отзывов (бот должен быть администратором канала)
REVIEWS_CHANNEL = "@cheezereviews"

# =====================================================================
# КАРТИНКИ ДЛЯ КАЖДОГО РАЗДЕЛА
# =====================================================================
# Чтобы добавить картинку — впишите ссылку (URL) или Telegram file_id
# вместо None. Картинка будет показана над текстом раздела.
# Если оставить None — раздел будет без картинки.
# Подсказка: длина текста с картинкой не должна превышать 1024 символа.

SECTION_IMAGES: dict[str, str | None] = {
    "welcome": WELCOME_IMAGE_URL,  # /start, главное меню
    "shop": "images/shop.png",  # «Купить донат» (локальный файл)
    "roblox_instant": "https://i.ibb.co/0R5v4qFd/A48-CAFDF-8-C92-4-CA1-9-C68-E462-DD8-A34-D9.png",  # Roblox моментально
    "roblox_gamepass": "https://i.ibb.co/0R5v4qFd/A48-CAFDF-8-C92-4-CA1-9-C68-E462-DD8-A34-D9.png",  # Roblox геймпассом
    "brawl": "https://i.ibb.co/xqp3GzrQ/F42684-C7-1395-43-EF-A49-B-E817-B7-E73551.png",  # Brawl Stars
    "tgstars": "https://i.ibb.co/N2JJ5SHb/42-E499-C6-7-FE1-4-E03-94-A4-47-E6-A705-FEAA.png",  # Telegram Stars
    "other": None,  # Другие приложения
    "profile": "https://i.ibb.co/5x2bNNJK/56-D6214-E-F8-F0-4-BFE-AF06-EBA573966-CD0.png",  # Профиль
    "orders": "https://i.ibb.co/Gvpsmbnq/886-D71-D9-F7-B0-4-A9-E-A540-180-FFB450321.png",  # Мои заказы
    "topup": None,  # Пополнение баланса
    "support": "https://i.ibb.co/pvcyff13/13-B5-AB68-5-FDA-4598-B1-D7-12-F98-A0188-D7.png",  # Поддержка
    "info": "https://i.ibb.co/n8gT4YRD/4-AF7623-F-D966-4130-9-D7-F-592-ED9-C22935.png",  # Информация о магазине
    "guarantee": None,  # Гарантия
    # Карточки товаров (общие для всех товаров категории):
    "card_roblox_instant": None,
    "card_brawl": None,
    # Динамические категории из админки:
    "pubg": "images/pubg.png",
}

# Курсы и лимиты теперь хранятся в таблице settings (управляются через /admin → Каталог и курсы).
# Начальные значения задаются в db_init() и применяются только при первом запуске.

# =====================================================================
# РЕФЕРАЛЬНАЯ ПРОГРАММА — настройки
# =====================================================================
REFERRAL_LEVELS = {
    0: ("Новичок",      0),
    1: ("🥉 Приятель",  2),
    2: ("🥈 Знакомый",  4),
    3: ("🥇 Партнёр",   6),
    4: ("💎 Советник",  8),
    5: ("👑 Легенда",  10),
}
REFERRAL_REFS_PER_LEVEL     = 2  # рефералов для перехода на следующий уровень
REFERRAL_DISCOUNT_PER_LEVEL = 2  # % скидки за каждый уровень
REFERRAL_PROMOS_PER_LEVELUP = 3  # промокодов за повышение уровня

# --- Тексты ---
WELCOME_TEXT = (
    "<b>Добро пожаловать в магазин цифровых товаров.</b>\n\n"
    "Здесь вы можете быстро и удобно купить донат, "
    "пополнить баланс и связаться с поддержкой."
)

INFO_TEXT = (
    "<b>О магазине</b>\n\n"
    "Мы предлагаем удобную покупку цифровых товаров и доната "
    "по выгодным ценам.\n\n"
    "<b>Почему выбирают нас:</b>\n"
    "— Быстрое оформление заказов\n"
    "— Удобное пополнение баланса\n"
    "— Поддержка клиентов\n"
    "— Широкий выбор услуг\n\n"
    "<b>Время работы поддержки:</b>\n"
    "Ежедневно: 10:00 – 22:00 МСК\n\n"
    f"По всем вопросам: {SUPPORT_USERNAME}"
)

SUPPORT_TEXT = (
    "<b>Поддержка</b>\n\n"
    "Если у вас возникли вопросы, напишите в поддержку: "
    f"{SUPPORT_USERNAME}"
)

GUARANTEE_TEXT = (
    "<b>🛡️ Наша гарантия</b>\n\n"
    "Почему нам можно доверять:\n"
    "— Работаем официально и долго\n"
    "— Тысячи довольных клиентов\n"
    "— Деньги возвращаются, если заказ не выполнен\n"
    "— Поддержка отвечает каждый день\n"
    "— Прозрачные цены без скрытых комиссий\n\n"
    "Все отзывы реальных покупателей доступны по кнопке ниже."
)

# =====================================================================
# Подсказки по данным для входа (для каждой категории — своя)
# =====================================================================
# Если значение задано — после оплаты бот попросит покупателя
# отправить эти данные сообщением, сохранит их к заказу и перешлёт
# модератору. Если оставить None — данные не запрашиваются.

# login_hint для категорий со специальными хендлерами (не управляются через DB)
LOGIN_HINTS = {
    "roblox_gamepass": "Отправьте ссылку на созданный геймпасс в Roblox.",
    "tgstars": "Отправьте @username аккаунта Telegram для зачисления звёзд.",
}
# Стандартные категории из DB теперь хранят login_hint и needs_code в таблице categories

# =====================================================================
# КАТАЛОГ ТОВАРОВ
# Каталог товаров теперь хранится в таблице products (управляется через /admin → Каталог и курсы).
# Начальные товары добавляются в db_init() и применяются только при первом запуске.

# =====================================================================
# Инициализация бота
# =====================================================================

logging.basicConfig(level=logging.INFO)
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# Пользователи, для которых прямо сейчас обрабатывается платёж —
# защита от двойного списания при повторном/двойном нажатии кнопки оплаты.
_processing_payments: set[int] = set()


class BlacklistMiddleware(BaseMiddleware):
    """Блокирует все действия пользователей из чёрного списка.
    Модератор всегда пропускается. Покупателю показывается сообщение
    с кнопкой связи с поддержкой.
    """

    async def __call__(self, handler, event, data):
        user = getattr(event, "from_user", None)
        if user is None:
            return await handler(event, data)
        if MODERATOR_CHAT_ID and user.id == MODERATOR_CHAT_ID:
            return await handler(event, data)
        try:
            blocked = await db_is_blacklisted(user.id)
        except Exception:
            blocked = False
        if not blocked:
            return await handler(event, data)
        text = (
            "🚫 <b>Вы в чёрном списке.</b>\n\n"
            "Доступ к боту ограничен. Если считаете, что это ошибка — "
            "свяжитесь с поддержкой."
        )
        kb = kb_support()
        try:
            from aiogram.types import CallbackQuery as _CQ, Message as _MSG

            if isinstance(event, _CQ):
                await event.answer("Вы в чёрном списке", show_alert=True)
                try:
                    await event.message.answer(text, reply_markup=kb, parse_mode="HTML")
                except Exception:
                    pass
            elif isinstance(event, _MSG):
                await event.answer(text, reply_markup=kb, parse_mode="HTML")
        except Exception as e:
            logging.warning(f"BlacklistMiddleware error: {e}")
        return None


dp.message.middleware(BlacklistMiddleware())
dp.callback_query.middleware(BlacklistMiddleware())

# =====================================================================
# База данных (PostgreSQL через asyncpg + Neon)
# =====================================================================

_pool: asyncpg.Pool | None = None


async def get_pool() -> asyncpg.Pool:
    """Возвращает глобальный пул соединений с Neon PostgreSQL."""
    global _pool
    if _pool is None:
        _pool = await asyncpg.create_pool(
            os.environ["CONNECTION_STRING"],
            min_size=1,
            max_size=5,
        )
    return _pool


async def db_init() -> None:
    """Создаёт таблицы в PostgreSQL, если их ещё нет."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                tg_id          BIGINT PRIMARY KEY,
                username       TEXT,
                first_name     TEXT,
                balance        INTEGER NOT NULL DEFAULT 0,
                is_blacklisted BOOLEAN NOT NULL DEFAULT FALSE,
                created_at     TEXT NOT NULL
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id         SERIAL PRIMARY KEY,
                tg_id      BIGINT NOT NULL,
                title      TEXT NOT NULL,
                price      INTEGER NOT NULL,
                status     TEXT NOT NULL,
                category   TEXT,
                contact    TEXT,
                login_data TEXT,
                login_code TEXT,
                created_at TEXT NOT NULL
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id         SERIAL PRIMARY KEY,
                tg_id      BIGINT NOT NULL,
                amount     INTEGER NOT NULL,
                kind       TEXT NOT NULL,
                reason     TEXT,
                created_at TEXT NOT NULL
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS reviews (
                id            SERIAL PRIMARY KEY,
                order_id      INTEGER NOT NULL UNIQUE,
                tg_id         BIGINT NOT NULL,
                photo_file_id TEXT,
                comment       TEXT,
                created_at    TEXT NOT NULL
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS promo_codes (
                id            SERIAL PRIMARY KEY,
                code          TEXT UNIQUE NOT NULL,
                game          TEXT NOT NULL,
                product_title TEXT NOT NULL,
                promo_price   INTEGER NOT NULL,
                starts_at     TEXT,
                expires_at    TEXT,
                is_active     BOOLEAN NOT NULL DEFAULT TRUE,
                created_at    TEXT NOT NULL
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS user_promos (
                id         SERIAL PRIMARY KEY,
                tg_id      BIGINT NOT NULL,
                promo_id   INTEGER NOT NULL,
                claimed_at TEXT NOT NULL,
                used_at    TEXT,
                UNIQUE(tg_id, promo_id)
            )
        """)
        await conn.execute(
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS referral_code TEXT"
        )
        await conn.execute(
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS referred_by BIGINT"
        )
        await conn.execute(
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS referral_level "
            "INTEGER NOT NULL DEFAULT 0"
        )
        await conn.execute(
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS referral_count "
            "INTEGER NOT NULL DEFAULT 0"
        )
        await conn.execute(
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS active_ref_promo INTEGER"
        )
        await conn.execute(
            "ALTER TABLE promo_codes ADD COLUMN IF NOT EXISTS "
            "discount_pct INTEGER NOT NULL DEFAULT 0"
        )
        await conn.execute(
            "ALTER TABLE promo_codes ADD COLUMN IF NOT EXISTS "
            "max_uses INTEGER DEFAULT NULL"
        )
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS referral_purchases (
                id          SERIAL PRIMARY KEY,
                buyer_id    BIGINT NOT NULL,
                referrer_id BIGINT NOT NULL,
                created_at  TEXT NOT NULL,
                UNIQUE(buyer_id)
            )
        """)
        # Каталог товаров
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id         SERIAL PRIMARY KEY,
                category   TEXT NOT NULL,
                key        TEXT NOT NULL UNIQUE,
                name       TEXT NOT NULL,
                price      NUMERIC(12,2) NOT NULL,
                delivery   TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0,
                active     BOOLEAN NOT NULL DEFAULT TRUE
            )
        """)
        # Настройки (курсы, лимиты)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        """)
        # Заполняем настройки начальными значениями
        await conn.execute("""
            INSERT INTO settings (key, value) VALUES
                ('robux_gamepass_rate',           '0.65'),
                ('robux_gamepass_pass_price_rate','0.7'),
                ('tg_stars_rate',                 '1.3'),
                ('min_topup',                     '10'),
                ('min_tg_stars',                  '50')
            ON CONFLICT (key) DO NOTHING
        """)
        # Таблица категорий (для динамического магазина)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                key             TEXT PRIMARY KEY,
                name            TEXT NOT NULL,
                emoji           TEXT NOT NULL DEFAULT '🎮',
                sort_order      INTEGER NOT NULL DEFAULT 0,
                disabled        BOOLEAN NOT NULL DEFAULT FALSE,
                disabled_reason TEXT DEFAULT NULL,
                login_hint      TEXT DEFAULT NULL,
                needs_code      BOOLEAN NOT NULL DEFAULT FALSE
            )
        """)
        await conn.execute("""
            INSERT INTO categories (key, name, emoji, sort_order, disabled, login_hint, needs_code) VALUES
                ('roblox_instant', 'Roblox — моментально', '🟦', 0, FALSE,
                 'Отправьте логин (никнейм) в Roblox.', TRUE),
                ('brawl', 'Brawl Stars', '⭐', 1, FALSE,
                 'Отправьте почту, привязанную к вашему аккаунту и ожидайте код.', TRUE)
            ON CONFLICT (key) DO NOTHING
        """)
        # Миграция: добавляем новые колонки категорий если таблица уже была
        for col_sql in [
            "ALTER TABLE categories ADD COLUMN IF NOT EXISTS disabled_reason TEXT DEFAULT NULL",
            "ALTER TABLE categories ADD COLUMN IF NOT EXISTS login_hint TEXT DEFAULT NULL",
            "ALTER TABLE categories ADD COLUMN IF NOT EXISTS needs_code BOOLEAN NOT NULL DEFAULT FALSE",
        ]:
            await conn.execute(col_sql)
        # Заполняем каталог начальными товарами (не перезаписываем существующие)
        await conn.execute("""
            INSERT INTO products (category, key, name, price, delivery, sort_order) VALUES
                ('roblox_instant','rb_40',   '40 робуксов',   79,   'моментально', 0),
                ('roblox_instant','rb_80',   '80 робуксов',   99,   'моментально', 1),
                ('roblox_instant','rb_200',  '200 робуксов',  279,  'моментально', 2),
                ('roblox_instant','rb_400',  '400 робуксов',  459,  'моментально', 3),
                ('roblox_instant','rb_500',  '500 робуксов',  499,  'моментально', 4),
                ('roblox_instant','rb_1000', '1000 робуксов', 909,  'моментально', 5),
                ('roblox_instant','rb_1700', '1700 робуксов', 1619, 'моментально', 6),
                ('roblox_instant','rb_2000', '2000 робуксов', 1819, 'моментально', 7),
                ('roblox_instant','rb_3600', '3600 робуксов', 3299, 'моментально', 8),
                ('brawl','bs_pass',      'Brawl Pass',      899,  'по согласованию через Telegram', 0),
                ('brawl','bs_pass_plus', 'Brawl Pass Plus', 1239, 'по согласованию через Telegram', 1),
                ('brawl','bs_pro',       'Pro Pass',        2249, 'по согласованию через Telegram', 2)
            ON CONFLICT (key) DO NOTHING
        """)
        # Миграция: добавляем флаг ожидания подтверждения получения заказа
        await conn.execute(
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS "
            "confirm_pending BOOLEAN NOT NULL DEFAULT FALSE"
        )
        # Миграция: переводим денежные колонки в NUMERIC для поддержки копеек
        await conn.execute("""
            DO $$
            BEGIN
                IF EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name='users' AND column_name='balance' AND data_type='integer'
                ) THEN
                    ALTER TABLE users ALTER COLUMN balance TYPE NUMERIC(12,2)
                        USING balance::NUMERIC(12,2);
                END IF;
                IF EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name='transactions' AND column_name='amount' AND data_type='integer'
                ) THEN
                    ALTER TABLE transactions ALTER COLUMN amount TYPE NUMERIC(12,2)
                        USING amount::NUMERIC(12,2);
                END IF;
                IF EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name='orders' AND column_name='price' AND data_type='integer'
                ) THEN
                    ALTER TABLE orders ALTER COLUMN price TYPE NUMERIC(12,2)
                        USING price::NUMERIC(12,2);
                END IF;
            END $$;
        """)


async def db_get_or_create_user(user) -> tuple[dict, bool]:
    """Возвращает запись пользователя, создавая её при первом обращении.
    Возвращает (запись, is_new_user) — is_new_user True, если запись только что создана."""
    pool = await get_pool()
    ref_code = secrets.token_urlsafe(8)
    row = await pool.fetchrow(
        """
        INSERT INTO users (tg_id, username, first_name, balance, referral_code, created_at)
        VALUES ($1, $2, $3, 0, $4, $5)
        ON CONFLICT (tg_id) DO UPDATE
            SET username      = EXCLUDED.username,
                first_name    = EXCLUDED.first_name,
                referral_code = COALESCE(users.referral_code, EXCLUDED.referral_code)
        RETURNING *, (xmax = 0) AS _is_new
        """,
        user.id,
        user.username,
        user.first_name,
        ref_code,
        datetime.utcnow().isoformat(timespec="seconds"),
    )
    data = dict(row)
    is_new = bool(data.pop("_is_new", False))
    return data, is_new


async def db_get_balance(tg_id: int) -> float:
    pool = await get_pool()
    val = await pool.fetchval("SELECT balance FROM users WHERE tg_id = $1", tg_id)
    return float(val) if val is not None else 0.0


async def db_add_balance(tg_id: int, amount: float) -> float:
    pool = await get_pool()
    await pool.execute(
        "UPDATE users SET balance = balance + $1 WHERE tg_id = $2",
        amount, tg_id,
    )
    return await db_get_balance(tg_id)


async def db_try_charge(tg_id: int, amount: float) -> bool:
    """Списывает amount с баланса, если денег достаточно. Возвращает True/False."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            row = await conn.fetchrow(
                "SELECT balance FROM users WHERE tg_id = $1 FOR UPDATE",
                tg_id,
            )
            if not row or float(row["balance"]) < amount:
                return False
            await conn.execute(
                "UPDATE users SET balance = balance - $1 WHERE tg_id = $2",
                amount, tg_id,
            )
            return True


async def db_create_order(
    tg_id: int,
    title: str,
    price: float,
    status: str = "Оплачен",
    category: str | None = None,
) -> int:
    pool = await get_pool()
    order_id = await pool.fetchval(
        "INSERT INTO orders (tg_id, title, price, status, category, created_at)"
        " VALUES ($1, $2, $3, $4, $5, $6) RETURNING id",
        tg_id,
        title,
        price,
        status,
        category,
        datetime.utcnow().isoformat(timespec="seconds"),
    )
    return order_id


async def db_set_order_status(order_id: int, status: str) -> None:
    pool = await get_pool()
    await pool.execute(
        "UPDATE orders SET status = $1 WHERE id = $2",
        status, order_id,
    )


async def db_set_confirm_pending(order_id: int, value: bool) -> None:
    pool = await get_pool()
    await pool.execute(
        "UPDATE orders SET confirm_pending = $1 WHERE id = $2",
        value, order_id,
    )


async def db_get_confirm_pending(order_id: int) -> bool:
    pool = await get_pool()
    val = await pool.fetchval(
        "SELECT confirm_pending FROM orders WHERE id = $1", order_id
    )
    return bool(val) if val is not None else False


# =====================================================================
# Каталог товаров и настройки (DB)
# =====================================================================

async def db_get_products(category: str) -> list[tuple]:
    """Возвращает активные товары категории: [(key, name, price, delivery), ...]."""
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT key, name, price, delivery FROM products "
        "WHERE category = $1 AND active = TRUE ORDER BY sort_order, id",
        category,
    )
    return [(r["key"], r["name"], float(r["price"]), r["delivery"]) for r in rows]


async def db_get_product_by_key(key: str) -> tuple | None:
    """Возвращает товар по ключу: (key, name, price, delivery) или None."""
    pool = await get_pool()
    row = await pool.fetchrow(
        "SELECT key, name, price, delivery FROM products WHERE key = $1", key
    )
    return (row["key"], row["name"], float(row["price"]), row["delivery"]) if row else None


async def db_all_products() -> list[dict]:
    """Все товары (для админ-панели)."""
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT id, category, key, name, price, delivery, sort_order, active "
        "FROM products ORDER BY category, sort_order, id"
    )
    return [dict(r) for r in rows]


async def db_update_product_price(key: str, price: float) -> None:
    pool = await get_pool()
    await pool.execute("UPDATE products SET price = $1 WHERE key = $2", price, key)


async def db_toggle_product(key: str) -> bool:
    """Переключает флаг active товара, возвращает новое значение."""
    pool = await get_pool()
    new_val = await pool.fetchval(
        "UPDATE products SET active = NOT active WHERE key = $1 RETURNING active", key
    )
    return bool(new_val)


async def db_add_product(
    category: str, key: str, name: str, price: float, delivery: str
) -> bool:
    """Добавляет товар. Возвращает False, если ключ уже существует."""
    pool = await get_pool()
    try:
        sort_max = await pool.fetchval(
            "SELECT COALESCE(MAX(sort_order), -1) + 1 FROM products WHERE category = $1",
            category,
        )
        await pool.execute(
            "INSERT INTO products (category, key, name, price, delivery, sort_order) "
            "VALUES ($1, $2, $3, $4, $5, $6)",
            category, key, name, price, delivery, int(sort_max or 0),
        )
        return True
    except Exception:
        return False


async def db_rename_product(key: str, name: str) -> None:
    pool = await get_pool()
    await pool.execute("UPDATE products SET name = $1 WHERE key = $2", name, key)


async def db_delete_product(key: str) -> None:
    pool = await get_pool()
    await pool.execute("DELETE FROM products WHERE key = $1", key)


async def db_rename_category(key: str, name: str) -> None:
    pool = await get_pool()
    await pool.execute("UPDATE categories SET name = $1 WHERE key = $2", name, key)


async def db_set_category_emoji(key: str, emoji: str) -> None:
    pool = await get_pool()
    await pool.execute("UPDATE categories SET emoji = $1 WHERE key = $2", emoji, key)


async def db_delete_category(key: str) -> None:
    """Удаляет категорию и все её товары (в одной транзакции)."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.execute("DELETE FROM products WHERE category = $1", key)
            await conn.execute("DELETE FROM categories WHERE key = $1", key)


async def db_get_setting(key: str, default: str = "") -> str:
    pool = await get_pool()
    val = await pool.fetchval("SELECT value FROM settings WHERE key = $1", key)
    return val if val is not None else default


async def db_set_setting(key: str, value: str) -> None:
    pool = await get_pool()
    await pool.execute(
        "INSERT INTO settings (key, value) VALUES ($1, $2) "
        "ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
        key, value,
    )


async def db_all_settings() -> list[dict]:
    pool = await get_pool()
    rows = await pool.fetch("SELECT key, value FROM settings ORDER BY key")
    return [dict(r) for r in rows]


async def db_all_categories() -> list[dict]:
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT key, name, emoji, sort_order, disabled, disabled_reason, login_hint, needs_code "
        "FROM categories ORDER BY sort_order, key"
    )
    return [dict(r) for r in rows]


async def db_get_category(key: str) -> dict | None:
    pool = await get_pool()
    row = await pool.fetchrow(
        "SELECT key, name, emoji, sort_order, disabled, disabled_reason, login_hint, needs_code "
        "FROM categories WHERE key = $1", key
    )
    return dict(row) if row else None


async def db_create_category(key: str, name: str, emoji: str) -> None:
    pool = await get_pool()
    sort_max = await pool.fetchval("SELECT COALESCE(MAX(sort_order), -1) + 1 FROM categories")
    await pool.execute(
        "INSERT INTO categories (key, name, emoji, sort_order) VALUES ($1, $2, $3, $4) "
        "ON CONFLICT (key) DO NOTHING",
        key, name, emoji, int(sort_max or 0),
    )


async def db_set_category_disabled(key: str, disabled: bool, reason: str | None = None) -> None:
    pool = await get_pool()
    await pool.execute(
        "UPDATE categories SET disabled = $1, disabled_reason = $2 WHERE key = $3",
        disabled, reason, key,
    )


async def db_update_category_login_hint(key: str, text: str | None) -> None:
    pool = await get_pool()
    await pool.execute(
        "UPDATE categories SET login_hint = $1 WHERE key = $2", text, key
    )


async def db_toggle_category_needs_code(key: str) -> bool:
    pool = await get_pool()
    new_val = await pool.fetchval(
        "UPDATE categories SET needs_code = NOT needs_code WHERE key = $1 RETURNING needs_code",
        key,
    )
    return bool(new_val)


async def db_get_order(order_id: int) -> dict | None:
    pool = await get_pool()
    row = await pool.fetchrow("SELECT * FROM orders WHERE id = $1", order_id)
    return dict(row) if row else None


async def db_update_order_status(order_id: int, status: str) -> None:
    pool = await get_pool()
    await pool.execute("UPDATE orders SET status = $1 WHERE id = $2", status, order_id)


async def db_set_order_contact(order_id: int, contact: str) -> None:
    pool = await get_pool()
    await pool.execute(
        "UPDATE orders SET contact = $1 WHERE id = $2",
        contact, order_id,
    )


async def db_set_order_login(order_id: int, login_data: str) -> None:
    pool = await get_pool()
    await pool.execute(
        "UPDATE orders SET login_data = $1 WHERE id = $2",
        login_data, order_id,
    )


async def db_set_order_login_code(order_id: int, code: str) -> None:
    pool = await get_pool()
    await pool.execute(
        "UPDATE orders SET login_code = $1 WHERE id = $2",
        code, order_id,
    )


async def db_add_review(
    order_id: int, tg_id: int, photo_file_id: str | None, comment: str
) -> int:
    pool = await get_pool()
    review_id = await pool.fetchval(
        """
        INSERT INTO reviews (order_id, tg_id, photo_file_id, comment, created_at)
        VALUES ($1, $2, $3, $4, $5)
        ON CONFLICT (order_id) DO UPDATE
            SET photo_file_id = EXCLUDED.photo_file_id,
                comment       = EXCLUDED.comment,
                created_at    = EXCLUDED.created_at
        RETURNING id
        """,
        order_id,
        tg_id,
        photo_file_id,
        comment,
        datetime.utcnow().isoformat(timespec="seconds"),
    )
    return review_id


async def db_has_review(order_id: int) -> bool:
    pool = await get_pool()
    val = await pool.fetchval(
        "SELECT 1 FROM reviews WHERE order_id = $1", order_id
    )
    return val is not None


async def db_add_transaction(
    tg_id: int, amount: float, kind: str, reason: str | None = None
) -> None:
    """Записывает движение по балансу. amount: +начисление / -списание."""
    pool = await get_pool()
    await pool.execute(
        "INSERT INTO transactions (tg_id, amount, kind, reason, created_at)"
        " VALUES ($1, $2, $3, $4, $5)",
        tg_id,
        amount,
        kind,
        reason,
        datetime.utcnow().isoformat(timespec="seconds"),
    )


async def db_get_transactions(tg_id: int, limit: int = 20) -> list[dict]:
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT * FROM transactions WHERE tg_id = $1 ORDER BY id DESC LIMIT $2",
        tg_id, limit,
    )
    return [dict(r) for r in rows]


async def db_set_balance(tg_id: int, value: float) -> float:
    pool = await get_pool()
    await pool.execute(
        "UPDATE users SET balance = $1 WHERE tg_id = $2",
        value, tg_id,
    )
    return await db_get_balance(tg_id)


async def db_set_blacklist(tg_id: int, value: bool) -> None:
    pool = await get_pool()
    await pool.execute(
        "UPDATE users SET is_blacklisted = $1 WHERE tg_id = $2",
        value, tg_id,
    )


async def db_is_blacklisted(tg_id: int) -> bool:
    pool = await get_pool()
    val = await pool.fetchval(
        "SELECT is_blacklisted FROM users WHERE tg_id = $1", tg_id
    )
    return bool(val) if val is not None else False


async def db_find_user(tg_id: int) -> dict | None:
    pool = await get_pool()
    row = await pool.fetchrow("SELECT * FROM users WHERE tg_id = $1", tg_id)
    return dict(row) if row else None


async def db_get_orders(tg_id: int, limit: int = 10) -> list[dict]:
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT * FROM orders WHERE tg_id = $1 ORDER BY id DESC LIMIT $2",
        tg_id, limit,
    )
    return [dict(r) for r in rows]


async def db_orders_count(tg_id: int) -> int:
    pool = await get_pool()
    val = await pool.fetchval(
        "SELECT COUNT(*) FROM orders WHERE tg_id = $1", tg_id
    )
    return int(val) if val else 0


async def db_users_count() -> int:
    pool = await get_pool()
    val = await pool.fetchval("SELECT COUNT(*) FROM users")
    return int(val) if val else 0


async def db_list_users(limit: int, offset: int) -> list[dict]:
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT tg_id, username, first_name, balance, is_blacklisted "
        "FROM users ORDER BY tg_id DESC LIMIT $1 OFFSET $2",
        limit, offset,
    )
    return [dict(r) for r in rows]


async def db_create_promo(
    code: str,
    game: str,
    product_title: str,
    promo_price: int,
    starts_at: str | None = None,
    expires_at: str | None = None,
    discount_pct: int = 0,
    max_uses: int | None = None,
) -> int:
    now = datetime.now(timezone.utc).isoformat()
    pool = await get_pool()
    promo_id = await pool.fetchval(
        "INSERT INTO promo_codes (code, game, product_title, promo_price, discount_pct,"
        " starts_at, expires_at, is_active, created_at, max_uses)"
        " VALUES ($1, $2, $3, $4, $5, $6, $7, TRUE, $8, $9) RETURNING id",
        code.upper(), game, product_title, promo_price, discount_pct,
        starts_at, expires_at, now, max_uses,
    )
    return promo_id


async def db_get_promo_by_code(code: str) -> dict | None:
    pool = await get_pool()
    row = await pool.fetchrow(
        "SELECT * FROM promo_codes WHERE UPPER(code) = UPPER($1)",
        code.strip(),
    )
    return dict(row) if row else None


async def db_get_promo_by_id(promo_id: int) -> dict | None:
    pool = await get_pool()
    row = await pool.fetchrow(
        "SELECT * FROM promo_codes WHERE id = $1", promo_id
    )
    return dict(row) if row else None


async def db_list_promos() -> list[dict]:
    pool = await get_pool()
    rows = await pool.fetch("SELECT * FROM promo_codes ORDER BY id DESC")
    return [dict(r) for r in rows]


async def db_delete_promo(promo_id: int) -> None:
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.execute("DELETE FROM user_promos WHERE promo_id = $1", promo_id)
            await conn.execute("DELETE FROM promo_codes WHERE id = $1", promo_id)


async def db_toggle_promo(promo_id: int, is_active: bool) -> None:
    pool = await get_pool()
    await pool.execute(
        "UPDATE promo_codes SET is_active = $1 WHERE id = $2",
        is_active, promo_id,
    )


async def db_claim_promo(tg_id: int, promo_id: int) -> bool:
    """Добавляет промокод в список пользователя. False — уже есть."""
    now = datetime.now(timezone.utc).isoformat()
    pool = await get_pool()
    try:
        await pool.execute(
            "INSERT INTO user_promos (tg_id, promo_id, claimed_at) VALUES ($1, $2, $3)",
            tg_id, promo_id, now,
        )
        return True
    except Exception:
        return False


async def db_use_promo(tg_id: int, promo_id: int) -> bool:
    """Помечает промокод как использованный.
    Если достигнут лимит max_uses — автоматически деактивирует промокод.
    Возвращает False, если уже использован или лимит исчерпан."""
    now = datetime.now(timezone.utc).isoformat()
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            # Проверяем, не использован ли уже промокод этим пользователем
            row = await conn.fetchrow(
                "SELECT used_at FROM user_promos WHERE tg_id = $1 AND promo_id = $2",
                tg_id, promo_id,
            )
            if not row or row["used_at"] is not None:
                return False
            # Получаем лимит активаций (с блокировкой строки)
            promo_row = await conn.fetchrow(
                "SELECT max_uses FROM promo_codes WHERE id = $1 FOR UPDATE",
                promo_id,
            )
            max_uses = promo_row["max_uses"] if promo_row else None
            # Проверяем лимит до записи
            if max_uses is not None:
                used_count = await conn.fetchval(
                    "SELECT COUNT(*) FROM user_promos "
                    "WHERE promo_id = $1 AND used_at IS NOT NULL",
                    promo_id,
                )
                if int(used_count) >= max_uses:
                    # Лимит исчерпан — деактивируем и отказываем
                    await conn.execute(
                        "UPDATE promo_codes SET is_active = FALSE WHERE id = $1",
                        promo_id,
                    )
                    return False
            # Помечаем как использованный
            await conn.execute(
                "UPDATE user_promos SET used_at = $1 WHERE tg_id = $2 AND promo_id = $3",
                now, tg_id, promo_id,
            )
            # Проверяем, исчерпан ли лимит теперь — деактивируем автоматически
            if max_uses is not None:
                new_count = await conn.fetchval(
                    "SELECT COUNT(*) FROM user_promos "
                    "WHERE promo_id = $1 AND used_at IS NOT NULL",
                    promo_id,
                )
                if int(new_count) >= max_uses:
                    await conn.execute(
                        "UPDATE promo_codes SET is_active = FALSE WHERE id = $1",
                        promo_id,
                    )
            return True


async def db_get_user_promos(tg_id: int) -> list[dict]:
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT up.id, up.tg_id, up.promo_id, up.claimed_at, up.used_at, "
        "pc.code, pc.game, pc.product_title, pc.promo_price, pc.discount_pct, "
        "pc.expires_at, pc.starts_at, pc.is_active "
        "FROM user_promos up "
        "JOIN promo_codes pc ON pc.id = up.promo_id "
        "WHERE up.tg_id = $1 ORDER BY up.id DESC",
        tg_id,
    )
    return [dict(r) for r in rows]


async def db_promo_usage_count(promo_id: int) -> int:
    pool = await get_pool()
    val = await pool.fetchval(
        "SELECT COUNT(*) FROM user_promos WHERE promo_id = $1 AND used_at IS NOT NULL",
        promo_id,
    )
    return int(val) if val else 0


async def db_get_all_user_ids() -> list[int]:
    """Возвращает tg_id всех не заблокированных пользователей (для рассылки)."""
    pool = await get_pool()
    rows = await pool.fetch("SELECT tg_id FROM users WHERE is_blacklisted = FALSE")
    return [r["tg_id"] for r in rows]


# =====================================================================
# Реферальная система — вспомогательные функции
# =====================================================================

def _ref_level_name(level: int) -> str:
    return REFERRAL_LEVELS.get(level, ("Новичок", 0))[0]


def _ref_discount_pct(level: int) -> int:
    return level * REFERRAL_DISCOUNT_PER_LEVEL


def _fmt_price(p: float | int) -> str:
    """Форматирует цену: без копеек если целое, с копейками если дробное."""
    v = float(p)
    if v == int(v):
        return str(int(v))
    return f"{v:.2f}"


async def db_get_user_by_ref_code(ref_code: str) -> dict | None:
    pool = await get_pool()
    row = await pool.fetchrow(
        "SELECT * FROM users WHERE referral_code = $1", ref_code
    )
    return dict(row) if row else None


async def db_set_referred_by(tg_id: int, referrer_id: int) -> None:
    pool = await get_pool()
    await pool.execute(
        "UPDATE users SET referred_by = $1 WHERE tg_id = $2 AND referred_by IS NULL",
        referrer_id, tg_id,
    )


async def db_set_active_ref_promo(tg_id: int, promo_id: int | None) -> None:
    pool = await get_pool()
    await pool.execute(
        "UPDATE users SET active_ref_promo = $1 WHERE tg_id = $2",
        promo_id, tg_id,
    )


async def db_process_referral(buyer_id: int) -> tuple[int | None, bool, int, bool]:
    """
    Вызывается когда заказ покупателя выполнен модератором.
    Если покупатель был приглашён — засчитывает реферал пригласившему.
    Возвращает: (referrer_id | None, level_up: bool, new_level: int, is_first_purchase: bool)
    is_first_purchase: True если первая покупка приглашённого (нужно выдать ему промокоды)
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            buyer = await conn.fetchrow(
                "SELECT referred_by FROM users WHERE tg_id = $1", buyer_id
            )
            if not buyer or not buyer["referred_by"]:
                return None, False, 0, False

            referrer_id = buyer["referred_by"]

            exists = await conn.fetchval(
                "SELECT 1 FROM referral_purchases WHERE buyer_id = $1", buyer_id
            )
            if exists:
                return None, False, 0, False

            await conn.execute(
                "INSERT INTO referral_purchases (buyer_id, referrer_id, created_at) "
                "VALUES ($1, $2, $3) ON CONFLICT DO NOTHING",
                buyer_id, referrer_id,
                datetime.utcnow().isoformat(timespec="seconds"),
            )

            new_row = await conn.fetchrow(
                "UPDATE users SET referral_count = referral_count + 1 "
                "WHERE tg_id = $1 RETURNING referral_level, referral_count",
                referrer_id,
            )
            if not new_row:
                return referrer_id, False, 0, True

            new_count = new_row["referral_count"]
            old_level = new_row["referral_level"]
            new_level = min(5, new_count // REFERRAL_REFS_PER_LEVEL)

            level_up = new_level > old_level
            if level_up:
                await conn.execute(
                    "UPDATE users SET referral_level = $1 WHERE tg_id = $2",
                    new_level, referrer_id,
                )

            return referrer_id, level_up, new_level, True


async def db_admin_set_ref_level(tg_id: int, new_level: int) -> int:
    """Принудительно устанавливает реферальный уровень пользователю.
    Возвращает старый уровень."""
    pool = await get_pool()
    old = await pool.fetchval(
        "SELECT referral_level FROM users WHERE tg_id = $1", tg_id
    )
    await pool.execute(
        "UPDATE users SET referral_level = $1 WHERE tg_id = $2",
        new_level, tg_id,
    )
    return int(old) if old is not None else 0


async def db_admin_get_user_ref_promos(tg_id: int) -> list[dict]:
    """Возвращает все реферальные промокоды пользователя (выданные и использованные)."""
    pool = await get_pool()
    rows = await pool.fetch(
        """
        SELECT up.id AS up_id, up.promo_id, up.used_at,
               pc.code, pc.discount_pct, pc.is_active
        FROM user_promos up
        JOIN promo_codes pc ON pc.id = up.promo_id
        WHERE up.tg_id = $1 AND pc.game = 'ref_discount'
        ORDER BY up.id DESC
        """,
        tg_id,
    )
    return [dict(r) for r in rows]


async def db_admin_remove_user_promo(tg_id: int, promo_id: int) -> bool:
    """Удаляет реферальный промокод из инвентаря пользователя."""
    pool = await get_pool()
    result = await pool.execute(
        "DELETE FROM user_promos WHERE tg_id = $1 AND promo_id = $2",
        tg_id, promo_id,
    )
    deleted = int(result.split()[-1])
    return deleted > 0


async def db_delete_all_user_ref_promos(tg_id: int) -> int:
    """Удаляет все НЕИСПОЛЬЗОВАННЫЕ реферальные промокоды из инвентаря пользователя."""
    pool = await get_pool()
    result = await pool.execute(
        "DELETE FROM user_promos "
        "WHERE tg_id = $1 AND used_at IS NULL AND promo_id IN "
        "(SELECT id FROM promo_codes WHERE game = 'ref_discount')",
        tg_id,
    )
    return int(result.split()[-1])


async def db_delete_all_user_regular_promos(tg_id: int) -> int:
    """Удаляет все обычные (не реферальные) промокоды из инвентаря пользователя."""
    pool = await get_pool()
    result = await pool.execute(
        "DELETE FROM user_promos "
        "WHERE tg_id = $1 AND promo_id IN "
        "(SELECT id FROM promo_codes WHERE game != 'ref_discount')",
        tg_id,
    )
    return int(result.split()[-1])


async def db_delete_all_used_user_promos(tg_id: int) -> int:
    """Удаляет только ИСПОЛЬЗОВАННЫЕ обычные промокоды из инвентаря пользователя."""
    pool = await get_pool()
    result = await pool.execute(
        "DELETE FROM user_promos "
        "WHERE tg_id = $1 AND used_at IS NOT NULL AND promo_id IN "
        "(SELECT id FROM promo_codes WHERE game != 'ref_discount')",
        tg_id,
    )
    return int(result.split()[-1])


async def db_list_regular_promos() -> list[dict]:
    """Возвращает только обычные промокоды (без реферальных ref_discount)."""
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT * FROM promo_codes WHERE game NOT IN ('ref_discount') ORDER BY id DESC"
    )
    return [dict(r) for r in rows]


async def db_delete_all_regular_promos() -> int:
    """Удаляет все обычные промокоды (не реферальные) вместе с записями использования."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            ids = await conn.fetch(
                "SELECT id FROM promo_codes WHERE game NOT IN ('ref_discount')"
            )
            if not ids:
                return 0
            id_list = [r["id"] for r in ids]
            await conn.execute(
                "DELETE FROM user_promos WHERE promo_id = ANY($1::int[])", id_list
            )
            result = await conn.execute(
                "DELETE FROM promo_codes WHERE game NOT IN ('ref_discount')"
            )
            return int(result.split()[-1])


async def db_create_ref_promo_codes(referrer_id: int, level: int) -> list[str]:
    """Создаёт 3 промокода на скидку при повышении уровня."""
    discount = level * REFERRAL_DISCOUNT_PER_LEVEL
    pool = await get_pool()
    codes = []
    now = datetime.now(timezone.utc).isoformat()
    for _ in range(REFERRAL_PROMOS_PER_LEVELUP):
        code = "REF-" + secrets.token_urlsafe(6).upper()[:8]
        try:
            promo_id = await pool.fetchval(
                "INSERT INTO promo_codes (code, game, product_title, promo_price, "
                "discount_pct, is_active, created_at) "
                "VALUES ($1, $2, $3, 0, $4, TRUE, $5) RETURNING id",
                code,
                "ref_discount",
                f"Скидка {discount}% на любой товар (не Telegram Stars)",
                discount,
                now,
            )
            await pool.execute(
                "INSERT INTO user_promos (tg_id, promo_id, claimed_at) "
                "VALUES ($1, $2, $3)",
                referrer_id, promo_id, now,
            )
            codes.append(code)
        except Exception as e:
            logging.warning(f"Ошибка создания реферального промокода: {e}")
    return codes


async def db_create_invite_promo_codes(buyer_id: int, count: int = 3) -> list[str]:
    """Создаёт count промокодов на 5% скидку (по умолчанию 3 — для приглашённого пользователя)."""
    INVITE_DISCOUNT = 5
    pool = await get_pool()
    codes: list[str] = []
    now = datetime.now(timezone.utc).isoformat()
    for _ in range(count):
        code = "INV-" + secrets.token_urlsafe(6).upper()[:8]
        try:
            promo_id = await pool.fetchval(
                "INSERT INTO promo_codes (code, game, product_title, promo_price, "
                "discount_pct, is_active, created_at) "
                "VALUES ($1, $2, $3, 0, $4, TRUE, $5) RETURNING id",
                code,
                "ref_discount",
                "Скидка 5% на любой товар (не Telegram Stars)",
                INVITE_DISCOUNT,
                now,
            )
            await pool.execute(
                "INSERT INTO user_promos (tg_id, promo_id, claimed_at) VALUES ($1, $2, $3)",
                buyer_id, promo_id, now,
            )
            codes.append(code)
        except Exception as e:
            logging.warning(f"Ошибка создания инвайт промокода: {e}")
    return codes


# =====================================================================
# Состояния FSM (для ввода чисел и почты)
# =====================================================================


class ShopStates(StatesGroup):
    waiting_topup_amount = State()
    waiting_topup_confirm = State()
    waiting_robux_amount = State()
    waiting_stars_amount = State()
    waiting_login_data = State()
    waiting_login_code = State()
    waiting_promo_input = State()
    waiting_review_photo = State()
    waiting_review_text = State()
    waiting_reject_screenshot = State()
    waiting_email_code = State()


class AdminStates(StatesGroup):
    waiting_user_id = State()
    waiting_credit_amount = State()
    waiting_credit_reason = State()
    waiting_reset_reason = State()
    waiting_block_reason = State()
    waiting_unblock_reason = State()
    waiting_gp_price = State()
    waiting_mod_reply = State()        # ответ покупателю из чата модератора
    waiting_broadcast = State()        # рассылка всем пользователям
    waiting_edit_product_price = State()   # изменение цены товара
    waiting_edit_setting_value = State()   # изменение настройки (курс/лимит)
    waiting_new_product_key = State()      # ключ нового товара
    waiting_new_product_name = State()     # название нового товара
    waiting_new_product_price = State()    # цена нового товара
    waiting_new_product_delivery = State() # способ выдачи нового товара
    # Управление категориями
    waiting_new_cat_key = State()          # ключ новой категории
    waiting_new_cat_name = State()         # название новой категории
    waiting_new_cat_emoji = State()        # эмодзи новой категории
    waiting_cat_disabled_reason = State()  # причина отключения категории
    waiting_cat_login_hint = State()       # текст после оплаты для категории
    waiting_rename_product = State()       # новое название товара
    waiting_rename_cat_name = State()      # новое название категории
    waiting_rename_cat_emoji = State()     # новый эмодзи категории
    waiting_cat_delete_confirm = State()   # подтверждение удаления категории (текст)


class PromoStates(StatesGroup):
    waiting_code_input = State()
    waiting_discount_type = State()
    waiting_product = State()
    waiting_price = State()
    waiting_discount_pct = State()
    waiting_dates = State()
    waiting_max_uses = State()


# =====================================================================
# Клавиатуры
# =====================================================================


def kb_main_menu() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🛒 Купить донат", callback_data="shop")],
            [
                InlineKeyboardButton(text="👤 Профиль", callback_data="profile"),
                InlineKeyboardButton(text="📦 Мои заказы", callback_data="orders"),
            ],
            [
                InlineKeyboardButton(text="🛡️ Гарантия", callback_data="guarantee"),
                InlineKeyboardButton(text="📖 О магазине", callback_data="info"),
            ],
            [
                InlineKeyboardButton(text="⭐ Отзывы", url=REVIEWS_URL),
                InlineKeyboardButton(text="🆘 Поддержка", callback_data="support"),
            ],
        ]
    )


async def kb_shop_dynamic() -> InlineKeyboardMarkup:
    """Динамическая клавиатура магазина из таблицы categories + хардкод-категории."""
    cats = await db_all_categories()
    rows: list[list[InlineKeyboardButton]] = []
    for cat in cats:
        icon = "⚠️ " if cat["disabled"] else ""
        rows.append([InlineKeyboardButton(
            text=f"{icon}{cat['emoji']} {cat['name']}",
            callback_data=f"cat:{cat['key']}",
        )])
    # Специальные категории с собственными хендлерами (не управляются через DB)
    rows.append([InlineKeyboardButton(text="🎮 Roblox, геймпассом (5 дней)", callback_data="cat:roblox_gamepass")])
    rows.append([InlineKeyboardButton(text="✨ Telegram Stars", callback_data="cat:tgstars")])
    rows.append([InlineKeyboardButton(text="📱 Другие приложения", callback_data="cat:other")])
    rows.append([InlineKeyboardButton(text="⬅️ Назад в меню", callback_data="main")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def kb_back_main(back_cb: str = "shop") -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="⬅️ Назад", callback_data=back_cb),
                InlineKeyboardButton(text="🏠 В меню", callback_data="main"),
            ],
        ]
    )


def kb_product_list(
    items: list[tuple], prefix: str, back_cb: str = "shop"
) -> InlineKeyboardMarkup:
    rows = [
        [
            InlineKeyboardButton(
                text=f"{name} — {price}₽",
                callback_data=f"{prefix}:{key}",
            )
        ]
        for key, name, price, _ in items
    ]
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data=back_cb)])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def kb_product_card(buy_cb: str, back_cb: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Оплатить с баланса", callback_data=buy_cb)],
            [InlineKeyboardButton(text="💳 Пополнить баланс", callback_data="topup")],
            [
                InlineKeyboardButton(text="⬅️ Назад", callback_data=back_cb),
                InlineKeyboardButton(text="🏠 В меню", callback_data="main"),
            ],
        ]
    )


def kb_after_purchase(
    order_id: int,
    needs_login: bool = False,
    needs_code: bool = False,
    login_label: str = "🔐 Отправить данные для входа",
) -> InlineKeyboardMarkup:
    rows = []
    if needs_login:
        rows.append(
            [
                InlineKeyboardButton(
                    text=login_label,
                    callback_data=f"send_login:{order_id}:{1 if needs_code else 0}",
                )
            ]
        )
    if needs_code:
        rows.append(
            [
                InlineKeyboardButton(
                    text="📨 Отправить код для входа",
                    callback_data=f"send_code:{order_id}",
                )
            ]
        )
    rows.extend(
        [
            [
                InlineKeyboardButton(
                    text="💬 Связаться с модератором через бота",
                    callback_data=f"contact_mod:{order_id}",
                )
            ],
            [InlineKeyboardButton(text="📦 Мои заказы", callback_data="orders")],
        ]
    )
    return InlineKeyboardMarkup(inline_keyboard=rows)


def kb_guarantee() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="⭐ Посмотреть отзывы", url=REVIEWS_URL)],
            [InlineKeyboardButton(text="💬 Поддержка", callback_data="support")],
            [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
        ]
    )


def kb_username_link(username: str, back_cb: str = "shop") -> InlineKeyboardMarkup:
    handle = username.lstrip("@")
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="↗️ Перейти по юзернейму",
                    url=f"https://t.me/{handle}",
                )
            ],
            [
                InlineKeyboardButton(text="⬅️ Назад", callback_data=back_cb),
                InlineKeyboardButton(text="🏠 В меню", callback_data="main"),
            ],
        ]
    )


def kb_support() -> InlineKeyboardMarkup:
    handle = SUPPORT_USERNAME.lstrip("@")
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="💬 Перейти в поддержку",
                    url=f"https://t.me/{handle}",
                )
            ],
            [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
        ]
    )


def kb_info() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="💬 Поддержка", callback_data="support")],
            [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
        ]
    )


def kb_profile() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="💳 Пополнить баланс", callback_data="topup")],
            [InlineKeyboardButton(text="📦 История заказов", callback_data="orders")],
            [
                InlineKeyboardButton(
                    text="📜 История транзакций", callback_data="transactions"
                )
            ],
            [
                InlineKeyboardButton(
                    text="🎟️ Мои промокоды", callback_data="my_promos"
                ),
                InlineKeyboardButton(
                    text="✏️ Ввести промокод", callback_data="enter_promo"
                ),
            ],
            [InlineKeyboardButton(text="👥 Реферальная программа", callback_data="referral_info")],
            [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
        ]
    )


def kb_topup_confirm() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Подтвердить", callback_data="topup_go")],
            [InlineKeyboardButton(text="✏️ Изменить сумму", callback_data="topup")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="main")],
        ]
    )


def kb_topup_pay(amount: int, code: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Я оплатил", callback_data=f"paid:{amount}:{code}")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="main")],
        ]
    )


def kb_calc_actions(buy_cb: str, change_cb: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Оплатить с баланса", callback_data=buy_cb)],
            [
                InlineKeyboardButton(
                    text="✏️ Изменить количество", callback_data=change_cb
                )
            ],
            [InlineKeyboardButton(text="💳 Пополнить баланс", callback_data="topup")],
            [
                InlineKeyboardButton(text="⬅️ Назад", callback_data="shop"),
                InlineKeyboardButton(text="🏠 В меню", callback_data="main"),
            ],
        ]
    )


async def kb_order_actions(order: dict) -> InlineKeyboardMarkup:
    order_id = int(order["id"])
    category = order.get("category")
    rows = []

    needs_login = await order_needs_login(category)
    needs_code  = await order_needs_code(category)

    if needs_login:
        login_label = (
            "🔗 Отправить ссылку на геймпасс"
            if category == "roblox_gamepass"
            else "🔐 Отправить данные для входа"
        )
        rows.append([InlineKeyboardButton(
            text=login_label,
            callback_data=f"send_login:{order_id}:{1 if needs_code else 0}",
        )])

    if needs_code:
        rows.append([InlineKeyboardButton(
            text="📨 Отправить код для входа",
            callback_data=f"send_code:{order_id}",
        )])

    rows.extend([
        [InlineKeyboardButton(
            text="💬 Связаться с модератором через бота",
            callback_data=f"contact_mod:{order_id}",
        )],
        [
            InlineKeyboardButton(text="⬅️ Назад", callback_data="orders"),
            InlineKeyboardButton(text="🏠 В меню", callback_data="main"),
        ],
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


# =====================================================================
# Утилиты
# =====================================================================


async def order_needs_login(category: str | None) -> bool:
    """Нужно ли запрашивать данные для входа после покупки."""
    if not category:
        return False
    if category in LOGIN_HINTS:          # специальные хардкод-категории
        return True
    cat = await db_get_category(category)
    return bool(cat and cat.get("login_hint"))


async def order_needs_code(category: str | None) -> bool:
    """Нужно ли показывать кнопку «Отправить код для входа»."""
    if not category:
        return False
    cat = await db_get_category(category)
    if cat is not None:
        return bool(cat.get("needs_code"))
    return False


async def _try_delete(msg: Message) -> None:
    """Тихо удаляет сообщение, игнорируя ошибки."""
    try:
        await msg.delete()
    except Exception:
        pass


async def _edit_prompt(state: FSMContext, text: str,
                       kb: InlineKeyboardMarkup) -> None:
    """Редактирует сохранённое в FSM сообщение-подсказку бота.

    Если редактирование невозможно (фото, устарело и т.д.) — отправляет новое.
    Перед этим обновляет сохранённый msg_id на новый.
    """
    data = await state.get_data()
    chat_id: int | None = data.get("_prompt_chat_id")
    msg_id: int | None = data.get("_prompt_msg_id")
    if chat_id and msg_id:
        try:
            edited = await bot.edit_message_text(
                text,
                chat_id=chat_id,
                message_id=msg_id,
                reply_markup=kb,
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
            await state.update_data(_prompt_msg_id=edited.message_id)
            return
        except Exception:
            pass
    # Если не получилось отредактировать — шлём новым сообщением
    sent = await bot.send_message(
        chat_id or 0,
        text,
        reply_markup=kb,
        parse_mode="HTML",
        disable_web_page_preview=True,
    )
    await state.update_data(_prompt_chat_id=sent.chat.id,
                            _prompt_msg_id=sent.message_id)


async def send_or_edit(target, text: str,
                       kb: InlineKeyboardMarkup) -> Message:
    """Для CallbackQuery удаляет старое сообщение и отправляет новое.
    Возвращает отправленное/отредактированное сообщение."""
    try:
        if isinstance(target, CallbackQuery):
            chat_id = target.message.chat.id
            try:
                await target.message.delete()
            except Exception:
                pass
            return await bot.send_message(
                chat_id,
                text,
                reply_markup=kb,
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
        else:
            return await target.answer(
                text,
                reply_markup=kb,
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
    except Exception:
        chat_id = (
            target.message.chat.id
            if isinstance(target, CallbackQuery)
            else target.chat.id
        )
        return await bot.send_message(
            chat_id,
            text,
            reply_markup=kb,
            parse_mode="HTML",
            disable_web_page_preview=True,
        )


async def show_section(
    call: CallbackQuery, key: str, text: str, kb: InlineKeyboardMarkup
) -> Message:
    """Всегда удаляет старое сообщение и открывает новое. Возвращает Message."""
    image = SECTION_IMAGES.get(key)

    if image and len(text) <= 1024:
        if isinstance(image, str) and not image.startswith(("http://", "https://")):
            # Пробуем найти файл относительно скрипта, затем relative to CWD
            script_dir = os.path.dirname(os.path.abspath(__file__))
            local_path = image if os.path.isabs(image) else os.path.join(script_dir, image)
            if not os.path.exists(local_path):
                local_path = os.path.abspath(image)
            if os.path.exists(local_path):
                try:
                    with open(local_path, "rb") as _f:
                        photo = BufferedInputFile(_f.read(), filename=os.path.basename(local_path))
                except Exception:
                    photo = None
            else:
                photo = None
        else:
            photo = image

        if photo is not None:
            chat_id = call.message.chat.id
            try:
                await call.message.delete()
            except Exception:
                pass

            try:
                sent = await bot.send_photo(
                    chat_id,
                    photo=photo,
                    caption=text,
                    reply_markup=kb,
                    parse_mode="HTML",
                )
                return sent
            except Exception as e:
                logging.warning(f"Не удалось отправить картинку '{key}': {e}")

    return await send_or_edit(call, text, kb)


async def notify_moderator(
    text: str, reply_markup: InlineKeyboardMarkup | None = None
) -> "Message | None":
    """Отправляет уведомление модератору, если задан MODERATOR_CHAT_ID."""
    if not MODERATOR_CHAT_ID:
        return None
    try:
        return await bot.send_message(
            MODERATOR_CHAT_ID, text, parse_mode="HTML", reply_markup=reply_markup
        )
    except Exception as e:
        logging.warning(f"Не удалось уведомить модератора: {e}")
        return None


async def notify_moderator_order(
    text: str, reply_markup: InlineKeyboardMarkup | None = None
) -> None:
    """Отправляет уведомление модератору о заказе и закрепляет сообщение."""
    msg = await notify_moderator(text, reply_markup=reply_markup)
    if msg is None:
        return
    try:
        await bot.pin_chat_message(
            msg.chat.id, msg.message_id, disable_notification=True
        )
    except Exception as e:
        logging.warning(f"Не удалось закрепить сообщение модератора: {e}")


def parse_positive_int(text: str) -> int | None:
    """Парсит положительное целое число из текста."""
    text = (text or "").strip().replace(" ", "")
    if not text.isdigit():
        return None
    value = int(text)
    if value <= 0:
        return None
    return value



# =====================================================================
# /start и главное меню
# =====================================================================


@dp.message(CommandStart())
async def handle_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    user_row, is_new_user = await db_get_or_create_user(message.from_user)

    # Обрабатываем реферальную ссылку
    parts = message.text.split() if message.text else []
    used_ref_link = False
    if len(parts) > 1 and parts[1].startswith("ref_"):
        ref_code = parts[1][4:]
        referrer = await db_get_user_by_ref_code(ref_code)
        if referrer and referrer["tg_id"] != message.from_user.id:
            used_ref_link = True
            already_referred = user_row.get("referred_by") is not None
            await db_set_referred_by(message.from_user.id, referrer["tg_id"])
            # Выдаём 3 INV-промокода только при первом входе по реф. ссылке
            if not already_referred:
                invite_codes = await db_create_invite_promo_codes(message.from_user.id, count=3)
                if invite_codes:
                    codes_text = "\n".join(f"<code>{c}</code>" for c in invite_codes)
                    try:
                        await message.answer(
                            f"🎁 <b>Добро пожаловать! Вам начислены промокоды.</b>\n\n"
                            f"Вы перешли по реферальной ссылке и получили "
                            f"<b>3 промокода</b> на скидку <b>5%</b> на любой заказ:\n\n"
                            f"{codes_text}\n\n"
                            f"Активируйте промокод в разделе "
                            f"<b>«Мои реф. промокоды»</b> перед покупкой.\n"
                            f"<i>Не действует на Telegram Stars.</i>",
                            parse_mode="HTML",
                        )
                    except Exception as e:
                        logging.warning(
                            f"Не удалось уведомить нового реферала "
                            f"{message.from_user.id}: {e}"
                        )

    # Обычным новым пользователям (без реф. ссылки) выдаём 1 промокод на 5%
    if not used_ref_link and is_new_user:
        welcome_codes = await db_create_invite_promo_codes(message.from_user.id, count=1)
        if welcome_codes:
            try:
                await message.answer(
                    f"🎁 <b>Добро пожаловать! Вам начислен промокод.</b>\n\n"
                    f"Вы получили промокод на скидку <b>5%</b> на любой заказ:\n\n"
                    f"<code>{welcome_codes[0]}</code>\n\n"
                    f"Активируйте промокод в разделе "
                    f"<b>«Мои реф. промокоды»</b> перед покупкой.\n"
                    f"<i>Не действует на Telegram Stars.</i>",
                    parse_mode="HTML",
                )
            except Exception as e:
                logging.warning(
                    f"Не удалось уведомить нового пользователя "
                    f"{message.from_user.id}: {e}"
                )

    try:
        await message.answer_photo(
            photo=WELCOME_IMAGE_URL,
            caption=WELCOME_TEXT,
            reply_markup=kb_main_menu(),
            parse_mode="HTML",
        )
    except Exception:
        await message.answer(
            WELCOME_TEXT, reply_markup=kb_main_menu(), parse_mode="HTML"
        )


@dp.callback_query(F.data == "main")
async def cb_main(call: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await show_section(call, "welcome", WELCOME_TEXT, kb_main_menu())
    await call.answer()


@dp.callback_query(F.data == "guarantee")
async def cb_guarantee(call: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await show_section(call, "guarantee", GUARANTEE_TEXT, kb_guarantee())
    await call.answer()


@dp.callback_query(F.data == "support")
async def cb_support(call: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await show_section(call, "support", SUPPORT_TEXT, kb_support())
    await call.answer()


@dp.callback_query(F.data == "info")
async def cb_info(call: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await show_section(call, "info", INFO_TEXT, kb_info())
    await call.answer()


# =====================================================================
# Раздел "Купить донат"
# =====================================================================


@dp.callback_query(F.data == "shop")
async def cb_shop(call: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await show_section(
        call, "shop", "<b>Купить донат</b>\n\nВыберите категорию:",
        await kb_shop_dynamic(),
    )
    await call.answer()


# --- Roblox моментально ---


@dp.callback_query(F.data == "cat:roblox_instant")
async def cb_roblox_instant(call: CallbackQuery) -> None:
    cat = await db_get_category("roblox_instant")
    if cat and cat["disabled"]:
        reason = cat.get("disabled_reason") or "Временно недоступно."
        await show_section(call, "roblox_instant",
            f"⚠️ <b>Категория временно недоступна.</b>\n\n{escape(reason)}",
            kb_back_main("shop"))
        await call.answer()
        return
    products = await db_get_products("roblox_instant")
    label = f"{cat['emoji']} {cat['name']}" if cat else "Roblox — моментально"
    await show_section(
        call,
        "roblox_instant",
        f"<b>{label}</b>\n\nВыберите количество робуксов:",
        kb_product_list(products, "rbi"),
    )
    await call.answer()


@dp.callback_query(F.data.startswith("rbi:"))
async def cb_roblox_instant_card(call: CallbackQuery) -> None:
    key = call.data.split(":", 1)[1]
    product = await db_get_product_by_key(key)
    if not product:
        await call.answer("Товар не найден.", show_alert=True)
        return
    _, name, price, delivery = product
    text = (
        f"<b>{escape(name)}</b>\n\n"
        f"Цена: <b>{price}₽</b>\n"
        f"Способ выдачи: {escape(delivery)}\n"
        "Способ оплаты: Оплата с внутреннего баланса бота\n\n"
        "Для покупки на вашем внутреннем балансе должно быть достаточно средств."
    )
    await show_section(
        call,
        "card_roblox_instant",
        text,
        kb_product_card(f"buy:rbi:{key}", "cat:roblox_instant"),
    )
    await call.answer()


# --- Roblox геймпассом ---


@dp.callback_query(F.data == "cat:roblox_gamepass")
async def cb_roblox_gamepass(call: CallbackQuery, state: FSMContext) -> None:
    text = (
        "<b>Roblox — геймпассом (до 5 дней)</b>\n\n"
        "⚠️ <b>Данный способ покупки временно недоступен.</b>\n\n"
        "На данный момент покупка робуксов через геймпасс невозможна "
        "в связи с ограничениями на платформе Roblox.\n\n"
        "Воспользуйтесь покупкой <b>моментально</b> — она работает в штатном режиме."
    )
    await show_section(call, "roblox_gamepass", text, kb_back_main("shop"))
    await call.answer()


@dp.message(ShopStates.waiting_robux_amount)
async def msg_robux_amount(message: Message, state: FSMContext) -> None:
    qty = parse_positive_int(message.text)
    if qty is None:
        await message.answer(
            "Введите положительное число робуксов (например: 1500).",
            reply_markup=kb_back_main("shop"),
        )
        return

    gp_rate = float(await db_get_setting("robux_gamepass_rate", "0.65"))
    gp_pass_rate = float(await db_get_setting("robux_gamepass_pass_price_rate", "0.7"))
    price = max(1, round(qty * gp_rate))
    gamepass_price = max(1, round(qty / gp_pass_rate))

    await state.update_data(
        robux_qty=qty,
        robux_price=price,
        robux_gamepass_price=gamepass_price,
    )

    text = (
        f"<b>Roblox геймпассом</b>\n\n"
        f"Количество: <b>{qty}</b> робуксов\n"
        f"Курс оплаты: 1 робукс = {gp_rate}₽\n"
        f"Итого к оплате: <b>{price}₽</b>\n"
        f"Цена для создания геймпасса: <b>{gamepass_price} R$</b>\n"
        "Срок: до 5 дней\n"
        "Способ оплаты: Оплата с внутреннего баланса бота"
    )
    await message.answer(
        text,
        reply_markup=kb_calc_actions("buy:robux_gp", "cat:roblox_gamepass"),
        parse_mode="HTML",
    )


# --- Brawl Stars ---


@dp.callback_query(F.data == "cat:brawl")
async def cb_brawl(call: CallbackQuery) -> None:
    cat = await db_get_category("brawl")
    if cat and cat["disabled"]:
        reason = cat.get("disabled_reason") or "Временно недоступно."
        await show_section(call, "brawl",
            f"⚠️ <b>Категория временно недоступна.</b>\n\n{escape(reason)}",
            kb_back_main("shop"))
        await call.answer()
        return
    brawl_products = await db_get_products("brawl")
    label = f"{cat['emoji']} {cat['name']}" if cat else "Brawl Stars"
    rows = [
        [InlineKeyboardButton(text=f"{n} — {p}₽", callback_data=f"bs:{k}")]
        for k, n, p, _ in brawl_products
    ]
    rows.append([InlineKeyboardButton(text="🎁 Другие акции", callback_data="bs:promo")])
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="shop")])
    await show_section(
        call, "brawl",
        f"<b>{label}</b>\n\nВыберите товар:",
        InlineKeyboardMarkup(inline_keyboard=rows),
    )
    await call.answer()


@dp.callback_query(F.data == "bs:promo")
async def cb_brawl_promo(call: CallbackQuery) -> None:
    text = (
        "Чтобы узнать об актуальных акциях в категории Brawl Stars, "
        f"напишите в Telegram: {BRAWL_PROMO_USERNAME}"
    )
    await show_section(
        call, "brawl", text, kb_username_link(BRAWL_PROMO_USERNAME, "cat:brawl")
    )
    await call.answer()


@dp.callback_query(F.data.startswith("bs:"))
async def cb_brawl_card(call: CallbackQuery) -> None:
    key = call.data.split(":", 1)[1]
    if key == "promo":
        return
    product = await db_get_product_by_key(key)
    if not product:
        await call.answer("Товар не найден.", show_alert=True)
        return
    _, name, price, delivery = product
    text = (
        f"<b>{escape(name)}</b>\n\n"
        f"Цена: <b>{price}₽</b>\n"
        f"Способ выдачи: {escape(delivery)}\n"
        "Способ оплаты: Оплата с внутреннего баланса бота"
    )
    await show_section(
        call, "card_brawl", text, kb_product_card(f"buy:bs:{key}", "cat:brawl")
    )
    await call.answer()


# --- Telegram Stars ---


@dp.callback_query(F.data == "cat:tgstars")
async def cb_tgstars(call: CallbackQuery, state: FSMContext) -> None:
    stars_rate = float(await db_get_setting("tg_stars_rate", "1.3"))
    text = (
        "<b>Telegram Stars</b>\n\n"
        f"Курс: 1 звезда = {stars_rate}₽\n\n"
        "Введите количество звёзд одним сообщением (например: 100):"
    )
    sent = await show_section(call, "tgstars", text, kb_back_main("shop"))
    await state.set_state(ShopStates.waiting_stars_amount)
    await state.update_data(_prompt_chat_id=sent.chat.id,
                            _prompt_msg_id=sent.message_id)
    await call.answer()


@dp.message(ShopStates.waiting_stars_amount)
async def msg_stars_amount(message: Message, state: FSMContext) -> None:
    await _try_delete(message)
    qty = parse_positive_int(message.text)
    if qty is None:
        await _edit_prompt(state,
                           "⚠️ Введите положительное число звёзд (например: 100).",
                           kb_back_main("shop"))
        return
    stars_rate = float(await db_get_setting("tg_stars_rate", "1.3"))
    min_stars = int(await db_get_setting("min_tg_stars", "50"))
    if qty < min_stars:
        await _edit_prompt(state,
                           f"⚠️ Минимальное количество — {min_stars} звёзд.\n"
                           "Попробуйте ещё раз:",
                           kb_back_main("shop"))
        return
    price = max(1, round(qty * stars_rate))
    await state.update_data(stars_qty=qty, stars_price=price)
    text = (
        f"<b>Telegram Stars</b>\n\n"
        f"Количество: <b>{qty}</b> ⭐\n"
        f"Курс: 1 звезда = {stars_rate}₽\n"
        f"Итого: <b>{price}₽</b>\n"
        f"Минимум: <b>{min_stars} звёзд</b>\n"
        "Способ выдачи: моментально\n"
        "Способ оплаты: Оплата с внутреннего баланса бота"
    )
    await _edit_prompt(state, text, kb_calc_actions("buy:stars", "cat:tgstars"))


# --- Другие приложения ---


@dp.callback_query(F.data == "cat:other")
async def cb_other(call: CallbackQuery) -> None:
    text = (
        "Для получения информации по донату в других приложениях "
        f"напишите в Telegram: {OTHER_APPS_USERNAME}"
    )
    await show_section(
        call, "other", text, kb_username_link(OTHER_APPS_USERNAME, "shop")
    )
    await call.answer()


# --- Универсальный хендлер для DB-категорий (создаваемых через админку) ---


# Набор ключей с собственными хендлерами — generic-хендлер их игнорирует
_BUILTIN_CAT_KEYS = frozenset({"roblox_instant", "roblox_gamepass", "brawl", "tgstars", "other"})


@dp.callback_query(F.data.startswith("cat:"))
async def cb_cat_generic(call: CallbackQuery) -> None:
    key = call.data.split(":", 1)[1]
    if key in _BUILTIN_CAT_KEYS:
        return  # обрабатывается отдельными хендлерами выше
    await call.answer()
    cat = await db_get_category(key)
    if not cat:
        await call.answer("Категория не найдена.", show_alert=True)
        return
    name = f"{cat['emoji']} {cat['name']}"
    if cat["disabled"]:
        reason = cat.get("disabled_reason") or "Временно недоступно."
        await show_section(call, key,
            f"⚠️ <b>{escape(name)}</b>\n\n<b>Категория временно недоступна.</b>\n\n{escape(reason)}",
            kb_back_main("shop"))
        return
    products = await db_get_products(key)
    if not products:
        await show_section(call, key,
            f"<b>{escape(name)}</b>\n\nТоваров пока нет.", kb_back_main("shop"))
        return
    await show_section(call, key,
        f"<b>{escape(name)}</b>\n\nВыберите товар:",
        kb_product_list(products, f"cp:{key}", "shop"))


@dp.callback_query(F.data.startswith("cp:"))
async def cb_cat_prod_card(call: CallbackQuery) -> None:
    """Карточка товара из DB-категории."""
    await call.answer()
    parts = call.data.split(":", 2)
    if len(parts) < 3:
        return
    cat_key, prod_key = parts[1], parts[2]
    product = await db_get_product_by_key(prod_key)
    if not product:
        await call.answer("Товар не найден.", show_alert=True)
        return
    _, name, price, delivery = product
    cat = await db_get_category(cat_key)
    cat_name = f"{cat['emoji']} {cat['name']}" if cat else cat_key
    text = (
        f"<b>{escape(name)}</b>\n\n"
        f"Цена: <b>{price}₽</b>\n"
        f"Способ выдачи: {escape(delivery)}\n"
        "Способ оплаты: Оплата с внутреннего баланса бота\n\n"
        "Для покупки на вашем внутреннем балансе должно быть достаточно средств."
    )
    await show_section(call, f"card_{cat_key}", text,
        kb_product_card(f"buy_cp:{cat_key}:{prod_key}", f"cat:{cat_key}"))


@dp.callback_query(F.data.startswith("buy_cp:"))
async def cb_buy_cat_prod(call: CallbackQuery, state: FSMContext) -> None:
    """Оплата товара из DB-категории."""
    parts = call.data.split(":", 2)
    if len(parts) < 3:
        return
    cat_key, prod_key = parts[1], parts[2]
    product = await db_get_product_by_key(prod_key)
    if not product:
        await call.answer("Товар не найден.", show_alert=True)
        return
    cat = await db_get_category(cat_key)
    _, name, price, _ = product
    cat_name = cat["name"] if cat else cat_key
    title = f"{cat_name} — {name}"
    await perform_purchase(call, title, float(price), category=cat_key, state=state)


# =====================================================================
# Покупка / списание с баланса
# =====================================================================


async def perform_purchase(
    call: CallbackQuery,
    title: str,
    price: float,
    category: str | None = None,
    extra: dict | None = None,
    state: FSMContext | None = None,
) -> None:
    """Показывает экран подтверждения с итоговой ценой (с учётом реф. скидки).
    Фактическое списание происходит в confirm_purchase."""
    await call.answer()
    user = call.from_user
    user_row, _ = await db_get_or_create_user(user)

    original_price = float(price)

    # Скидка по активному реф./pct промокоду (для приглашённых пользователей)
    # Проверяем тип заранее, чтобы не применять level_disc вместе с pct_discount
    _active_promo_id_check = user_row.get("active_ref_promo") if category != "tgstars" else None
    _active_promo_for_check = (
        await db_get_promo_by_id(_active_promo_id_check)
        if _active_promo_id_check else None
    )
    _has_active_pct = (
        _active_promo_for_check is not None
        and _active_promo_for_check.get("game") == "pct_discount"
        and _active_promo_for_check.get("discount_pct", 0) > 0
    )

    # Автоматическая скидка по реферальному уровню (не для TG Stars, не если активен pct_discount)
    level = (user_row.get("referral_level") or 0) if (category != "tgstars" and not _has_active_pct) else 0
    level_disc = level * REFERRAL_DISCOUNT_PER_LEVEL

    promo_disc = 0
    applied_ref_promo_id = None
    if category != "tgstars":
        active_promo_id = _active_promo_id_check
        if active_promo_id:
            ref_promo = await db_get_promo_by_id(active_promo_id)
            if ref_promo and ref_promo.get("discount_pct", 0) > 0:
                user_promo_list = await db_get_user_promos(user.id)
                up = next(
                    (p for p in user_promo_list
                     if p["promo_id"] == active_promo_id and p.get("used_at") is None),
                    None,
                )
                if up:
                    promo_disc = ref_promo["discount_pct"]
                    applied_ref_promo_id = active_promo_id

    final_price = original_price * (1 - level_disc / 100) * (1 - promo_disc / 100)

    if state is not None:
        await state.update_data(
            _pnd_title=title,
            _pnd_final=final_price,
            _pnd_orig=original_price,
            _pnd_cat=category,
            _pnd_extra=extra,
            _pnd_ldsc=level_disc,
            _pnd_pdsc=promo_disc,
            _pnd_pid=applied_ref_promo_id,
        )

    balance = await db_get_balance(user.id)

    if level_disc or promo_disc:
        disc_parts = []
        if level_disc:
            disc_parts.append(f"🏆 Реф. уровень: <b>-{level_disc}%</b>")
        if promo_disc:
            disc_parts.append(f"🎟️ Промокод: <b>-{promo_disc}%</b>")
        price_block = (
            f"💵 Цена без скидки: <s>{_fmt_price(original_price)}₽</s>\n"
            + "\n".join(disc_parts) + "\n"
            + f"💰 Итого к оплате: <b>{_fmt_price(final_price)}₽</b>"
        )
    else:
        price_block = f"💰 К оплате: <b>{_fmt_price(final_price)}₽</b>"

    enough = balance >= final_price
    text = (
        "🛒 <b>Подтверждение оплаты</b>\n\n"
        f"📦 {escape(title)}\n\n"
        f"{price_block}\n"
        f"💼 Ваш баланс: <b>{_fmt_price(balance)}₽</b>\n\n"
        + ("✅ Баланса достаточно для оплаты." if enough
           else "❌ Недостаточно средств на балансе!")
    )
    if enough:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Подтвердить оплату",
                                  callback_data="confirm_purchase")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="main")],
        ])
    else:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="💳 Пополнить баланс", callback_data="topup")],
            [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
        ])
    await send_or_edit(call, text, kb)


@dp.callback_query(F.data == "confirm_purchase")
async def cb_confirm_purchase(call: CallbackQuery, state: FSMContext) -> None:
    """Выполняет покупку, данные которой сохранены в FSM perform_purchase."""
    await call.answer()
    data = await state.get_data()

    title: str | None = data.get("_pnd_title")
    final_price: float | None = data.get("_pnd_final")
    original_price: float = data.get("_pnd_orig") or 0.0
    category: str | None = data.get("_pnd_cat")
    extra: dict | None = data.get("_pnd_extra")
    level_disc: int = data.get("_pnd_ldsc") or 0
    promo_disc: int = data.get("_pnd_pdsc") or 0
    applied_ref_promo_id = data.get("_pnd_pid")

    if not title or final_price is None:
        await call.answer("Данные заказа устарели. Начните заново.", show_alert=True)
        return

    user = call.from_user

    if user.id in _processing_payments:
        await call.answer("Платёж уже обрабатывается, подождите...", show_alert=True)
        return
    _processing_payments.add(user.id)
    try:
        ok = await db_try_charge(user.id, final_price)
        if not ok:
            balance = await db_get_balance(user.id)
            text = (
                "❌ <b>Недостаточно средств на балансе.</b>\n\n"
                f"Сумма заказа: {_fmt_price(final_price)}₽\n"
                f"Ваш баланс: {_fmt_price(balance)}₽\n\n"
                "Пополните баланс и повторите попытку."
            )
            kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="💳 Пополнить баланс", callback_data="topup")],
                [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
            ])
            await send_or_edit(call, text, kb)
            return

        order_id = await db_create_order(user.id, title, final_price,
                                         status="Оплачен", category=category)
        await db_add_transaction(user.id, -final_price, kind="purchase",
                                 reason=f"Заказ #{order_id}: {title}")

        if applied_ref_promo_id:
            await db_use_promo(user.id, applied_ref_promo_id)
            await db_set_active_ref_promo(user.id, None)

        # Сбрасываем pending-данные, оставляем прочее состояние FSM
        await state.update_data(
            _pnd_title=None, _pnd_final=None, _pnd_orig=None,
            _pnd_cat=None, _pnd_extra=None, _pnd_ldsc=None,
            _pnd_pdsc=None, _pnd_pid=None,
        )
    finally:
        _processing_payments.discard(user.id)

    new_balance = await db_get_balance(user.id)
    # Текст после оплаты: для специальных категорий — хардкод, для DB-категорий — из таблицы
    if category in LOGIN_HINTS:
        login_hint = LOGIN_HINTS.get(category)
        needs_code = False
    elif category:
        _cat_obj = await db_get_category(category)
        login_hint = _cat_obj.get("login_hint") if _cat_obj else None
        needs_code = bool(_cat_obj.get("needs_code")) if _cat_obj else False
    else:
        login_hint = None
        needs_code = False
    login_label = "🔐 Отправить данные для входа"

    disc_line = ""
    if level_disc:
        disc_line += f"🏆 Скидка реф. уровень: <b>-{level_disc}%</b>\n"
    if promo_disc:
        disc_line += f"🎟️ Промокод: <b>-{promo_disc}%</b>\n"
    if disc_line and original_price != final_price:
        disc_line += f"💵 Цена без скидки: <s>{_fmt_price(original_price)}₽</s>\n"

    text = (
        "✅ <b>Оплата прошла успешно. Заказ принят в обработку.</b>\n\n"
        f"🧾 Номер заказа: <code>#{order_id}</code>\n"
        f"🎁 Товар: {escape(title)}\n"
        f"{disc_line}"
        f"💰 Сумма: <b>{_fmt_price(final_price)}₽</b>\n"
        f"💼 Остаток на балансе: <b>{_fmt_price(new_balance)}₽</b>\n\n"
    )

    if category == "roblox_gamepass":
        login_label = "🔗 Отправить ссылку на геймпасс"
        if extra:
            gp_price = extra.get("gamepass_price")
            if gp_price:
                text += f"🎮 <b>Цена для создания геймпасса:</b> <b>{gp_price} R$</b>\n\n"
        text += (
            "🔗 Создайте геймпасс на указанную сумму и нажмите "
            "«Отправить ссылку на геймпасс» — пришлите ссылку одним "
            "сообщением. Никакие данные от аккаунта и коды отправлять не нужно."
        )
    elif login_hint:
        text += (
            "🔐 <b>Для выполнения заказа нужны данные для входа.</b>\n"
            f"{escape(login_hint)}\n\n"
            "Нажмите «Отправить данные для входа» и пришлите их одним "
            "сообщением — модератор получит их вместе с заказом."
        )
        if needs_code:
            if category and category.startswith("roblox"):
                text += (
                    "\n\n📨 <b>Если у вас подключён двухфакторный вход</b> "
                    "(2FA) — на почту придёт код, когда модератор будет "
                    "заходить на аккаунт. Пришлите код кнопкой «Отправить "
                    "код для входа»."
                )
            else:
                text += (
                    "\n\n📨 Если на почту придёт код подтверждения, когда "
                    "модератор будет заходить — пришлите его кнопкой "
                    "«Отправить код для входа»."
                )
    else:
        text += "Выберите удобный способ связи с модератором:"

    await send_or_edit(
        call,
        text,
        kb_after_purchase(
            order_id,
            needs_login=bool(login_hint),
            needs_code=needs_code,
            login_label=login_label,
        ),
    )
    await call.answer("Заказ создан")

    username = f"@{user.username}" if user.username else "—"
    admin_rows: list[list[InlineKeyboardButton]] = [
        [InlineKeyboardButton(text="✅ Завершить заказ",
                              callback_data=f"ordone:{order_id}:{user.id}")],
        [InlineKeyboardButton(text="📧 Запросить код из почты",
                              callback_data=f"mod:req_email:{order_id}:{user.id}")],
        [InlineKeyboardButton(text="💸 Возврат",
                              callback_data=f"mod:refund:{order_id}")],
    ]
    if category == "roblox_gamepass" and extra:
        gp_price = int(extra.get("gamepass_price", 0))
        if gp_price > 0:
            admin_rows.insert(0, [InlineKeyboardButton(
                text="🎮 Попросить изменить цену геймпасса",
                callback_data=f"gpfix:{order_id}:{user.id}:{gp_price}",
            )])

    admin_kb = InlineKeyboardMarkup(inline_keyboard=admin_rows)
    disc_note = f" (скидка {level_disc + promo_disc}%)" if level_disc or promo_disc else ""
    admin_text = (
        f"🆕 <b>Новый заказ #{order_id}</b>\n\n"
        f"Покупатель: {escape(user.first_name or '')} ({escape(username)})\n"
        f"Telegram ID: <code>{user.id}</code>\n"
        f"Товар: {escape(title)}\n"
        f"Сумма: {_fmt_price(final_price)}₽{disc_note}\n"
        f"Статус: Оплачен"
    )
    if category == "roblox_gamepass" and extra:
        gp_price = extra.get("gamepass_price")
        if gp_price:
            admin_text += f"\nЦена геймпасса: <b>{gp_price} R$</b> (допуск ±5 R$)"
    await notify_moderator_order(admin_text, reply_markup=admin_kb)


# --- Запрос данных для входа ---


@dp.callback_query(F.data.startswith("send_login:"))
async def cb_send_login(call: CallbackQuery, state: FSMContext) -> None:
    parts = call.data.split(":")
    order_id = int(parts[1])
    needs_code = len(parts) > 2 and parts[2] == "1"
    await state.set_state(ShopStates.waiting_login_data)
    text = (
        f"🔐 Отправьте данные для входа по заказу <b>#{order_id}</b> "
        "одним сообщением.\n\n"
        "Например: логин/пароль, ссылка на геймпасс, Supercell ID, "
        "@username и т.д."
    )
    sent = await send_or_edit(call, text, kb_back_main(f"order_actions:{order_id}"))
    await state.update_data(login_order_id=order_id, login_needs_code=needs_code,
                            _prompt_chat_id=sent.chat.id, _prompt_msg_id=sent.message_id)
    await call.answer()


@dp.message(ShopStates.waiting_login_data)
async def msg_login_data(message: Message, state: FSMContext) -> None:
    await _try_delete(message)
    payload = (message.text or "").strip()
    data = await state.get_data()
    order_id = int(data.get("login_order_id", 0))
    needs_code = bool(data.get("login_needs_code", False))
    if not payload:
        await _edit_prompt(state, "⚠️ Пустое сообщение. Пришлите данные текстом.",
                           kb_back_main(f"order_actions:{order_id}"))
        return
    await state.clear()

    if order_id:
        await db_set_order_login(order_id, payload)

    user = message.from_user
    username = f"@{user.username}" if user.username else "—"
    await notify_moderator(
        f"🔐 Данные для входа по заказу <b>#{order_id}</b>\n\n"
        f"<pre>{escape(payload)}</pre>\n"
        f"Покупатель: {escape(user.first_name or '')} ({escape(username)})\n"
        f"Telegram ID: <code>{user.id}</code>"
    )

    text = (
        f"✅ Спасибо! Данные по заказу <b>#{order_id}</b> переданы модератору.\n"
        "Скоро с вами свяжутся."
    )
    if needs_code:
        text += (
            "\n\n📨 Когда модератор будет заходить на аккаунт — вам "
            "может прийти код подтверждения. Пришлите его кнопкой "
            "<b>«Отправить код для входа»</b> ниже."
        )
    await message.answer(
        text,
        reply_markup=kb_after_purchase(
            order_id, needs_login=False, needs_code=needs_code
        ),
        parse_mode="HTML",
    )


@dp.callback_query(F.data.startswith("send_code:"))
async def cb_send_code(call: CallbackQuery, state: FSMContext) -> None:
    order_id = int(call.data.split(":", 1)[1])
    await state.set_state(ShopStates.waiting_login_code)
    text = (
        f"📨 Пришлите код для входа по заказу <b>#{order_id}</b> "
        "одним сообщением.\n\n"
        "Это код, который пришёл вам на почту, когда модератор "
        "заходил на аккаунт."
    )
    sent = await send_or_edit(call, text, kb_back_main(f"order_actions:{order_id}"))
    await state.update_data(code_order_id=order_id,
                            _prompt_chat_id=sent.chat.id, _prompt_msg_id=sent.message_id)
    await call.answer()


@dp.message(ShopStates.waiting_login_code)
async def msg_login_code(message: Message, state: FSMContext) -> None:
    await _try_delete(message)
    code = (message.text or "").strip()
    data = await state.get_data()
    order_id = int(data.get("code_order_id", 0))
    if not code:
        await _edit_prompt(state, "⚠️ Пустое сообщение. Пришлите код текстом.",
                           kb_back_main(f"order_actions:{order_id}"))
        return
    await state.clear()

    if order_id:
        await db_set_order_login_code(order_id, code)

    user = message.from_user
    username = f"@{user.username}" if user.username else "—"
    await notify_moderator(
        f"📨 Код для входа по заказу <b>#{order_id}</b>\n\n"
        f"<pre>{escape(code)}</pre>\n"
        f"Покупатель: {escape(user.first_name or '')} ({escape(username)})\n"
        f"Telegram ID: <code>{user.id}</code>"
    )

    await message.answer(
        f"✅ Спасибо! Код по заказу <b>#{order_id}</b> передан модератору.\n\n"
        "Если придёт ещё один код — пришлите его той же кнопкой "
        "<b>«Отправить код для входа»</b>.",
        reply_markup=kb_after_purchase(order_id, needs_login=False, needs_code=True),
        parse_mode="HTML",
    )


@dp.callback_query(F.data.startswith("buy:rbi:"))
async def cb_buy_rbi(call: CallbackQuery, state: FSMContext) -> None:
    key = call.data.split(":", 2)[2]
    product = await db_get_product_by_key(key)
    if not product:
        await call.answer("Товар не найден.", show_alert=True)
        return
    _, name, price, _ = product
    await perform_purchase(
        call, f"Roblox — {name} (моментально)", price,
        category="roblox_instant", state=state,
    )


@dp.callback_query(F.data.startswith("buy:bs:"))
async def cb_buy_bs(call: CallbackQuery, state: FSMContext) -> None:
    key = call.data.split(":", 2)[2]
    product = await db_get_product_by_key(key)
    if not product:
        await call.answer("Товар не найден.", show_alert=True)
        return
    _, name, price, _ = product
    await perform_purchase(
        call, f"Brawl Stars — {name}", price, category="brawl", state=state,
    )


@dp.callback_query(F.data == "buy:robux_gp")
async def cb_buy_robux_gp(call: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    qty = data.get("robux_qty")
    price = data.get("robux_price")
    gamepass_price = data.get("robux_gamepass_price")

    if not qty or not price or not gamepass_price:
        await call.answer("Сначала введите количество.", show_alert=True)
        return

    await perform_purchase(
        call,
        f"Roblox геймпассом — {qty} робуксов (до 5 дней)",
        int(price),
        category="roblox_gamepass",
        extra={"qty": int(qty), "gamepass_price": int(gamepass_price)},
        state=state,
    )


@dp.callback_query(F.data == "buy:stars")
async def cb_buy_stars(call: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    qty = data.get("stars_qty")
    price = data.get("stars_price")
    if not qty or not price:
        await call.answer("Сначала введите количество.", show_alert=True)
        return
    await perform_purchase(
        call,
        f"Telegram Stars — {qty} звёзд",
        int(price),
        category="tgstars",
        state=state,
    )


# =====================================================================
# Связь с модератором после покупки
# =====================================================================


@dp.callback_query(F.data.startswith("contact_mod:"))
async def cb_contact_mod(call: CallbackQuery) -> None:
    order_id = call.data.split(":", 1)[1]
    user = call.from_user
    username = f"@{user.username}" if user.username else "—"

    await notify_moderator(
        f"📨 Покупатель просит связь по заказу <b>#{order_id}</b>\n\n"
        f"Имя: {escape(user.first_name or '')}\n"
        f"Username: {escape(username)}\n"
        f"Telegram ID: <code>{user.id}</code>\n"
        "Свяжитесь с клиентом в Telegram."
    )

    handle = SUPPORT_USERNAME.lstrip("@")
    text = (
        f"✅ Модератор уведомлён о вашем заказе <b>#{order_id}</b> "
        "и скоро свяжется с вами.\n\n"
        f"Если хотите написать сами — нажмите кнопку ниже."
    )
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="💬 Написать модератору", url=f"https://t.me/{handle}"
                )
            ],
            [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
        ]
    )
    await send_or_edit(call, text, kb)
    await call.answer()


# =====================================================================
# Профиль и история заказов
# =====================================================================


@dp.callback_query(F.data == "profile")
async def cb_profile(call: CallbackQuery) -> None:
    user_row, _ = await db_get_or_create_user(call.from_user)
    orders_cnt = await db_orders_count(call.from_user.id)
    user = call.from_user
    username = f"@{user.username}" if user.username else "не указан"
    created = user_row["created_at"].split("T")[0]
    level = user_row.get("referral_level", 0)
    level_name = _ref_level_name(level)
    discount = _ref_discount_pct(level)
    discount_line = f" · скидка <b>{discount}%</b>" if discount else ""

    text = (
        "<b>👤 Ваш профиль</b>\n\n"
        f"Telegram ID: <code>{user.id}</code>\n"
        f"Username: {escape(username)}\n"
        f"Имя: {escape(user.first_name or '—')}\n"
        f"Баланс: <b>{user_row['balance']}₽</b>\n"
        f"Дата регистрации: {created}\n"
        f"Заказов: <b>{orders_cnt}</b>\n"
        f"Статус: <b>{level_name}</b>{discount_line}"
    )
    await show_section(call, "profile", text, kb_profile())
    await call.answer()


@dp.callback_query(F.data == "orders")
async def cb_orders(call: CallbackQuery) -> None:
    orders = await db_get_orders(call.from_user.id, limit=10)

    rows = []

    if not orders:
        text = "📦 <b>История заказов</b>\n\nУ вас пока нет заказов."
    else:
        lines = ["📦 <b>История заказов</b>\n"]
        for o in orders:
            date = _fmt_msk(o["created_at"])
            lines.append(
                f"<b>#{o['id']}</b>  •  {escape(o['title'])}\n"
                f"Сумма: {o['price']}₽  •  Статус: {escape(o['status'])}\n"
                f"<i>{date}</i>\n"
            )

            if o["status"] != "Выполнен":
                rows.append(
                    [
                        InlineKeyboardButton(
                            text=f"⚙️ Действия по заказу #{o['id']}",
                            callback_data=f"order_actions:{o['id']}",
                        )
                    ]
                )

        text = "\n".join(lines)

    rows.append(
        [
            InlineKeyboardButton(text="⬅️ Назад", callback_data="profile"),
            InlineKeyboardButton(text="🏠 В меню", callback_data="main"),
        ]
    )

    kb = InlineKeyboardMarkup(inline_keyboard=rows)
    await show_section(call, "orders", text, kb)
    await call.answer()


@dp.callback_query(F.data.startswith("order_actions:"))
async def cb_order_actions(call: CallbackQuery) -> None:
    try:
        order_id = int(call.data.split(":", 1)[1])
    except ValueError:
        await call.answer("Некорректный заказ.", show_alert=True)
        return

    order = await db_get_order(order_id)
    if not order or order["tg_id"] != call.from_user.id:
        await call.answer("Заказ не найден.", show_alert=True)
        return

    if order["status"] == "Выполнен":
        has_review = await db_has_review(order_id)
        review_line = (
            "\n\n⭐ Вы уже оставили отзыв по этому заказу. Спасибо!"
            if has_review
            else "\n\n⭐ Если вам понравилось — оставьте, пожалуйста, отзыв."
        )
        text = (
            f"📦 <b>Заказ #{order_id}</b>\n\n"
            f"Товар: {escape(order['title'])}\n"
            f"Сумма: {order['price']}₽\n"
            f"Статус: <b>{escape(order['status'])}</b>\n\n"
            "Этот заказ уже выполнен."
            f"{review_line}"
        )
        rows = []
        if not has_review:
            rows.append(
                [
                    InlineKeyboardButton(
                        text="⭐ Оставить отзыв",
                        callback_data=f"review:{order_id}",
                    )
                ]
            )
        rows.append(
            [
                InlineKeyboardButton(text="⬅️ Назад", callback_data="orders"),
                InlineKeyboardButton(text="🏠 В меню", callback_data="main"),
            ]
        )
        kb = InlineKeyboardMarkup(inline_keyboard=rows)
        await send_or_edit(call, text, kb)
        await call.answer()
        return

    text = (
        f"📦 <b>Заказ #{order_id}</b>\n\n"
        f"Товар: {escape(order['title'])}\n"
        f"Сумма: {order['price']}₽\n"
        f"Статус: <b>{escape(order['status'])}</b>\n\n"
        "Выберите действие по заказу:"
    )
    await send_or_edit(call, text, await kb_order_actions(order))
    await call.answer()


def _format_tx_kind(kind: str) -> str:
    return {
        "topup": "Пополнение",
        "purchase": "Покупка",
        "admin_add": "Начисление администратором",
        "admin_reset": "Обнуление баланса администратором",
    }.get(kind, kind)


@dp.callback_query(F.data == "transactions")
async def cb_transactions(call: CallbackQuery) -> None:
    txs = await db_get_transactions(call.from_user.id, limit=20)
    if not txs:
        text = (
            "📜 <b>История транзакций</b>\n\n"
            "Здесь будут отображаться все начисления и списания по "
            "вашему балансу. Пока операций нет."
        )
    else:
        lines = ["📜 <b>История транзакций</b>\n"]
        for t in txs:
            date = _fmt_msk(t["created_at"])
            sign = "➕" if t["amount"] > 0 else "➖"
            lines.append(
                f"{sign} <b>{abs(t['amount'])}₽</b> — "
                f"{escape(_format_tx_kind(t['kind']))}\n"
                + (f"<i>{escape(t['reason'])}</i>\n" if t["reason"] else "")
                + f"<i>{date}</i>\n"
            )
        text = "\n".join(lines)
        if len(text) > 3800:
            text = text[:3800] + "\n…"
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="⬅️ Назад", callback_data="profile"),
                InlineKeyboardButton(text="🏠 В меню", callback_data="main"),
            ],
        ]
    )
    await show_section(call, "orders", text, kb)
    await call.answer()


# =====================================================================
# Пополнение баланса
# =====================================================================


@dp.callback_query(F.data == "topup")
async def cb_topup(call: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(ShopStates.waiting_topup_amount)
    min_topup = int(await db_get_setting("min_topup", "10"))
    text = (
        "💳 <b>Пополнение баланса</b>\n\n"
        f"Введите сумму пополнения в рублях (минимум {min_topup}₽).\n"
        "Например: 500"
    )
    sent = await show_section(call, "topup", text, kb_back_main("profile"))
    await state.update_data(_prompt_chat_id=sent.chat.id,
                            _prompt_msg_id=sent.message_id)
    await call.answer()


@dp.message(ShopStates.waiting_topup_amount)
async def msg_topup_amount(message: Message, state: FSMContext) -> None:
    await _try_delete(message)
    amount = parse_positive_int(message.text)
    if amount is None:
        await _edit_prompt(state,
                           "⚠️ Введите положительное число рублей, например: 500.",
                           kb_back_main("profile"))
        return
    min_topup = int(await db_get_setting("min_topup", "10"))
    if amount < min_topup:
        await _edit_prompt(state,
                           f"⚠️ Минимальная сумма пополнения — {min_topup}₽\n"
                           "Попробуйте ещё раз:",
                           kb_back_main("profile"))
        return
    await state.update_data(topup_amount=amount)
    await state.set_state(ShopStates.waiting_topup_confirm)
    await _edit_prompt(state,
                       f"Вы хотите пополнить баланс на <b>{amount}₽</b>?",
                       kb_topup_confirm())


@dp.callback_query(F.data == "topup_go")
async def cb_topup_go(call: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    amount = int(data.get("topup_amount", 0))
    await state.clear()
    if amount < int(await db_get_setting("min_topup", "10")):
        await call.answer("Сумма не задана.", show_alert=True)
        return
    code = f"{secrets.randbelow(1_000_000):06d}"
    text = (
        f"💳 <b>Оплатите {amount}₽</b> по реквизитам ниже:\n\n"
        f"{CARD_DETAILS}\n\n"
        f"💰 <b>Сумма к переводу: {amount}₽</b>\n\n"
        f"🔢 <b>Код для комментария к переводу:</b>\n"
        f"<code>{code}</code>\n\n"
        "⚠️ Обязательно укажите этот код в <b>комментарии</b> к переводу — "
        "без него платёж не будет засчитан.\n\n"
        "⚠️ Переведите <b>точную сумму</b>, иначе платёж не будет засчитан.\n\n"
        "После оплаты нажмите «Я оплатил» — заявка отправится модератору "
        "на проверку, и баланс начислится после подтверждения."
    )
    await show_section(call, "topup", text, kb_topup_pay(amount, code))
    await call.answer()


@dp.callback_query(F.data.startswith("paid:"))
async def cb_paid(call: CallbackQuery) -> None:
    await call.answer("Заявка отправлена")
    parts = call.data.split(":")
    try:
        amount = int(parts[1])
    except (ValueError, IndexError):
        return
    code = parts[2] if len(parts) > 2 else ""

    user = call.from_user
    username = f"@{user.username}" if user.username else "—"

    text = (
        "🕒 <b>Заявка на пополнение отправлена</b>\n\n"
        f"Сумма: <b>{amount}₽</b>\n"
        + (f"Код в комментарии: <code>{code}</code>\n\n" if code else "\n")
        + "Мы проверим поступление средств и начислим баланс. "
        "Вы получите уведомление, как только заявка будет обработана."
    )
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
        ]
    )
    await send_or_edit(call, text, kb)

    admin_kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="✅ Подтвердить",
                    callback_data=f"tpconf:{user.id}:{amount}:{code}",
                ),
                InlineKeyboardButton(
                    text="❌ Отклонить",
                    callback_data=f"tprej:{user.id}:{amount}:{code}",
                ),
            ],
        ]
    )
    await notify_moderator(
        f"💰 <b>Новая заявка на пополнение</b>\n\n"
        f"Покупатель: {escape(user.first_name or '')} ({escape(username)})\n"
        f"Telegram ID: <code>{user.id}</code>\n"
        f"Сумма: <b>{amount}₽</b>\n"
        + (f"Код в комментарии: <code>{code}</code>\n\n" if code else "\n")
        + "Проверьте поступление средств (сумма + код в комментарии) "
        "и нажмите соответствующую кнопку.",
        reply_markup=admin_kb,
    )


def _is_moderator(user_id: int) -> bool:
    return bool(MODERATOR_CHAT_ID) and user_id == MODERATOR_CHAT_ID


@dp.callback_query(F.data.startswith("tpconf:"))
async def cb_topup_admin_confirm(call: CallbackQuery) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer("Только для модератора.", show_alert=True)
        return
    try:
        parts = call.data.split(":")
        target_id = int(parts[1])
        amount = int(parts[2])
        code = parts[3] if len(parts) > 3 else ""
    except (ValueError, IndexError):
        await call.answer("Некорректные данные.", show_alert=True)
        return

    new_balance = await db_add_balance(target_id, amount)
    reason = f"Пополнение баланса (код: {code})" if code else "Пополнение баланса"
    await db_add_transaction(
        target_id,
        amount,
        kind="topup",
        reason=reason,
    )

    try:
        old = call.message.text or call.message.caption or ""
        await call.message.edit_text(
            f"{old}\n\n✅ <b>Подтверждено.</b> Баланс пользователя: {_fmt_price(new_balance)}₽",
            parse_mode="HTML",
        )
    except Exception:
        pass

    try:
        user_kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="🛒 Купить донат", callback_data="shop")],
                [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
            ]
        )
        await bot.send_message(
            target_id,
            f"✅ <b>Баланс пополнен на {_fmt_price(amount)}₽</b>\n\n"
            f"Текущий баланс: <b>{_fmt_price(new_balance)}₽</b>",
            parse_mode="HTML",
            reply_markup=user_kb,
        )
    except Exception as e:
        logging.warning(f"Не удалось уведомить пользователя {target_id}: {e}")

    await call.answer("Баланс начислен")


@dp.callback_query(F.data.startswith("tprej:"))
async def cb_topup_admin_reject(call: CallbackQuery) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer("Только для модератора.", show_alert=True)
        return
    try:
        parts = call.data.split(":")
        target_id = int(parts[1])
        amount = int(parts[2])
    except (ValueError, IndexError):
        await call.answer("Некорректные данные.", show_alert=True)
        return

    try:
        old = call.message.text or call.message.caption or ""
        await call.message.edit_text(
            f"{old}\n\n❌ <b>Отклонено.</b>",
            parse_mode="HTML",
        )
    except Exception:
        pass

    try:
        await bot.send_message(
            target_id,
            f"❌ <b>Заявка на пополнение на {amount}₽ отклонена.</b>\n\n"
            "Если вы уверены, что оплата была произведена — свяжитесь с поддержкой.",
            parse_mode="HTML",
            reply_markup=kb_support(),
        )
    except Exception as e:
        logging.warning(f"Не удалось уведомить пользователя {target_id}: {e}")

    await call.answer("Отклонено")


# =====================================================================
# Авто-подтверждение заказа через 30 минут
# =====================================================================

RECEIPT_CONFIRM_TIMEOUT = 30 * 60  # секунд


async def _auto_confirm_receipt(order_id: int, tg_id: int) -> None:
    """Через 30 минут автоматически подтверждает заказ, если покупатель не ответил."""
    await asyncio.sleep(RECEIPT_CONFIRM_TIMEOUT)
    still_pending = await db_get_confirm_pending(order_id)
    if not still_pending:
        return  # пользователь уже нажал кнопку

    await db_set_confirm_pending(order_id, False)

    # Уведомляем модератора
    try:
        await notify_moderator(
            f"⏱ <b>Заказ #{order_id} подтверждён автоматически</b>\n\n"
            f"Покупатель (ID <code>{tg_id}</code>) не ответил в течение 30 минут — "
            "заказ засчитан выполненным."
        )
    except Exception as e:
        logging.warning(f"Не удалось уведомить модератора об авто-подтверждении: {e}")

    # Сообщаем покупателю
    try:
        order = await db_get_order(order_id)
        has_review = await db_has_review(order_id)
        review_btn = [] if has_review else [
            [InlineKeyboardButton(text="⭐ Оставить отзыв", callback_data=f"review:{order_id}")]
        ]
        kb = InlineKeyboardMarkup(inline_keyboard=review_btn + [
            [InlineKeyboardButton(text="📦 Мои заказы", callback_data="orders")],
            [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
        ])
        await bot.send_message(
            tg_id,
            f"⏱ <b>Заказ #{order_id} подтверждён автоматически.</b>\n\n"
            "Вы не выбрали ответ в течение 30 минут, поэтому получение засчитано автоматически.\n\n"
            + ("⭐ Если хотите — оставьте отзыв о заказе." if not has_review else ""),
            parse_mode="HTML",
            reply_markup=kb,
        )
    except Exception as e:
        logging.warning(f"Не удалось отправить авто-подтверждение пользователю {tg_id}: {e}")


# =====================================================================
# Действия модератора по заказу
# =====================================================================


@dp.callback_query(F.data.startswith("ordone:"))
async def cb_order_done(call: CallbackQuery) -> None:
    """Модератор отмечает заказ как выполненный."""
    if not _is_moderator(call.from_user.id):
        await call.answer("Только для модератора.", show_alert=True)
        return
    try:
        _, oid_s, uid_s = call.data.split(":")
        order_id = int(oid_s)
        target_id = int(uid_s)
    except ValueError:
        await call.answer("Некорректные данные.", show_alert=True)
        return

    order = await db_get_order(order_id)
    if not order:
        await call.answer("Заказ не найден.", show_alert=True)
        return
    if order.get("status") == "Выполнен":
        await call.answer("Этот заказ уже отмечен как выполненный.", show_alert=True)
        return

    await db_set_order_status(order_id, "Выполнен")

    # Обработка реферальной программы
    referrer_id, level_up, new_level, _ = await db_process_referral(target_id)

    if referrer_id:
        ref_row = await db_find_user(referrer_id)
        ref_count = ref_row["referral_count"] if ref_row else 0
        next_level_refs = (new_level + 1) * REFERRAL_REFS_PER_LEVEL if new_level < 5 else None
        refs_left = (next_level_refs - ref_count) if next_level_refs else None

        if level_up:
            level_name = _ref_level_name(new_level)
            discount = _ref_discount_pct(new_level)
            try:
                await bot.send_message(
                    referrer_id,
                    f"🎉 <b>Поздравляем! Вы достигли нового уровня!</b>\n\n"
                    f"🏆 Ваш новый статус: <b>{level_name}</b>\n"
                    f"💰 Скидка на все заказы: <b>{discount}%</b>\n\n"
                    f"Скидка применяется автоматически при каждой покупке.\n"
                    f"<i>Не действует на Telegram Stars.</i>",
                    parse_mode="HTML",
                )
            except Exception as e:
                logging.warning(f"Не удалось уведомить о повышении уровня {referrer_id}: {e}")
        else:
            progress = f"\nДо следующего уровня: ещё <b>{refs_left}</b> реферал(ов)." if refs_left else ""
            try:
                await bot.send_message(
                    referrer_id,
                    f"✅ <b>+1 реферал засчитан!</b>\n\n"
                    f"Ваш приглашённый совершил первую покупку.\n"
                    f"Всего рефералов: <b>{ref_count}</b>.{progress}",
                    parse_mode="HTML",
                )
            except Exception as e:
                logging.warning(f"Не удалось уведомить реферера {referrer_id}: {e}")

    try:
        old = call.message.text or call.message.caption or ""
        await call.message.edit_text(
            f"{old}\n\n✅ <b>Заказ #{order_id} отмечен как выполненный.</b>",
            parse_mode="HTML",
        )
    except Exception:
        pass

    try:
        user_kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="✅ Подтвердить получение",
                        callback_data=f"confirm_receipt:{order_id}",
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="❌ Отклонить",
                        callback_data=f"reject_receipt:{order_id}",
                    )
                ],
            ]
        )
        await bot.send_message(
            target_id,
            f"🎉 <b>Ваш заказ #{order_id} выполнен!</b>\n\n"
            "⚠️ <b>Перед тем как нажать кнопку — проверьте наличие цифрового товара в игре.</b>\n\n"
            "✅ Нажмите <b>«Подтвердить получение»</b>, если товар зачислен.\n"
            "❌ Нажмите <b>«Отклонить»</b>, если товар не поступил. В этом случае потребуется "
            "прислать скриншот из игры в качестве доказательства.\n\n"
            "⚠️ <b>Внимание:</b> ложное отклонение может привести к блокировке в боте.",
            parse_mode="HTML",
            reply_markup=user_kb,
        )
        await db_set_confirm_pending(order_id, True)
        asyncio.create_task(_auto_confirm_receipt(order_id, target_id))
    except Exception as e:
        logging.warning(f"Не удалось уведомить пользователя {target_id}: {e}")

    try:
        await bot.unpin_chat_message(call.message.chat.id, call.message.message_id)
    except Exception:
        pass

    await call.answer("Заказ выполнен")


# =====================================================================
# Подтверждение / отклонение получения заказа
# =====================================================================


@dp.callback_query(F.data.startswith("confirm_receipt:"))
async def cb_confirm_receipt(call: CallbackQuery) -> None:
    """Покупатель подтверждает получение товара → предлагаем оставить отзыв."""
    try:
        order_id = int(call.data.split(":", 1)[1])
    except (ValueError, IndexError):
        await call.answer("Некорректные данные.", show_alert=True)
        return

    order = await db_get_order(order_id)
    if not order or order["tg_id"] != call.from_user.id:
        await call.answer("Заказ не найден.", show_alert=True)
        return

    # Если таймер уже сработал — флаг сброшен, просто показываем успех
    await db_set_confirm_pending(order_id, False)

    user = call.from_user
    username = f"@{user.username}" if user.username else "—"
    # Уведомляем модератора
    try:
        await notify_moderator(
            f"✅ <b>Покупатель подтвердил получение заказа #{order_id}</b>\n\n"
            f"Имя: {escape(user.first_name or '—')}\n"
            f"Username: {escape(username)}\n"
            f"Telegram ID: <code>{user.id}</code>"
        )
    except Exception as e:
        logging.warning(f"Не удалось уведомить модератора о подтверждении: {e}")

    has_review = await db_has_review(order_id)
    review_btn = [] if has_review else [
        [InlineKeyboardButton(text="⭐ Оставить отзыв", callback_data=f"review:{order_id}")]
    ]
    kb = InlineKeyboardMarkup(inline_keyboard=review_btn + [
        [InlineKeyboardButton(text="📦 Мои заказы", callback_data="orders")],
        [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
    ])
    await send_or_edit(
        call,
        f"✅ <b>Спасибо! Получение заказа #{order_id} подтверждено.</b>\n\n"
        "Рады, что всё прошло хорошо. Будем ждать вас снова!\n\n"
        + ("⭐ Если вам понравилось — оставьте, пожалуйста, отзыв. "
           "Это помогает другим покупателям и нам становиться лучше." if not has_review else ""),
        kb,
    )
    await call.answer()


@dp.callback_query(F.data.startswith("reject_receipt:"))
async def cb_reject_receipt(call: CallbackQuery, state: FSMContext) -> None:
    """Покупатель отклоняет получение → просим прислать скриншот из игры."""
    try:
        order_id = int(call.data.split(":", 1)[1])
    except (ValueError, IndexError):
        await call.answer("Некорректные данные.", show_alert=True)
        return

    order = await db_get_order(order_id)
    if not order or order["tg_id"] != call.from_user.id:
        await call.answer("Заказ не найден.", show_alert=True)
        return

    await state.set_state(ShopStates.waiting_reject_screenshot)
    await state.update_data(reject_order_id=order_id)

    kb_cancel = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ Отмена", callback_data="main")],
    ])
    await send_or_edit(
        call,
        f"⚠️ <b>Отклонение заказа #{order_id}</b>\n\n"
        "Пожалуйста, пришлите <b>скриншот из игры</b>, подтверждающий, "
        "что товар не был зачислен (например, скриншот инвентаря или баланса).\n\n"
        "Скриншот будет отправлен модератору для разбора ситуации.\n\n"
        "⚠️ Напоминаем: ложное отклонение может привести к блокировке.",
        kb_cancel,
    )
    await call.answer()


@dp.message(ShopStates.waiting_reject_screenshot)
async def msg_reject_screenshot(message: Message, state: FSMContext) -> None:
    """Получаем скриншот от покупателя и пересылаем модератору."""
    await _try_delete(message)
    data = await state.get_data()
    order_id = int(data.get("reject_order_id", 0))

    file_id: str | None = None
    is_document = False
    if message.photo:
        file_id = message.photo[-1].file_id
    elif message.document and (message.document.mime_type or "").startswith("image/"):
        file_id = message.document.file_id
        is_document = True

    if not file_id:
        kb_cancel = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data="main")],
        ])
        await message.answer(
            "⚠️ Пришлите изображение (фото или картинку файлом).",
            reply_markup=kb_cancel,
            parse_mode="HTML",
        )
        return

    await state.clear()
    await db_set_confirm_pending(order_id, False)

    user = message.from_user
    username = f"@{user.username}" if user.username else "—"

    # Уведомляем модератора
    caption = (
        f"🚫 <b>Покупатель отклонил получение заказа #{order_id}</b>\n\n"
        f"Имя: {escape(user.first_name or '—')}\n"
        f"Username: {escape(username)}\n"
        f"Telegram ID: <code>{user.id}</code>\n\n"
        "Скриншот от покупателя прикреплён ниже."
    )
    try:
        if is_document:
            await bot.send_document(
                MODERATOR_CHAT_ID, file_id, caption=caption, parse_mode="HTML"
            )
        else:
            await bot.send_photo(
                MODERATOR_CHAT_ID, file_id, caption=caption, parse_mode="HTML"
            )
    except Exception as e:
        logging.warning(f"Не удалось отправить скриншот отклонения модератору: {e}")

    await message.answer(
        f"✅ <b>Скриншот отправлен модератору.</b>\n\n"
        f"Мы рассмотрим ситуацию по заказу <b>#{order_id}</b> и свяжемся с вами.\n\n"
        "Если хотите написать напрямую — нажмите кнопку ниже.",
        parse_mode="HTML",
        reply_markup=kb_support(),
    )


# =====================================================================
# Отзывы покупателей
# =====================================================================


def kb_review_skip_photo(order_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="➡️ Без фото",
                    callback_data=f"review_skip_photo:{order_id}",
                )
            ],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="review_cancel")],
        ]
    )


def kb_review_cancel() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data="review_cancel")],
        ]
    )


def kb_review_mod(order_id: int, buyer_id: int, *msg_ids: int) -> InlineKeyboardMarkup:
    """Клавиатура модератора под управляющим сообщением отзыва.
    buyer_id  — tg_id покупателя для кнопки «Ответить».
    msg_ids   — id всех сообщений в чате модератора (заголовок + пересланные),
                передаются в callback_data для форварда в канал.
    """
    ids_str = ":".join(str(m) for m in msg_ids)
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="✅ Опубликовать в канал",
                    callback_data=f"pub_review:{order_id}:{ids_str}",
                ),
            ],
            [
                InlineKeyboardButton(
                    text="💬 Ответить покупателю",
                    callback_data=f"reply_buyer:{buyer_id}",
                ),
            ],
            [
                InlineKeyboardButton(
                    text="❌ Отклонить",
                    callback_data=f"rej_review:{order_id}",
                ),
            ],
        ]
    )


@dp.callback_query(F.data.startswith("review:"))
async def cb_review_start(call: CallbackQuery, state: FSMContext) -> None:
    try:
        order_id = int(call.data.split(":", 1)[1])
    except (ValueError, IndexError):
        await call.answer("Некорректные данные.", show_alert=True)
        return

    order = await db_get_order(order_id)
    if not order or order["tg_id"] != call.from_user.id:
        await call.answer("Заказ не найден.", show_alert=True)
        return
    if order["status"] != "Выполнен":
        await call.answer(
            "Отзыв можно оставить только после выполнения заказа.",
            show_alert=True,
        )
        return
    if await db_has_review(order_id):
        await call.answer("Вы уже оставили отзыв по этому заказу.", show_alert=True)
        return

    await state.clear()
    await state.set_state(ShopStates.waiting_review_photo)
    text = (
        f"⭐ <b>Отзыв по заказу #{order_id}</b>\n\n"
        "📷 Пришлите <b>фото</b> выполненного заказа одним сообщением "
        "(скриншот зачисления, скриншот игры и т.п.).\n\n"
        "Если фото нет — нажмите «Без фото»."
    )
    sent = await send_or_edit(call, text, kb_review_skip_photo(order_id))
    await state.update_data(review_order_id=order_id,
                            _prompt_chat_id=sent.chat.id, _prompt_msg_id=sent.message_id)
    await call.answer()


@dp.callback_query(F.data.startswith("review_skip_photo:"))
async def cb_review_skip_photo(call: CallbackQuery, state: FSMContext) -> None:
    try:
        order_id = int(call.data.split(":", 1)[1])
    except (ValueError, IndexError):
        await call.answer("Некорректные данные.", show_alert=True)
        return
    await state.set_state(ShopStates.waiting_review_text)
    await state.update_data(review_order_id=order_id, review_photo_id=None)
    text = (
        f"⭐ <b>Отзыв по заказу #{order_id}</b>\n\n"
        "✍️ Напишите ваш комментарий к отзыву одним сообщением."
    )
    await send_or_edit(call, text, kb_review_cancel())
    await call.answer()


@dp.callback_query(F.data == "review_cancel")
async def cb_review_cancel(call: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await send_or_edit(
        call,
        "❌ Отправка отзыва отменена.",
        InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="📦 Мои заказы", callback_data="orders")],
                [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
            ]
        ),
    )
    await call.answer()


@dp.message(ShopStates.waiting_review_photo)
async def msg_review_photo(message: Message, state: FSMContext) -> None:
    # Фото НЕ удаляем здесь — оно нужно для copy_message модератору.
    # Удалим его позже, в msg_review_text, после пересылки.
    file_id: str | None = None
    is_document = False
    if message.photo:
        file_id = message.photo[-1].file_id
    elif message.document and (message.document.mime_type or "").startswith("image/"):
        file_id = message.document.file_id
        is_document = True

    data = await state.get_data()
    order_id = int(data.get("review_order_id", 0))

    if not file_id:
        await _try_delete(message)
        await _edit_prompt(state,
                           "⚠️ Пришлите изображение (фото или картинку файлом). "
                           "Если фото нет — нажмите «Без фото».",
                           kb_review_skip_photo(order_id))
        return

    await state.update_data(
        review_photo_id=file_id,
        review_photo_is_document=is_document,
        # Сохраняем координаты фото для copy_message
        review_photo_msg_id=message.message_id,
        review_photo_chat_id=message.chat.id,
    )
    await state.set_state(ShopStates.waiting_review_text)
    sent = await message.answer(
        f"✅ Фото получено.\n\n"
        f"✍️ Теперь напишите комментарий к отзыву по заказу <b>#{order_id}</b> "
        "одним сообщением.",
        parse_mode="HTML",
        reply_markup=kb_review_cancel(),
    )
    await state.update_data(_prompt_chat_id=sent.chat.id, _prompt_msg_id=sent.message_id)


@dp.message(ShopStates.waiting_review_text)
async def msg_review_text(message: Message, state: FSMContext) -> None:
    # НЕ удаляем сообщение здесь — оно нужно для forward_message модератору.
    # Для ошибочных случаев удаляем вручную перед возвратом.
    comment = (message.text or "").strip()
    if not comment:
        await _try_delete(message)
        await _edit_prompt(state, "⚠️ Пустое сообщение. Напишите текст отзыва.",
                           kb_review_cancel())
        return
    if len(comment) > 2000:
        await _try_delete(message)
        await _edit_prompt(state, "⚠️ Комментарий слишком длинный (макс. 2000 символов).\n"
                           "Напишите покороче:",
                           kb_review_cancel())
        return

    data = await state.get_data()
    order_id = int(data.get("review_order_id", 0))
    photo_id = data.get("review_photo_id")
    photo_msg_id = data.get("review_photo_msg_id")
    photo_chat_id = data.get("review_photo_chat_id")
    await state.clear()

    if not order_id:
        await message.answer("Не удалось определить заказ. Попробуйте ещё раз.")
        return

    await db_add_review(order_id, message.from_user.id, photo_id, comment)

    order = await db_get_order(order_id)
    title = order["title"] if order else "—"

    # ----------------------------------------------------------------
    # Пересылаем отзыв модератору через forward_message ("Forwarded from").
    # forward_message не поддерживает reply_markup, поэтому кнопки
    # публикации отправляем отдельным управляющим сообщением после форварда.
    # ID пересланных сообщений сохраняем в callback_data для последующего
    # форварда в канал при нажатии «Опубликовать».
    # ----------------------------------------------------------------
    try:
        header = (
            f"⭐ <b>Новый отзыв</b>\n"
            f"🎁 Товар: <b>{escape(str(title))}</b>"
        )
        # Заголовок тоже сохраняем — он пересылается в канал вместе с отзывом
        header_msg = await bot.send_message(MODERATOR_CHAT_ID, header, parse_mode="HTML")
        buyer_id = message.from_user.id

        if photo_id and photo_msg_id and photo_chat_id:
            # Сначала форвардим — потом удаляем оригиналы
            photo_fwd = await bot.forward_message(
                chat_id=MODERATOR_CHAT_ID,
                from_chat_id=photo_chat_id,
                message_id=photo_msg_id,
            )
            text_fwd = await bot.forward_message(
                chat_id=MODERATOR_CHAT_ID,
                from_chat_id=message.chat.id,
                message_id=message.message_id,
            )
            await bot.delete_message(photo_chat_id, photo_msg_id)
            await _try_delete(message)
            await bot.send_message(
                MODERATOR_CHAT_ID,
                "⬆️ Управление отзывом:",
                reply_markup=kb_review_mod(
                    order_id, buyer_id,
                    header_msg.message_id, photo_fwd.message_id, text_fwd.message_id,
                ),
            )
        else:
            # Текстовый отзыв: форвардим сначала, удаляем после
            text_fwd = await bot.forward_message(
                chat_id=MODERATOR_CHAT_ID,
                from_chat_id=message.chat.id,
                message_id=message.message_id,
            )
            await _try_delete(message)
            await bot.send_message(
                MODERATOR_CHAT_ID,
                "⬆️ Управление отзывом:",
                reply_markup=kb_review_mod(
                    order_id, buyer_id,
                    header_msg.message_id, text_fwd.message_id,
                ),
            )
    except Exception as e:
        logging.warning(f"Не удалось переслать отзыв модератору: {e}")

    thanks_kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="⭐ Все отзывы", url=REVIEWS_URL)],
            [InlineKeyboardButton(text="📦 Мои заказы", callback_data="orders")],
            [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
        ]
    )
    await message.answer(
        f"✅ <b>Спасибо за отзыв по заказу #{order_id}!</b>\n\n"
        "Ваш отзыв передан модератору и скоро появится в нашем канале.",
        parse_mode="HTML",
        reply_markup=thanks_kb,
    )


@dp.callback_query(F.data.startswith("pub_review:"))
async def cb_publish_review(call: CallbackQuery) -> None:
    """Модератор публикует отзыв в канал — пересылает через forward_message."""
    if not _is_moderator(call.from_user.id):
        await call.answer("Нет доступа.", show_alert=True)
        return
    # callback_data: pub_review:{order_id}:{msg_id1}[:{msg_id2}]
    parts = call.data.split(":")
    # parts[0]="pub_review", parts[1]=order_id, parts[2+]=msg_ids
    mod_msg_ids = [int(p) for p in parts[2:] if p]
    try:
        for mid in mod_msg_ids:
            await bot.forward_message(
                chat_id=REVIEWS_CHANNEL,
                from_chat_id=call.message.chat.id,
                message_id=mid,
            )
        published_kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(
                    text="✅ Опубликовано в канале",
                    url=REVIEWS_URL,
                )]
            ]
        )
        await call.message.edit_reply_markup(reply_markup=published_kb)
        await call.answer("✅ Отзыв опубликован в канале!", show_alert=True)
    except Exception as e:
        logging.warning(f"Ошибка публикации отзыва: {e}")
        await call.answer(f"Ошибка: {e}", show_alert=True)


@dp.callback_query(F.data.startswith("rej_review:"))
async def cb_reject_review(call: CallbackQuery) -> None:
    """Модератор отклоняет отзыв."""
    if not _is_moderator(call.from_user.id):
        await call.answer("Нет доступа.", show_alert=True)
        return
    rejected_kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отклонён", callback_data="noop")]
        ]
    )
    await call.message.edit_reply_markup(reply_markup=rejected_kb)
    await call.answer("Отзыв отклонён.", show_alert=True)


@dp.callback_query(F.data.startswith("gpfix:"))
async def cb_gamepass_fix(call: CallbackQuery, state: FSMContext) -> None:
    """Модератор запускает запрос на изменение цены геймпасса."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        await call.answer("Только для модератора.", show_alert=True)
        return
    try:
        _, oid_s, uid_s, gp_s = call.data.split(":")
        order_id = int(oid_s)
        target_id = int(uid_s)
        suggested_price = int(gp_s)
    except ValueError:
        await call.answer("Некорректные данные.", show_alert=True)
        return

    await state.set_state(AdminStates.waiting_gp_price)
    await state.update_data(
        gp_order_id=order_id,
        gp_target_id=target_id,
        gp_admin_chat_id=call.message.chat.id,
        gp_admin_msg_id=call.message.message_id,
    )

    prompt = await call.message.answer(
        f"🎮 Введите новую цену геймпасса в R$ для заказа #{order_id}.\n"
        f"Расчётная цена: <b>{suggested_price} R$</b>.\n\n"
        "Отправьте число или /cancel для отмены.",
        parse_mode="HTML",
    )
    await state.update_data(gp_prompt_msg_id=prompt.message_id)


@dp.message(AdminStates.waiting_gp_price, Command("cancel"))
async def msg_gp_price_cancel(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer("Отменено.")


@dp.message(AdminStates.waiting_gp_price)
async def msg_gp_price_input(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    raw = (message.text or "").strip().replace(",", ".")
    try:
        gp_price = int(round(float(raw)))
        if gp_price <= 0:
            raise ValueError
    except ValueError:
        await message.answer("Введите положительное целое число (R$).")
        return

    data = await state.get_data()
    order_id = data.get("gp_order_id")
    target_id = data.get("gp_target_id")
    admin_chat_id = data.get("gp_admin_chat_id")
    admin_msg_id = data.get("gp_admin_msg_id")
    prompt_msg_id = data.get("gp_prompt_msg_id")
    await state.clear()

    if not (order_id and target_id):
        await message.answer("Контекст потерян. Откройте заявку заново.")
        return

    try:
        user_kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="⚙️ Действия по заказу",
                        callback_data=f"order_actions:{order_id}",
                    )
                ],
                [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
            ]
        )
        await bot.send_message(
            target_id,
            f"⚠️ <b>Просьба по заказу #{order_id}</b>\n\n"
            f"Пожалуйста, измените цену вашего геймпасса в Roblox на "
            f"<b>{gp_price} R$</b> и пришлите обновлённую ссылку через кнопку "
            "«🔗 Отправить ссылку на геймпасс».",
            parse_mode="HTML",
            reply_markup=user_kb,
        )
    except Exception as e:
        logging.warning(f"Не удалось уведомить пользователя {target_id}: {e}")
        await message.answer("Не удалось отправить запрос покупателю.")
        return

    if admin_chat_id and admin_msg_id:
        keep_kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="✅ Завершить заказ",
                        callback_data=f"ordone:{order_id}:{target_id}",
                    )
                ]
            ]
        )
        try:
            await bot.edit_message_reply_markup(
                chat_id=admin_chat_id,
                message_id=admin_msg_id,
                reply_markup=keep_kb,
            )
        except Exception:
            pass

    if prompt_msg_id:
        try:
            await bot.delete_message(admin_chat_id, prompt_msg_id)
        except Exception:
            pass
    try:
        await message.delete()
    except Exception:
        pass

    await bot.send_message(
        admin_chat_id,
        f"📨 Запрос на изменение цены геймпасса по заказу #{order_id} "
        f"отправлен покупателю: <b>{gp_price} R$</b>.",
        parse_mode="HTML",
    )


# =====================================================================
# Админ-панель
# =====================================================================


USERS_PAGE_SIZE = 8


def kb_admin_main() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🔍 Найти пользователя", callback_data="adm:find")],
            [
                InlineKeyboardButton(
                    text="👥 Список пользователей", callback_data="adm:users:0"
                )
            ],
            [InlineKeyboardButton(text="🎟️ Промокоды", callback_data="adm:promos")],
            [InlineKeyboardButton(text="🛒 Каталог и курсы", callback_data="adm:catalog")],
            [InlineKeyboardButton(text="❌ Закрыть", callback_data="adm:close")],
        ]
    )


# ---- вспомогательные словари для каталога ----

_SETTING_LABELS: dict[str, str] = {
    "robux_gamepass_rate":            "💎 Курс геймпасс (₽/робукс)",
    "robux_gamepass_pass_price_rate": "🎮 Делитель цены геймпасса",
    "tg_stars_rate":                  "⭐ Курс Telegram Stars (₽/звезда)",
    "min_topup":                      "💳 Минимум пополнения (₽)",
    "min_tg_stars":                   "⭐ Минимум Telegram Stars",
}


async def kb_catalog_categories() -> InlineKeyboardMarkup:
    cats = await db_all_categories()
    rows: list[list[InlineKeyboardButton]] = []
    for cat in cats:
        status = "⚠️ " if cat["disabled"] else ""
        rows.append([InlineKeyboardButton(
            text=f"{status}{cat['emoji']} {cat['name']}",
            callback_data=f"adm:catalog:cat:{cat['key']}",
        )])
    rows.append([InlineKeyboardButton(text="➕ Создать категорию", callback_data="adm:catalog:new_cat")])
    rows.append([InlineKeyboardButton(text="⚙️ Курсы и лимиты", callback_data="adm:settings")])
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="adm:main")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def kb_catalog_products_admin(products: list[dict], category: str) -> InlineKeyboardMarkup:
    """Клавиатура товаров категории + кнопка настроек категории."""
    rows = []
    for p in products:
        status = "✅" if p["active"] else "❌"
        price_str = f"{int(p['price'])}₽" if float(p["price"]) == int(p["price"]) else f"{p['price']}₽"
        rows.append([
            InlineKeyboardButton(
                text=f"{status} {p['name']} — {price_str}",
                callback_data=f"adm:prod:{p['key']}",
            ),
            InlineKeyboardButton(text="🔄", callback_data=f"adm:catalog:toggle:{p['key']}"),
        ])
    rows.append([InlineKeyboardButton(text="➕ Добавить товар", callback_data=f"adm:catalog:add:{category}")])
    rows.append([InlineKeyboardButton(text="⚙️ Настройки категории", callback_data=f"adm:catcfg:{category}")])
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="adm:catalog")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def kb_product_manage(prod_key: str, cat_key: str) -> InlineKeyboardMarkup:
    """Клавиатура управления конкретным товаром."""
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✏️ Переименовать", callback_data=f"adm:prod:rename:{prod_key}")],
        [InlineKeyboardButton(text="💰 Изменить цену",  callback_data=f"adm:catalog:price:{prod_key}")],
        [InlineKeyboardButton(text="🗑 Удалить товар",  callback_data=f"adm:prod:delask:{prod_key}")],
        [InlineKeyboardButton(text="⬅️ Назад к товарам", callback_data=f"adm:catalog:cat:{cat_key}")],
    ])


def kb_cat_config(cat: dict) -> InlineKeyboardMarkup:
    """Клавиатура настроек категории."""
    key = cat["key"]
    disabled = cat.get("disabled", False)
    needs_code = cat.get("needs_code", False)
    rows: list[list[InlineKeyboardButton]] = []
    if disabled:
        rows.append([InlineKeyboardButton(text="✅ Включить категорию", callback_data=f"adm:catdis:enable:{key}")])
    else:
        rows.append([InlineKeyboardButton(text="🔴 Отключить категорию", callback_data=f"adm:catdis:start:{key}")])
    rows.append([InlineKeyboardButton(text="📝 Изменить текст после оплаты", callback_data=f"adm:catlogin:{key}")])
    nc_label = "🔔 Выключить запрос кода" if needs_code else "🔔 Включить запрос кода"
    rows.append([InlineKeyboardButton(text=nc_label, callback_data=f"adm:catnc:{key}")])
    rows.append([
        InlineKeyboardButton(text="✏️ Переименовать", callback_data=f"adm:catname:{key}"),
        InlineKeyboardButton(text="🖼 Эмодзи",        callback_data=f"adm:catemoji:{key}"),
    ])
    rows.append([InlineKeyboardButton(text="🗑 Удалить категорию", callback_data=f"adm:catdel:ask:{key}")])
    rows.append([InlineKeyboardButton(text="⬅️ К товарам", callback_data=f"adm:catalog:cat:{key}")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def kb_settings_list(settings: list[dict]) -> InlineKeyboardMarkup:
    rows = []
    for s in settings:
        label = _SETTING_LABELS.get(s["key"], s["key"])
        rows.append([
            InlineKeyboardButton(
                text=f"{label}: {s['value']}",
                callback_data=f"adm:settings:edit:{s['key']}",
            )
        ])
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="adm:catalog")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _format_user_row(u: dict) -> str:
    handle = u.get("username")
    name = u.get("first_name") or "—"
    label = f"@{handle}" if handle else (name or f"id{u['tg_id']}")
    if u.get("is_blacklisted"):
        label = f"🚫 {label}"
    return f"{label} · {u['balance']}₽"


def kb_admin_users(users: list[dict], page: int, total: int) -> InlineKeyboardMarkup:
    rows = [
        [
            InlineKeyboardButton(
                text=_format_user_row(u),
                callback_data=f"adm:user:{u['tg_id']}",
            )
        ]
        for u in users
    ]
    nav = []
    if page > 0:
        nav.append(
            InlineKeyboardButton(text="⬅️", callback_data=f"adm:users:{page - 1}")
        )
    last_page = max(0, (total - 1) // USERS_PAGE_SIZE)
    nav.append(
        InlineKeyboardButton(
            text=f"{page + 1}/{last_page + 1}", callback_data="adm:noop"
        )
    )
    if page < last_page:
        nav.append(
            InlineKeyboardButton(text="➡️", callback_data=f"adm:users:{page + 1}")
        )
    if nav:
        rows.append(nav)
    rows.append(
        [
            InlineKeyboardButton(text="🔍 Найти", callback_data="adm:find"),
            InlineKeyboardButton(text="❌ Закрыть", callback_data="adm:close"),
        ]
    )
    return InlineKeyboardMarkup(inline_keyboard=rows)


def kb_admin_user(
    target_id: int, blocked: bool, from_list_page: int | None = None
) -> InlineKeyboardMarkup:
    block_btn = (
        InlineKeyboardButton(
            text="✅ Убрать из ЧС", callback_data=f"adm:unblock:{target_id}"
        )
        if blocked
        else InlineKeyboardButton(
            text="🚫 В чёрный список", callback_data=f"adm:block:{target_id}"
        )
    )
    if from_list_page is not None:
        nav_row = [
            InlineKeyboardButton(
                text="⬅️ К списку", callback_data=f"adm:users:{from_list_page}"
            ),
            InlineKeyboardButton(text="❌ Закрыть", callback_data="adm:close"),
        ]
    else:
        nav_row = [
            InlineKeyboardButton(
                text="🔍 Другой пользователь", callback_data="adm:find"
            ),
            InlineKeyboardButton(text="❌ Закрыть", callback_data="adm:close"),
        ]
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="➕ Начислить баланс", callback_data=f"adm:credit:{target_id}"
                )
            ],
            [
                InlineKeyboardButton(
                    text="♻️ Обнулить баланс", callback_data=f"adm:reset:{target_id}"
                )
            ],
            [block_btn],
            [
                InlineKeyboardButton(
                    text="📜 История транзакций",
                    callback_data=f"adm:tx:{target_id}",
                )
            ],
            [
                InlineKeyboardButton(
                    text="👑 Выдать реф. статус",
                    callback_data=f"adm:ref_level:{target_id}",
                )
            ],
            [
                InlineKeyboardButton(
                    text="🎟️ Реф. промокоды",
                    callback_data=f"adm:ref_promos:{target_id}",
                )
            ],
            nav_row,
        ]
    )


async def _get_admin_card_source(state: FSMContext) -> int | None:
    """Возвращает номер страницы списка, если карточка открыта из списка, иначе None."""
    data = await state.get_data()
    page = data.get("admin_user_from_list_page")
    return int(page) if page is not None else None


async def _send_admin_user_card(
    call: CallbackQuery, target_id: int, state: FSMContext
) -> None:
    user_row = await db_find_user(target_id)
    if not user_row:
        await send_or_edit(
            call,
            f"❌ Пользователь с ID <code>{target_id}</code> не найден.",
            kb_admin_main(),
        )
        return
    blocked = bool(user_row.get("is_blacklisted"))
    orders_cnt = await db_orders_count(target_id)
    from_list_page = await _get_admin_card_source(state)
    level = user_row.get("referral_level", 0) or 0
    ref_count = user_row.get("referral_count", 0) or 0
    level_name = _ref_level_name(level)
    discount = _ref_discount_pct(level)
    text = (
        "<b>👤 Карточка пользователя</b>\n\n"
        f"Telegram ID: <code>{user_row['tg_id']}</code>\n"
        f"Username: @{escape(user_row.get('username') or '—')}\n"
        f"Имя: {escape(user_row.get('first_name') or '—')}\n"
        f"Баланс: <b>{user_row['balance']}₽</b>\n"
        f"Заказов: <b>{orders_cnt}</b>\n"
        f"Реф. статус: <b>{level_name}</b> (скидка {discount}%)\n"
        f"Рефералов засчитано: <b>{ref_count}</b>\n"
        f"Чёрный список: {'<b>да</b>' if blocked else 'нет'}"
    )
    await send_or_edit(
        call, text, kb_admin_user(target_id, blocked, from_list_page)
    )


@dp.message(Command("admin"))
async def cmd_admin(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await state.clear()
    await message.answer(
        "<b>🛠️ Админ-панель</b>\n\nВыберите действие:",
        reply_markup=kb_admin_main(),
        parse_mode="HTML",
    )


@dp.callback_query(F.data == "adm:close")
async def cb_adm_close(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    await state.clear()
    try:
        await call.message.delete()
    except Exception:
        pass
    await call.answer("Закрыто")


@dp.callback_query(F.data == "adm:noop")
async def cb_adm_noop(call: CallbackQuery) -> None:
    await call.answer()


# =====================================================================
# Управление каталогом и настройками (модератор)
# =====================================================================

@dp.callback_query(F.data == "adm:main")
async def cb_adm_main(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    await state.clear()
    await send_or_edit(call, "<b>🛠️ Админ-панель</b>\n\nВыберите действие:", kb_admin_main())
    await call.answer()


@dp.callback_query(F.data == "adm:catalog")
async def cb_adm_catalog(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    await state.clear()
    await send_or_edit(
        call,
        "<b>🛒 Каталог и курсы</b>\n\nВыберите категорию или откройте настройки:",
        await kb_catalog_categories(),
    )
    await call.answer()


@dp.callback_query(F.data.startswith("adm:catalog:cat:"))
async def cb_adm_catalog_cat(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    category = call.data.split(":", 3)[3]
    await state.update_data(catalog_category=category)
    cat = await db_get_category(category)
    label = f"{cat['emoji']} {cat['name']}" if cat else category
    all_prods = await db_all_products()
    cat_prods = [p for p in all_prods if p["category"] == category]
    dis_note = "\n⚠️ <i>Категория отключена</i>" if cat and cat.get("disabled") else ""
    if not cat_prods:
        text = f"<b>{escape(label)}</b>{dis_note}\n\nТоваров пока нет."
    else:
        lines = [
            ("✅" if p["active"] else "❌") + f" {p['name']} — {p['price']}₽"
            for p in cat_prods
        ]
        text = (f"<b>{escape(label)}</b>{dis_note}\n\n" + "\n".join(lines) +
                "\n\n✅ = активен  ❌ = скрыт\nНажмите на товар, чтобы изменить цену.\n🔄 — переключить активность.")
    await send_or_edit(call, text, kb_catalog_products_admin(cat_prods, category))
    await call.answer()


@dp.callback_query(F.data.startswith("adm:catalog:toggle:"))
async def cb_adm_catalog_toggle(call: CallbackQuery) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    key = call.data.split(":", 3)[3]
    new_state = await db_toggle_product(key)
    status_text = "активирован ✅" if new_state else "скрыт ❌"
    await call.answer(f"Товар {status_text}", show_alert=True)
    all_prods = await db_all_products()
    prod = next((p for p in all_prods if p["key"] == key), None)
    if prod:
        category = prod["category"]
        cat = await db_get_category(category)
        label = f"{cat['emoji']} {cat['name']}" if cat else category
        cat_prods = [p for p in all_prods if p["category"] == category]
        dis_note = "\n⚠️ <i>Категория отключена</i>" if cat and cat.get("disabled") else ""
        lines = [("✅" if p["active"] else "❌") + f" {p['name']} — {p['price']}₽" for p in cat_prods]
        text = (f"<b>{escape(label)}</b>{dis_note}\n\n" + "\n".join(lines) +
                "\n\n✅ = активен  ❌ = скрыт\nНажмите на товар, чтобы изменить цену.\n🔄 — переключить активность.")
        try:
            await call.message.edit_text(text, reply_markup=kb_catalog_products_admin(cat_prods, category), parse_mode="HTML")
        except Exception:
            pass


@dp.callback_query(F.data.startswith("adm:catalog:price:"))
async def cb_adm_catalog_price(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    key = call.data.split(":", 3)[3]
    product = await db_get_product_by_key(key)
    if not product:
        await call.answer("Товар не найден.", show_alert=True)
        return
    _, name, current_price, _ = product
    await state.set_state(AdminStates.waiting_edit_product_price)
    await state.update_data(edit_product_key=key, edit_product_name=name)
    await send_or_edit(
        call,
        f"✏️ <b>Изменение цены: {escape(name)}</b>\n\n"
        f"Текущая цена: <b>{current_price}₽</b>\n\n"
        "Введите новую цену в рублях (целое число):",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data=f"adm:catalog:cat:none")],
        ]),
    )
    await call.answer()


@dp.message(AdminStates.waiting_edit_product_price)
async def msg_adm_edit_product_price(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    price = parse_positive_int(message.text)
    if price is None or price <= 0:
        await message.answer("⚠️ Введите положительное целое число (цену в рублях).")
        return
    data = await state.get_data()
    key = data.get("edit_product_key", "")
    name = data.get("edit_product_name", key)
    await state.clear()
    await db_update_product_price(key, float(price))
    await message.answer(
        f"✅ Цена <b>{escape(name)}</b> обновлена: <b>{price}₽</b>",
        parse_mode="HTML",
        reply_markup=kb_admin_main(),
    )


@dp.callback_query(F.data.startswith("adm:catalog:add:"))
async def cb_adm_catalog_add(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    category = call.data.split(":", 3)[3]
    await state.set_state(AdminStates.waiting_new_product_key)
    await state.update_data(new_product_category=category)
    cat = await db_get_category(category)
    label = f"{cat['emoji']} {cat['name']}" if cat else category
    await send_or_edit(
        call,
        f"➕ <b>Новый товар в «{escape(label)}»</b>\n\n"
        "Шаг 1/4. Введите <b>уникальный ключ</b> товара (латинские буквы, цифры, _).\n"
        "Пример: <code>rb_5000</code>",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:catalog")],
        ]),
    )
    await call.answer()


@dp.message(AdminStates.waiting_new_product_key)
async def msg_adm_new_product_key(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    raw = (message.text or "").strip()
    if not raw or not all(c.isalnum() or c == "_" for c in raw):
        await message.answer("⚠️ Ключ может содержать только латинские буквы, цифры и _.\nПопробуйте ещё раз:")
        return
    await state.update_data(new_product_key=raw)
    await state.set_state(AdminStates.waiting_new_product_name)
    await message.answer(
        "Шаг 2/4. Введите <b>название</b> товара (отображается покупателям).\n"
        "Пример: <code>5000 робуксов</code>",
        parse_mode="HTML",
    )


@dp.message(AdminStates.waiting_new_product_name)
async def msg_adm_new_product_name(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    name = (message.text or "").strip()
    if not name:
        await message.answer("⚠️ Название не может быть пустым. Попробуйте ещё раз:")
        return
    await state.update_data(new_product_name=name)
    await state.set_state(AdminStates.waiting_new_product_price)
    await message.answer("Шаг 3/4. Введите <b>цену</b> в рублях (целое число):", parse_mode="HTML")


@dp.message(AdminStates.waiting_new_product_price)
async def msg_adm_new_product_price(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    price = parse_positive_int(message.text)
    if price is None or price <= 0:
        await message.answer("⚠️ Введите положительное целое число. Попробуйте ещё раз:")
        return
    await state.update_data(new_product_price=price)
    await state.set_state(AdminStates.waiting_new_product_delivery)
    await message.answer(
        "Шаг 4/4. Введите <b>способ выдачи</b> товара.\n"
        "Пример: <code>моментально</code>",
        parse_mode="HTML",
    )


@dp.message(AdminStates.waiting_new_product_delivery)
async def msg_adm_new_product_delivery(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    delivery = (message.text or "").strip()
    if not delivery:
        await message.answer("⚠️ Способ выдачи не может быть пустым. Попробуйте ещё раз:")
        return
    data = await state.get_data()
    category = data.get("new_product_category", "roblox_instant")
    key = data.get("new_product_key", "")
    name = data.get("new_product_name", "")
    price = float(data.get("new_product_price", 0))
    await state.clear()
    ok = await db_add_product(category, key, name, price, delivery)
    if ok:
        await message.answer(
            f"✅ Товар <b>{escape(name)}</b> добавлен!\n"
            f"Ключ: <code>{key}</code>  Цена: <b>{int(price)}₽</b>",
            parse_mode="HTML",
            reply_markup=kb_admin_main(),
        )
    else:
        await message.answer(
            f"❌ Товар с ключом <code>{key}</code> уже существует. Выберите другой ключ.",
            parse_mode="HTML",
            reply_markup=kb_admin_main(),
        )


# ---- Управление категориями ----


@dp.callback_query(F.data == "adm:catalog:new_cat")
async def cb_adm_catalog_new_cat(call: CallbackQuery, state: FSMContext) -> None:
    """Начало создания новой категории."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    await state.set_state(AdminStates.waiting_new_cat_key)
    await send_or_edit(
        call,
        "➕ <b>Новая категория — Шаг 1/3</b>\n\n"
        "Введите <b>уникальный ключ</b> категории (латинские буквы, цифры, _).\n"
        "Это внутренний идентификатор, покупатели его не видят.\n"
        "Пример: <code>minecraft</code>",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:catalog")],
        ]),
    )


@dp.message(AdminStates.waiting_new_cat_key)
async def msg_adm_new_cat_key(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    raw = (message.text or "").strip().lower()
    reserved = {"roblox_instant", "roblox_gamepass", "brawl", "tgstars", "other"}
    if not raw or not all(c.isalnum() or c == "_" for c in raw):
        await message.answer(
            "⚠️ Ключ может содержать только латинские буквы, цифры и _. Попробуйте ещё раз:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:catalog")],
            ]),
        )
        return
    if raw in reserved:
        await message.answer(
            f"⚠️ Ключ <code>{raw}</code> зарезервирован. Выберите другой:",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:catalog")],
            ]),
        )
        return
    existing = await db_get_category(raw)
    if existing:
        await message.answer(
            f"⚠️ Категория с ключом <code>{raw}</code> уже существует.",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:catalog")],
            ]),
        )
        return
    await state.update_data(new_cat_key=raw)
    await state.set_state(AdminStates.waiting_new_cat_name)
    await message.answer(
        f"Ключ: <code>{raw}</code>\n\n"
        "➕ <b>Новая категория — Шаг 2/3</b>\n\n"
        "Введите <b>название категории</b> (видят покупатели).\n"
        "Пример: <code>Minecraft</code>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:catalog")],
        ]),
    )


@dp.message(AdminStates.waiting_new_cat_name)
async def msg_adm_new_cat_name(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    name = (message.text or "").strip()
    if not name:
        await message.answer("⚠️ Название не может быть пустым. Попробуйте ещё раз:")
        return
    await state.update_data(new_cat_name=name)
    await state.set_state(AdminStates.waiting_new_cat_emoji)
    data = await state.get_data()
    await message.answer(
        f"Ключ: <code>{data['new_cat_key']}</code>  Название: <b>{escape(name)}</b>\n\n"
        "➕ <b>Новая категория — Шаг 3/3</b>\n\n"
        "Введите <b>эмодзи</b> для категории (одним символом).\n"
        "Пример: <code>⛏️</code>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🎮 Пропустить (по умолчанию)", callback_data="adm:newcat:emoji_default")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:catalog")],
        ]),
    )


@dp.callback_query(F.data == "adm:newcat:emoji_default")
async def cb_adm_newcat_emoji_default(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    await _finish_create_category(call, state, emoji="🎮")


@dp.message(AdminStates.waiting_new_cat_emoji)
async def msg_adm_new_cat_emoji(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    emoji = (message.text or "").strip()
    if not emoji:
        emoji = "🎮"
    await _finish_create_category(message, state, emoji=emoji)


async def _finish_create_category(target, state: FSMContext, emoji: str) -> None:
    data = await state.get_data()
    key = data.get("new_cat_key", "")
    name = data.get("new_cat_name", "")
    await state.clear()
    await db_create_category(key, name, emoji)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🛒 В каталог", callback_data="adm:catalog")],
        [InlineKeyboardButton(text="⚙️ Настроить категорию", callback_data=f"adm:catcfg:{key}")],
    ])
    msg_text = (
        f"✅ <b>Категория создана!</b>\n\n"
        f"Ключ: <code>{key}</code>\n"
        f"Название: <b>{escape(name)}</b>\n"
        f"Эмодзи: {emoji}\n\n"
        "Теперь добавьте товары и настройте текст после оплаты."
    )
    if hasattr(target, "message"):  # CallbackQuery
        await send_or_edit(target, msg_text, kb)
    else:
        await target.answer(msg_text, parse_mode="HTML", reply_markup=kb)


@dp.callback_query(F.data.startswith("adm:catcfg:"))
async def cb_adm_catcfg(call: CallbackQuery, state: FSMContext) -> None:
    """Страница настроек категории."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    key = call.data.split(":", 2)[2]
    cat = await db_get_category(key)
    if not cat:
        await call.answer("Категория не найдена.", show_alert=True)
        return
    dis_icon = "⚠️ " if cat["disabled"] else "✅ "
    hint_preview = (cat.get("login_hint") or "—")[:60]
    if len(cat.get("login_hint") or "") > 60:
        hint_preview += "…"
    nc_icon = "🔔✅" if cat["needs_code"] else "🔔❌"
    text = (
        f"⚙️ <b>Настройки: {cat['emoji']} {escape(cat['name'])}</b>\n\n"
        f"Статус: {dis_icon}{'отключена' if cat['disabled'] else 'активна'}\n"
        f"Текст после оплаты: <i>{escape(hint_preview)}</i>\n"
        f"Кнопка кода: {nc_icon} {'включена' if cat['needs_code'] else 'выключена'}"
    )
    if cat["disabled"] and cat.get("disabled_reason"):
        text += f"\nПричина отключения: <i>{escape(cat['disabled_reason'])}</i>"
    await send_or_edit(call, text, kb_cat_config(cat))


@dp.callback_query(F.data.startswith("adm:catdis:enable:"))
async def cb_adm_catdis_enable(call: CallbackQuery) -> None:
    """Включить категорию."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    key = call.data.split(":", 3)[3]
    await db_set_category_disabled(key, False, None)
    await call.answer("✅ Категория включена.", show_alert=True)
    cat = await db_get_category(key)
    if cat:
        await send_or_edit(call, (
            f"⚙️ <b>Настройки: {cat['emoji']} {escape(cat['name'])}</b>\n\n"
            f"Статус: ✅ активна"
        ), kb_cat_config(cat))


@dp.callback_query(F.data.startswith("adm:catdis:start:"))
async def cb_adm_catdis_start(call: CallbackQuery, state: FSMContext) -> None:
    """Начало отключения категории — запрашиваем причину."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    key = call.data.split(":", 3)[3]
    await state.set_state(AdminStates.waiting_cat_disabled_reason)
    await state.update_data(cat_disable_key=key)
    await send_or_edit(
        call,
        "🔴 <b>Отключение категории</b>\n\n"
        "Введите <b>причину</b>, которую увидит покупатель при попытке открыть эту категорию.\n\n"
        "Или нажмите кнопку, чтобы отключить без объяснения:",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⚡ Без объяснения", callback_data=f"adm:catdis:noreason:{key}")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data=f"adm:catcfg:{key}")],
        ]),
    )


@dp.callback_query(F.data.startswith("adm:catdis:noreason:"))
async def cb_adm_catdis_noreason(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    key = call.data.split(":", 3)[3]
    await state.clear()
    await db_set_category_disabled(key, True, None)
    await call.answer("🔴 Категория отключена.", show_alert=True)
    cat = await db_get_category(key)
    if cat:
        await send_or_edit(call, (
            f"⚙️ <b>Настройки: {cat['emoji']} {escape(cat['name'])}</b>\n\n"
            f"Статус: ⚠️ отключена"
        ), kb_cat_config(cat))


@dp.message(AdminStates.waiting_cat_disabled_reason)
async def msg_adm_cat_disabled_reason(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    reason = (message.text or "").strip()
    if not reason:
        await message.answer("⚠️ Введите причину или нажмите «Без объяснения».")
        return
    data = await state.get_data()
    key = data.get("cat_disable_key", "")
    await state.clear()
    await db_set_category_disabled(key, True, reason)
    cat = await db_get_category(key)
    name = f"{cat['emoji']} {cat['name']}" if cat else key
    await message.answer(
        f"🔴 Категория <b>{escape(name)}</b> отключена.\n"
        f"Причина: <i>{escape(reason)}</i>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⚙️ Настройки категории", callback_data=f"adm:catcfg:{key}")],
            [InlineKeyboardButton(text="🛒 Каталог", callback_data="adm:catalog")],
        ]),
    )


@dp.callback_query(F.data.startswith("adm:catnc:"))
async def cb_adm_cat_needscode(call: CallbackQuery) -> None:
    """Переключить флаг 'нужен код' для категории."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    key = call.data.split(":", 2)[2]
    new_val = await db_toggle_category_needs_code(key)
    label = "включён ✅" if new_val else "выключен ❌"
    await call.answer(f"Запрос кода {label}", show_alert=True)
    cat = await db_get_category(key)
    if cat:
        await send_or_edit(call, (
            f"⚙️ <b>Настройки: {cat['emoji']} {escape(cat['name'])}</b>\n\n"
            f"Кнопка «Отправить код»: {'✅ включена' if new_val else '❌ выключена'}"
        ), kb_cat_config(cat))


@dp.callback_query(F.data.startswith("adm:catlogin:"))
async def cb_adm_catlogin(call: CallbackQuery, state: FSMContext) -> None:
    """Начало изменения текста после оплаты."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    key = call.data.split(":", 2)[2]
    cat = await db_get_category(key)
    if not cat:
        await call.answer("Категория не найдена.", show_alert=True)
        return
    current = cat.get("login_hint") or "—"
    await state.set_state(AdminStates.waiting_cat_login_hint)
    await state.update_data(cat_login_key=key)
    await send_or_edit(
        call,
        f"📝 <b>Текст после оплаты: {cat['emoji']} {escape(cat['name'])}</b>\n\n"
        f"Текущий текст:\n<i>{escape(current)}</i>\n\n"
        "Введите новый текст. Он появится в сообщении покупателю после успешной оплаты "
        "рядом с кнопкой «Отправить данные для входа».\n\n"
        "Если текст пустой — кнопка входа не показывается.",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🗑 Убрать текст (кнопка пропадёт)", callback_data=f"adm:catlogin:clear:{key}")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data=f"adm:catcfg:{key}")],
        ]),
    )


@dp.callback_query(F.data.startswith("adm:catlogin:clear:"))
async def cb_adm_catlogin_clear(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    key = call.data.split(":", 3)[3]
    await state.clear()
    await db_update_category_login_hint(key, None)
    await call.answer("🗑 Текст убран.", show_alert=True)
    cat = await db_get_category(key)
    if cat:
        await send_or_edit(call, (
            f"⚙️ <b>Настройки: {cat['emoji']} {escape(cat['name'])}</b>\n\n"
            "Текст после оплаты: —"
        ), kb_cat_config(cat))


@dp.message(AdminStates.waiting_cat_login_hint)
async def msg_adm_cat_login_hint(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    text = (message.text or "").strip()
    if not text:
        await message.answer("⚠️ Введите текст или нажмите «Убрать текст».")
        return
    data = await state.get_data()
    key = data.get("cat_login_key", "")
    await state.clear()
    await db_update_category_login_hint(key, text)
    cat = await db_get_category(key)
    name = f"{cat['emoji']} {cat['name']}" if cat else key
    await message.answer(
        f"✅ Текст после оплаты для <b>{escape(name)}</b> обновлён:\n\n"
        f"<i>{escape(text)}</i>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⚙️ Настройки категории", callback_data=f"adm:catcfg:{key}")],
            [InlineKeyboardButton(text="🛒 Каталог", callback_data="adm:catalog")],
        ]),
    )


# ---- Управление товарами (страница товара) ----


@dp.callback_query(F.data.startswith("adm:prod:rename:"))
async def cb_adm_prod_rename(call: CallbackQuery, state: FSMContext) -> None:
    """Начало переименования товара."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    prod_key = call.data.split(":", 3)[3]
    product = await db_get_product_by_key(prod_key)
    if not product:
        await call.answer("Товар не найден.", show_alert=True)
        return
    _, current_name, _, _ = product
    await state.set_state(AdminStates.waiting_rename_product)
    await state.update_data(rename_prod_key=prod_key)
    await send_or_edit(
        call,
        f"✏️ <b>Переименование товара</b>\n\n"
        f"Текущее название: <b>{escape(current_name)}</b>\n\n"
        "Введите новое название:",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data=f"adm:prod:{prod_key}")],
        ]),
    )


@dp.message(AdminStates.waiting_rename_product)
async def msg_adm_rename_product(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    new_name = (message.text or "").strip()
    if not new_name:
        await message.answer("⚠️ Название не может быть пустым. Попробуйте ещё раз:")
        return
    data = await state.get_data()
    prod_key = data.get("rename_prod_key", "")
    cat_key = data.get("prod_manage_cat", "")
    await state.clear()
    await db_rename_product(prod_key, new_name)
    await message.answer(
        f"✅ Товар переименован: <b>{escape(new_name)}</b>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📦 К товару", callback_data=f"adm:prod:{prod_key}")],
            [InlineKeyboardButton(text="⬅️ К списку", callback_data=f"adm:catalog:cat:{cat_key}")],
        ]),
    )


@dp.callback_query(F.data.startswith("adm:prod:delask:"))
async def cb_adm_prod_delask(call: CallbackQuery, state: FSMContext) -> None:
    """Подтверждение удаления товара."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    prod_key = call.data.split(":", 3)[3]
    product = await db_get_product_by_key(prod_key)
    if not product:
        await call.answer("Товар не найден.", show_alert=True)
        return
    _, name, price, _ = product
    price_str = f"{int(price)}₽" if price == int(price) else f"{price}₽"
    await send_or_edit(
        call,
        f"🗑 <b>Удалить товар?</b>\n\n"
        f"<b>{escape(name)}</b> — {price_str}\n\n"
        "⚠️ Это действие необратимо. Все данные о товаре будут удалены.\n"
        "История заказов на этот товар сохранится.",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Да, удалить", callback_data=f"adm:prod:delyes:{prod_key}")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data=f"adm:prod:{prod_key}")],
        ]),
    )


@dp.callback_query(F.data.startswith("adm:prod:delyes:"))
async def cb_adm_prod_delyes(call: CallbackQuery, state: FSMContext) -> None:
    """Выполнить удаление товара."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    prod_key = call.data.split(":", 3)[3]
    product = await db_get_product_by_key(prod_key)
    name = product[1] if product else prod_key
    all_prods = await db_all_products()
    prod_row = next((p for p in all_prods if p["key"] == prod_key), None)
    cat_key = prod_row["category"] if prod_row else ""
    await db_delete_product(prod_key)
    await send_or_edit(
        call,
        f"🗑 Товар <b>{escape(name)}</b> удалён.",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ К списку товаров", callback_data=f"adm:catalog:cat:{cat_key}")],
        ]),
    )


@dp.callback_query(F.data.startswith("adm:prod:"))
async def cb_adm_prod(call: CallbackQuery, state: FSMContext) -> None:
    """Страница управления конкретным товаром."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    parts = call.data.split(":", 2)
    prod_key = parts[2]

    product = await db_get_product_by_key(prod_key)
    if not product:
        await call.answer("Товар не найден.", show_alert=True)
        return
    key, name, price, delivery = product
    # Определяем категорию
    all_prods = await db_all_products()
    prod_row = next((p for p in all_prods if p["key"] == prod_key), None)
    cat_key = prod_row["category"] if prod_row else ""
    cat = await db_get_category(cat_key) if cat_key else None
    cat_label = f"{cat['emoji']} {cat['name']}" if cat else cat_key
    status = "✅ активен" if (prod_row and prod_row.get("active")) else "❌ скрыт"
    price_str = f"{int(price)}₽" if price == int(price) else f"{price}₽"
    text = (
        f"📦 <b>{escape(name)}</b>\n\n"
        f"Цена: <b>{price_str}</b>\n"
        f"Статус: {status}\n"
        f"Выдача: {escape(delivery)}\n"
        f"Категория: {escape(cat_label)}"
    )
    await send_or_edit(call, text, kb_product_manage(prod_key, cat_key))
    await state.update_data(prod_manage_cat=cat_key)


# ---- Переименование категории ----


@dp.callback_query(F.data.startswith("adm:catname:"))
async def cb_adm_catname(call: CallbackQuery, state: FSMContext) -> None:
    """Начало переименования категории."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    key = call.data.split(":", 2)[2]
    cat = await db_get_category(key)
    if not cat:
        await call.answer("Категория не найдена.", show_alert=True)
        return
    await state.set_state(AdminStates.waiting_rename_cat_name)
    await state.update_data(rename_cat_key=key)
    await send_or_edit(
        call,
        f"✏️ <b>Переименование категории</b>\n\n"
        f"Текущее название: <b>{escape(cat['name'])}</b>\n\n"
        "Введите новое название (покупатели увидят его в магазине):",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data=f"adm:catcfg:{key}")],
        ]),
    )


@dp.message(AdminStates.waiting_rename_cat_name)
async def msg_adm_rename_cat_name(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    new_name = (message.text or "").strip()
    if not new_name:
        await message.answer("⚠️ Название не может быть пустым. Попробуйте ещё раз:")
        return
    data = await state.get_data()
    key = data.get("rename_cat_key", "")
    await state.clear()
    await db_rename_category(key, new_name)
    cat = await db_get_category(key)
    label = f"{cat['emoji']} {new_name}" if cat else new_name
    await message.answer(
        f"✅ Категория переименована: <b>{escape(label)}</b>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⚙️ Настройки категории", callback_data=f"adm:catcfg:{key}")],
            [InlineKeyboardButton(text="🛒 Каталог", callback_data="adm:catalog")],
        ]),
    )


@dp.callback_query(F.data.startswith("adm:catemoji:"))
async def cb_adm_catemoji(call: CallbackQuery, state: FSMContext) -> None:
    """Начало смены эмодзи категории."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    key = call.data.split(":", 2)[2]
    cat = await db_get_category(key)
    if not cat:
        await call.answer("Категория не найдена.", show_alert=True)
        return
    await state.set_state(AdminStates.waiting_rename_cat_emoji)
    await state.update_data(rename_cat_key=key)
    await send_or_edit(
        call,
        f"🖼 <b>Эмодзи категории</b>\n\n"
        f"Текущий эмодзи: {cat['emoji']}\n\n"
        "Отправьте новый эмодзи одним сообщением:",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data=f"adm:catcfg:{key}")],
        ]),
    )


@dp.message(AdminStates.waiting_rename_cat_emoji)
async def msg_adm_rename_cat_emoji(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    emoji = (message.text or "").strip()
    if not emoji:
        await message.answer("⚠️ Эмодзи не может быть пустым. Попробуйте ещё раз:")
        return
    data = await state.get_data()
    key = data.get("rename_cat_key", "")
    await state.clear()
    await db_set_category_emoji(key, emoji)
    cat = await db_get_category(key)
    label = f"{emoji} {cat['name']}" if cat else emoji
    await message.answer(
        f"✅ Эмодзи категории обновлён: <b>{escape(label)}</b>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⚙️ Настройки категории", callback_data=f"adm:catcfg:{key}")],
            [InlineKeyboardButton(text="🛒 Каталог", callback_data="adm:catalog")],
        ]),
    )


# ---- Удаление категории ----


@dp.callback_query(F.data.startswith("adm:catdel:ask:"))
async def cb_adm_catdel_ask(call: CallbackQuery, state: FSMContext) -> None:
    """Предупреждение перед удалением категории."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    key = call.data.split(":", 3)[3]
    cat = await db_get_category(key)
    if not cat:
        await call.answer("Категория не найдена.", show_alert=True)
        return
    all_prods = await db_all_products()
    prod_count = sum(1 for p in all_prods if p["category"] == key)
    await state.set_state(AdminStates.waiting_cat_delete_confirm)
    await state.update_data(catdel_key=key)
    await send_or_edit(
        call,
        f"🗑 <b>Удаление категории: {cat['emoji']} {escape(cat['name'])}</b>\n\n"
        f"Вместе с категорией будут удалены <b>{prod_count} товаров</b>.\n"
        "История заказов сохранится.\n\n"
        "⚠️ Это действие <b>необратимо</b>.\n\n"
        "Для подтверждения напишите слово <code>УДАЛИТЬ</code>:",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data=f"adm:catcfg:{key}")],
        ]),
    )


@dp.message(AdminStates.waiting_cat_delete_confirm)
async def msg_adm_catdel_confirm(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    data = await state.get_data()
    key = data.get("catdel_key", "")
    if (message.text or "").strip() != "УДАЛИТЬ":
        await message.answer(
            "⚠️ Неверное слово. Напишите ровно <code>УДАЛИТЬ</code> для подтверждения:",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="❌ Отмена", callback_data=f"adm:catcfg:{key}")],
            ]),
        )
        return
    cat = await db_get_category(key)
    name = f"{cat['emoji']} {cat['name']}" if cat else key
    await state.clear()
    await db_delete_category(key)
    await message.answer(
        f"🗑 Категория <b>{escape(name)}</b> и все её товары удалены.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🛒 Каталог", callback_data="adm:catalog")],
        ]),
    )


# ---- Настройки (курсы и лимиты) ----

@dp.callback_query(F.data == "adm:settings")
async def cb_adm_settings(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    await state.clear()
    settings = await db_all_settings()
    lines = [
        f"• {_SETTING_LABELS.get(s['key'], s['key'])}: <b>{s['value']}</b>"
        for s in settings
    ]
    text = "<b>⚙️ Курсы и лимиты</b>\n\n" + "\n".join(lines) + "\n\nНажмите на строку, чтобы изменить значение."
    await send_or_edit(call, text, kb_settings_list(settings))
    await call.answer()


@dp.callback_query(F.data.startswith("adm:settings:edit:"))
async def cb_adm_settings_edit(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    key = call.data.split(":", 3)[3]
    current = await db_get_setting(key, "—")
    label = _SETTING_LABELS.get(key, key)
    await state.set_state(AdminStates.waiting_edit_setting_value)
    await state.update_data(edit_setting_key=key, edit_setting_label=label)
    await send_or_edit(
        call,
        f"✏️ <b>{escape(label)}</b>\n\n"
        f"Текущее значение: <b>{current}</b>\n\n"
        "Введите новое значение (число, можно дробное через точку):",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:settings")],
        ]),
    )
    await call.answer()


@dp.message(AdminStates.waiting_edit_setting_value)
async def msg_adm_edit_setting(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    raw = (message.text or "").strip().replace(",", ".")
    try:
        val = float(raw)
        if val <= 0:
            raise ValueError
    except ValueError:
        await message.answer("⚠️ Введите положительное число (например: 1.5 или 10). Попробуйте ещё раз:")
        return
    data = await state.get_data()
    key = data.get("edit_setting_key", "")
    label = data.get("edit_setting_label", key)
    await state.clear()
    # Для целочисленных настроек сохраняем без дробной части
    save_val = str(int(val)) if val == int(val) else str(val)
    await db_set_setting(key, save_val)
    await message.answer(
        f"✅ <b>{escape(label)}</b> обновлено: <b>{save_val}</b>",
        parse_mode="HTML",
        reply_markup=kb_admin_main(),
    )


@dp.callback_query(F.data.startswith("adm:users:"))
async def cb_adm_users(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    await state.clear()
    try:
        page = max(0, int(call.data.split(":", 2)[2]))
    except (ValueError, IndexError):
        page = 0
    total = await db_users_count()
    users = await db_list_users(USERS_PAGE_SIZE, page * USERS_PAGE_SIZE)
    if not users and page > 0:
        page = 0
        users = await db_list_users(USERS_PAGE_SIZE, 0)
    await state.update_data(admin_users_last_page=page)
    if total == 0:
        await send_or_edit(
            call,
            "<b>👥 Пользователей пока нет.</b>",
            kb_admin_main(),
        )
        await call.answer()
        return
    text = (
        f"<b>👥 Пользователи бота</b>\n"
        f"Всего: <b>{total}</b>\n\n"
        "Нажмите на пользователя, чтобы открыть карточку."
    )
    await send_or_edit(call, text, kb_admin_users(users, page, total))
    await call.answer()


@dp.callback_query(F.data.startswith("adm:user:"))
async def cb_adm_user_open(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    try:
        target_id = int(call.data.split(":", 2)[2])
    except (ValueError, IndexError):
        await call.answer("Некорректные данные.", show_alert=True)
        return
    data = await state.get_data()
    page = int(data.get("admin_users_last_page", 0) or 0)
    await state.clear()
    await state.update_data(admin_user_from_list_page=page)
    await _send_admin_user_card(call, target_id, state)
    await call.answer()


@dp.callback_query(F.data == "adm:find")
async def cb_adm_find(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    await state.clear()
    await state.set_state(AdminStates.waiting_user_id)
    kb_find = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:close")]
    ])
    sent = await send_or_edit(
        call, "🔍 Отправьте Telegram ID пользователя одним сообщением.", kb_find,
    )
    await state.update_data(_prompt_chat_id=sent.chat.id, _prompt_msg_id=sent.message_id)
    await call.answer()


@dp.message(AdminStates.waiting_user_id)
async def msg_adm_user_id(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    target_id = parse_positive_int(message.text)
    if target_id is None:
        await _edit_prompt(state, "⚠️ Введите корректный числовой Telegram ID.",
                           InlineKeyboardMarkup(inline_keyboard=[[
                               InlineKeyboardButton(text="❌ Отмена", callback_data="adm:close")
                           ]]))
        return
    await state.clear()
    user_row = await db_find_user(target_id)
    if not user_row:
        await message.answer(
            f"❌ Пользователь с ID <code>{target_id}</code> не найден.",
            reply_markup=kb_admin_main(),
            parse_mode="HTML",
        )
        return
    blocked = bool(user_row.get("is_blacklisted"))
    orders_cnt = await db_orders_count(target_id)
    text = (
        "<b>👤 Карточка пользователя</b>\n\n"
        f"Telegram ID: <code>{user_row['tg_id']}</code>\n"
        f"Username: @{escape(user_row.get('username') or '—')}\n"
        f"Имя: {escape(user_row.get('first_name') or '—')}\n"
        f"Баланс: <b>{user_row['balance']}₽</b>\n"
        f"Заказов: <b>{orders_cnt}</b>\n"
        f"Чёрный список: {'<b>да</b>' if blocked else 'нет'}"
    )
    await message.answer(
        text,
        reply_markup=kb_admin_user(target_id, blocked, from_list_page=None),
        parse_mode="HTML",
    )


@dp.callback_query(F.data.startswith("adm:credit:"))
async def cb_adm_credit(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    target_id = int(call.data.split(":")[2])
    prev = await state.get_data()
    saved_source = prev.get("admin_user_from_list_page")
    await state.set_state(AdminStates.waiting_credit_amount)
    kb_cancel = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:close")]
    ])
    sent = await send_or_edit(
        call,
        f"➕ Сколько рублей начислить пользователю <code>{target_id}</code>?\n"
        "Отправьте число одним сообщением.",
        kb_cancel,
    )
    await state.update_data(
        adm_target_id=target_id, admin_user_from_list_page=saved_source,
        _prompt_chat_id=sent.chat.id, _prompt_msg_id=sent.message_id,
    )
    await call.answer()


@dp.message(AdminStates.waiting_credit_amount)
async def msg_adm_credit(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await _try_delete(message)
    amount = parse_positive_int(message.text)
    if amount is None:
        await _edit_prompt(state, "⚠️ Введите положительное число рублей.",
                           InlineKeyboardMarkup(inline_keyboard=[[
                               InlineKeyboardButton(text="❌ Отмена", callback_data="adm:close")
                           ]]))
        return
    data = await state.get_data()
    target_id = int(data.get("adm_target_id", 0))
    saved_source = data.get("admin_user_from_list_page")
    await state.clear()
    if saved_source is not None:
        await state.update_data(admin_user_from_list_page=saved_source)
    if not target_id:
        await message.answer("Не указан пользователь.", reply_markup=kb_admin_main())
        return
    new_balance = await db_add_balance(target_id, amount)
    await db_add_transaction(
        target_id, amount, kind="admin_add", reason="Начисление администратором"
    )
    try:
        await bot.send_message(
            target_id,
            f"✅ <b>Администратор начислил вам {_fmt_price(amount)}₽</b>\n"
            f"Текущий баланс: <b>{_fmt_price(new_balance)}₽</b>",
            parse_mode="HTML",
        )
    except Exception:
        pass
    from_list_page = saved_source if saved_source is not None else None
    from_list_page = int(from_list_page) if from_list_page is not None else None
    await message.answer(
        f"✅ Начислено {_fmt_price(amount)}₽. Новый баланс: <b>{_fmt_price(new_balance)}₽</b>",
        parse_mode="HTML",
        reply_markup=kb_admin_user(
            target_id,
            await db_is_blacklisted(target_id),
            from_list_page=from_list_page,
        ),
    )


@dp.callback_query(F.data.startswith("adm:reset:"))
async def cb_adm_reset(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    target_id = int(call.data.split(":")[2])
    old_balance = await db_get_balance(target_id)
    await db_set_balance(target_id, 0)
    if old_balance:
        await db_add_transaction(
            target_id,
            -old_balance,
            kind="admin_reset",
            reason="Обнуление администратором",
        )
    try:
        await bot.send_message(
            target_id,
            "♻️ <b>Ваш баланс был обнулён администратором.</b>",
            parse_mode="HTML",
        )
    except Exception:
        pass
    await _send_admin_user_card(call, target_id, state)
    await call.answer("Баланс обнулён")


@dp.callback_query(F.data.startswith("adm:block:"))
async def cb_adm_block(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    target_id = int(call.data.split(":")[2])
    await db_set_blacklist(target_id, True)
    try:
        await bot.send_message(
            target_id,
            "🚫 <b>Вы добавлены в чёрный список магазина.</b>\n"
            "Доступ к боту ограничен.",
            parse_mode="HTML",
        )
    except Exception:
        pass
    await _send_admin_user_card(call, target_id, state)
    await call.answer("В чёрном списке")


@dp.callback_query(F.data.startswith("adm:unblock:"))
async def cb_adm_unblock(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    target_id = int(call.data.split(":")[2])
    await db_set_blacklist(target_id, False)
    try:
        await bot.send_message(
            target_id,
            "✅ <b>Вы исключены из чёрного списка.</b> Доступ к боту восстановлен.",
            parse_mode="HTML",
        )
    except Exception:
        pass
    await _send_admin_user_card(call, target_id, state)
    await call.answer("Убран из ЧС")


@dp.callback_query(F.data.startswith("adm:tx:"))
async def cb_adm_transactions(call: CallbackQuery) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        await call.answer("Только для модератора.", show_alert=True)
        return
    try:
        target_id = int(call.data.split(":")[2])
    except (ValueError, IndexError):
        await call.answer("Некорректные данные.", show_alert=True)
        return

    txs = await db_get_transactions(target_id, limit=30)
    if not txs:
        text = (
            f"📜 <b>История транзакций</b>\n"
            f"Пользователь: <code>{target_id}</code>\n\n"
            "Операций нет."
        )
    else:
        lines = [
            f"📜 <b>История транзакций</b>",
            f"Пользователь: <code>{target_id}</code>\n",
        ]
        for t in txs:
            date = _fmt_msk(t["created_at"])
            sign = "➕" if t["amount"] > 0 else "➖"
            lines.append(
                f"{sign} <b>{abs(t['amount'])}₽</b> — "
                f"{escape(_format_tx_kind(t['kind']))}\n"
                + (f"<i>{escape(t['reason'])}</i>\n" if t["reason"] else "")
                + f"<i>{date}</i>"
            )
        text = "\n".join(lines)
        if len(text) > 3800:
            text = text[:3800] + "\n…"

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="⬅️ К пользователю",
                    callback_data=f"adm:back:{target_id}",
                )
            ],
            [InlineKeyboardButton(text="❌ Закрыть", callback_data="adm:close")],
        ]
    )
    await send_or_edit(call, text, kb)


@dp.callback_query(F.data.startswith("adm:back:"))
async def cb_adm_back(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    try:
        target_id = int(call.data.split(":")[2])
    except (ValueError, IndexError):
        return
    await _send_admin_user_card(call, target_id, state)


# =====================================================================
# Реферальные статусы и промокоды — Административная панель
# =====================================================================


@dp.callback_query(F.data.startswith("adm:ref_level:"))
async def cb_adm_ref_level(call: CallbackQuery) -> None:
    """Показывает список уровней для выбора."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    try:
        target_id = int(call.data.split(":")[2])
    except (ValueError, IndexError):
        return

    user_row = await db_find_user(target_id)
    if not user_row:
        await call.answer("Пользователь не найден.", show_alert=True)
        return

    current_level = user_row.get("referral_level", 0) or 0
    current_name = _ref_level_name(current_level)

    text = (
        f"<b>👑 Выдача реферального статуса</b>\n\n"
        f"Пользователь: <code>{target_id}</code>\n"
        f"Текущий статус: <b>{current_name}</b> (ур. {current_level})\n\n"
        "Выберите новый уровень.\n"
        "<i>⚠️ При повышении уровня пользователь получит уведомление "
        "и промокоды на скидку. При понижении промокоды не отзываются.</i>"
    )

    rows = []
    for lvl, (name, disc) in REFERRAL_LEVELS.items():
        marker = " ✓" if lvl == current_level else ""
        rows.append([InlineKeyboardButton(
            text=f"{name} — {disc}%{marker}",
            callback_data=f"adm:set_ref_level:{target_id}:{lvl}",
        )])
    rows.append([InlineKeyboardButton(
        text="⬅️ К пользователю", callback_data=f"adm:back:{target_id}"
    )])
    kb = InlineKeyboardMarkup(inline_keyboard=rows)
    await send_or_edit(call, text, kb)


@dp.callback_query(F.data.startswith("adm:set_ref_level:"))
async def cb_adm_set_ref_level(call: CallbackQuery, state: FSMContext) -> None:
    """Устанавливает реферальный уровень пользователю."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    try:
        parts = call.data.split(":")
        target_id = int(parts[2])
        new_level = int(parts[3])
    except (ValueError, IndexError):
        return

    if new_level not in REFERRAL_LEVELS:
        await call.answer("Некорректный уровень.", show_alert=True)
        return

    old_level = await db_admin_set_ref_level(target_id, new_level)
    level_name = _ref_level_name(new_level)
    discount = _ref_discount_pct(new_level)

    if new_level > old_level:
        try:
            await bot.send_message(
                target_id,
                f"🎉 <b>Вам выдан новый реферальный статус!</b>\n\n"
                f"🏆 Ваш статус: <b>{level_name}</b>\n"
                f"💰 Скидка на все заказы: <b>{discount}%</b>\n\n"
                f"Скидка применяется автоматически при каждой покупке.\n"
                f"<i>Не действует на Telegram Stars.</i>",
                parse_mode="HTML",
            )
        except Exception as e:
            logging.warning(f"Не удалось уведомить {target_id} о новом уровне: {e}")
    elif new_level < old_level:
        try:
            await bot.send_message(
                target_id,
                f"ℹ️ Ваш реферальный статус изменён.\n\n"
                f"Новый статус: <b>{level_name}</b>\n"
                f"💰 Скидка на все заказы: <b>{discount}%</b>",
                parse_mode="HTML",
            )
        except Exception:
            pass

    action = "повышен" if new_level > old_level else ("понижен" if new_level < old_level else "не изменён")
    text = (
        f"✅ <b>Статус {action}.</b>\n\n"
        f"Пользователь: <code>{target_id}</code>\n"
        f"Новый статус: <b>{level_name}</b> (ур. {new_level})"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🎟️ Реф. промокоды", callback_data=f"adm:ref_promos:{target_id}")],
        [InlineKeyboardButton(text="⬅️ К пользователю", callback_data=f"adm:back:{target_id}")],
    ])
    await send_or_edit(call, text, kb)


@dp.callback_query(F.data.startswith("adm:ref_promos:"))
async def cb_adm_ref_promos(call: CallbackQuery) -> None:
    """Просмотр реферальных промокодов пользователя."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    try:
        target_id = int(call.data.split(":")[2])
    except (ValueError, IndexError):
        return

    promos = await db_admin_get_user_ref_promos(target_id)

    if not promos:
        text = (
            f"<b>🎟️ Реферальные промокоды</b>\n"
            f"Пользователь: <code>{target_id}</code>\n\n"
            "Промокодов нет."
        )
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ К пользователю", callback_data=f"adm:back:{target_id}")],
        ])
        await send_or_edit(call, text, kb)
        return

    lines = [f"<b>🎟️ Реферальные промокоды</b>\nПользователь: <code>{target_id}</code>\n"]
    rows = []
    active_count = 0
    for p in promos:
        status = "✅ активен" if not p["used_at"] and p["is_active"] else "🔴 использован/отозван"
        lines.append(f"• <code>{p['code']}</code> — скидка {p['discount_pct']}% — {status}")
        if not p["used_at"]:
            active_count += 1
            rows.append([InlineKeyboardButton(
                text=f"❌ Удалить {p['code']}",
                callback_data=f"adm:del_ref_promo:{target_id}:{p['promo_id']}",
            )])

    if active_count > 0:
        rows.append([InlineKeyboardButton(
            text=f"🗑️ Удалить все активные ({active_count})",
            callback_data=f"adm:del_all_ref_promos:{target_id}",
        )])
    rows.append([InlineKeyboardButton(
        text="⬅️ К пользователю", callback_data=f"adm:back:{target_id}"
    )])
    text = "\n".join(lines)
    if len(text) > 3800:
        text = text[:3800] + "\n…"
    kb = InlineKeyboardMarkup(inline_keyboard=rows)
    await send_or_edit(call, text, kb)


@dp.callback_query(F.data.startswith("adm:del_ref_promo:"))
async def cb_adm_del_ref_promo(call: CallbackQuery) -> None:
    """Удаляет реферальный промокод из инвентаря пользователя."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    try:
        parts = call.data.split(":")
        target_id = int(parts[2])
        promo_id = int(parts[3])
    except (ValueError, IndexError):
        return

    ok = await db_admin_remove_user_promo(target_id, promo_id)
    if ok:
        await call.answer("Промокод удалён.", show_alert=False)
    else:
        await call.answer("Промокод не найден или уже удалён.", show_alert=True)

    # Обновляем список
    promos = await db_admin_get_user_ref_promos(target_id)
    if not promos:
        text = (
            f"<b>🎟️ Реферальные промокоды</b>\n"
            f"Пользователь: <code>{target_id}</code>\n\n"
            "Промокодов нет."
        )
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ К пользователю", callback_data=f"adm:back:{target_id}")],
        ])
    else:
        lines = [f"<b>🎟️ Реферальные промокоды</b>\nПользователь: <code>{target_id}</code>\n"]
        rows = []
        active_count = 0
        for p in promos:
            status = "✅ активен" if not p["used_at"] and p["is_active"] else "🔴 использован/отозван"
            lines.append(f"• <code>{p['code']}</code> — скидка {p['discount_pct']}% — {status}")
            if not p["used_at"]:
                active_count += 1
                rows.append([InlineKeyboardButton(
                    text=f"❌ Удалить {p['code']}",
                    callback_data=f"adm:del_ref_promo:{target_id}:{p['promo_id']}",
                )])
        if active_count > 0:
            rows.append([InlineKeyboardButton(
                text=f"🗑️ Удалить все активные ({active_count})",
                callback_data=f"adm:del_all_ref_promos:{target_id}",
            )])
        rows.append([InlineKeyboardButton(
            text="⬅️ К пользователю", callback_data=f"adm:back:{target_id}"
        )])
        text = "\n".join(lines)
        kb = InlineKeyboardMarkup(inline_keyboard=rows)

    await send_or_edit(call, text, kb)


@dp.callback_query(F.data.startswith("adm:del_all_ref_promos:"))
async def cb_adm_del_all_ref_promos(call: CallbackQuery) -> None:
    """Модератор удаляет все активные реферальные промокоды пользователя."""
    if not _is_moderator(call.from_user.id):
        await call.answer()
        return
    try:
        target_id = int(call.data.split(":")[2])
    except (ValueError, IndexError):
        await call.answer()
        return

    deleted = await db_delete_all_user_ref_promos(target_id)
    await call.answer(f"Удалено: {deleted} промокодов", show_alert=True)

    text = (
        f"<b>🎟️ Реферальные промокоды</b>\n"
        f"Пользователь: <code>{target_id}</code>\n\n"
        f"✅ Удалено активных промокодов: <b>{deleted}</b>\n\n"
        "Промокодов нет."
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⬅️ К пользователю", callback_data=f"adm:back:{target_id}")],
    ])
    await send_or_edit(call, text, kb)


# =====================================================================
# Промокоды — вспомогательные функции
# =====================================================================


def _promo_is_valid_now(promo: dict) -> bool:
    """Проверяет, активен ли промокод в данный момент."""
    if not promo.get("is_active"):
        return False
    now = datetime.now(timezone.utc)
    starts = promo.get("starts_at")
    expires = promo.get("expires_at")
    if starts:
        try:
            if now < datetime.fromisoformat(starts):
                return False
        except ValueError:
            pass
    if expires:
        try:
            if now > datetime.fromisoformat(expires):
                return False
        except ValueError:
            pass
    return True


def _fmt_promo_dates(promo: dict) -> str:
    starts = promo.get("starts_at")
    expires = promo.get("expires_at")
    if not starts and not expires:
        return "Бессрочно"
    parts = []
    if starts:
        try:
            dt = datetime.fromisoformat(starts)
            parts.append(f"с {dt.strftime('%d.%m.%Y')}")
        except ValueError:
            parts.append(f"с {starts}")
    if expires:
        try:
            dt = datetime.fromisoformat(expires)
            parts.append(f"по {dt.strftime('%d.%m.%Y')}")
        except ValueError:
            parts.append(f"по {expires}")
    return " ".join(parts)


def _parse_date_iso(date_str: str) -> str | None:
    try:
        d = datetime.strptime(date_str.strip(), "%d.%m.%Y")
        return d.replace(tzinfo=timezone.utc).isoformat()
    except ValueError:
        return None


# =====================================================================
# Промокоды — Административная панель
# =====================================================================


def kb_admin_promos(promos: list[dict]) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for p in promos:
        icon = "✅" if p["is_active"] else "🔴"
        if p.get("discount_pct", 0) > 0 and p.get("promo_price", 0) == 0:
            label = f"{icon} {p['code']} — скидка {p['discount_pct']}% ({p['product_title']})"
        else:
            label = f"{icon} {p['code']} — {p['product_title']} ({p['promo_price']}₽)"
        rows.append([InlineKeyboardButton(text=label, callback_data=f"adm:promo:{p['id']}")])
    rows.append([InlineKeyboardButton(text="➕ Создать промокод", callback_data="adm:promo_create")])
    if promos:
        rows.append([InlineKeyboardButton(text="🗑️ Удалить все промокоды", callback_data="adm:del_all_promos")])
    rows.append([InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adm:panel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def kb_promo_game_select() -> InlineKeyboardMarkup:
    # Захардкоженные категории (ключи)
    _HARDCODED_KEYS = {"roblox_instant", "roblox_gamepass", "brawl", "tgstars", "other"}
    rows = [
        [InlineKeyboardButton(text="🟦 Roblox", callback_data="adm:promo_game:Roblox")],
        [InlineKeyboardButton(text="⭐ Brawl Stars", callback_data="adm:promo_game:Brawl Stars")],
        [InlineKeyboardButton(text="✨ Telegram Stars", callback_data="adm:promo_game:Telegram Stars")],
        [InlineKeyboardButton(text="📦 Другое", callback_data="adm:promo_game:Другое")],
    ]
    # Добавляем категории, созданные через админку (не зарезервированные, не отключённые)
    try:
        cats = await db_all_categories()
        for cat in cats:
            if cat["key"] not in _HARDCODED_KEYS and not cat.get("disabled"):
                label = f"{cat['emoji']} {cat['name']}"
                rows.append([InlineKeyboardButton(
                    text=label,
                    callback_data=f"adm:promo_game:{cat['name']}",
                )])
    except Exception:
        pass
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@dp.callback_query(F.data == "adm:panel")
async def cb_adm_panel(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    await state.clear()
    await send_or_edit(
        call,
        "<b>🛠️ Админ-панель</b>\n\nВыберите действие:",
        kb_admin_main(),
    )


@dp.callback_query(F.data == "adm:promos")
async def cb_adm_promos(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    await state.clear()
    promos = await db_list_regular_promos()
    text = "<b>🎟️ Промокоды</b>\n\n"
    if promos:
        text += f"Всего промокодов: <b>{len(promos)}</b>\n\nНажмите на промокод для управления."
    else:
        text += "Промокодов пока нет. Создайте первый!"
    await send_or_edit(call, text, kb_admin_promos(promos))


@dp.callback_query(F.data == "adm:del_all_promos")
async def cb_adm_del_all_promos(call: CallbackQuery) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    deleted = await db_delete_all_regular_promos()
    await call.answer(f"Удалено промокодов: {deleted}", show_alert=True)
    text = (
        "<b>🎟️ Промокоды</b>\n\n"
        f"✅ Удалено промокодов: <b>{deleted}</b>\n\n"
        "Промокодов пока нет. Создайте первый!"
    )
    await send_or_edit(call, text, kb_admin_promos([]))


@dp.callback_query(F.data == "adm:promo_create")
async def cb_adm_promo_create(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    await state.set_state(PromoStates.waiting_code_input)
    await send_or_edit(
        call,
        "📝 <b>Создание промокода</b>\n\n"
        "Введите <b>код промокода</b> (только латинские буквы, цифры и знак «_», без пробелов).\n\n"
        "Например: <code>SALE20</code>",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")]
        ])
    )


@dp.message(PromoStates.waiting_code_input)
async def msg_promo_code_input(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    code = (message.text or "").strip().upper()
    if not code or not code.replace("_", "").isalnum():
        await message.answer(
            "⚠️ Код может содержать только латинские буквы, цифры и «_». Попробуйте ещё раз.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")]
            ])
        )
        return
    existing = await db_get_promo_by_code(code)
    if existing:
        await message.answer(
            f"⚠️ Промокод <code>{escape(code)}</code> уже существует. Введите другой.",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")]
            ])
        )
        return
    await state.update_data(promo_code=code)
    await state.set_state(PromoStates.waiting_discount_type)
    await message.answer(
        f"Код: <code>{escape(code)}</code>\n\n"
        "📋 Выберите <b>тип промокода</b>:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🏷️ Фиксированная цена", callback_data="adm:promo_type:price")],
            [InlineKeyboardButton(text="💸 Скидка %", callback_data="adm:promo_type:pct")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")],
        ])
    )


@dp.callback_query(F.data.startswith("adm:promo_type:"))
async def cb_adm_promo_type(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    data = await state.get_data()
    if not data.get("promo_code"):
        return
    promo_type = call.data.split(":")[2]
    code = data.get("promo_code", "?")

    if promo_type == "pct":
        await state.update_data(promo_game="pct_discount")
        await state.set_state(PromoStates.waiting_product)
        await send_or_edit(
            call,
            f"Код: <code>{escape(code)}</code>  |  Тип: <b>Скидка %</b>\n\n"
            "📦 Введите <b>описание промокода</b>\n"
            "Например: <code>Скидка на любой товар</code>",
            InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")]
            ])
        )
    else:
        await state.set_state(None)
        await send_or_edit(
            call,
            f"Код: <code>{escape(code)}</code>\n\n"
            "🎮 Выберите <b>игру</b>, для которой действует промокод:",
            await kb_promo_game_select()
        )


@dp.callback_query(F.data.startswith("adm:promo_game:"))
async def cb_adm_promo_game(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    game = call.data.split(":", 2)[2]
    await state.update_data(promo_game=game)
    await state.set_state(PromoStates.waiting_product)
    data = await state.get_data()
    code = data.get("promo_code", "?")
    await send_or_edit(
        call,
        f"Код: <code>{escape(code)}</code>  |  Игра: <b>{escape(game)}</b>\n\n"
        "📦 Введите <b>название товара</b>\n"
        "Например: <code>800 Robux</code> или <code>Brawl Pass Plus</code>",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")]
        ])
    )


@dp.message(PromoStates.waiting_product)
async def msg_promo_product(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    product = (message.text or "").strip()
    if not product:
        await message.answer("⚠️ Название товара не может быть пустым. Введите ещё раз.")
        return
    await state.update_data(promo_product=product)
    data = await state.get_data()
    code = data.get("promo_code", "?")
    game = data.get("promo_game", "?")

    if game == "pct_discount":
        await state.set_state(PromoStates.waiting_discount_pct)
        await message.answer(
            f"Код: <code>{escape(code)}</code>  |  Тип: <b>Скидка %</b>\n"
            f"Описание: <b>{escape(product)}</b>\n\n"
            "💸 Введите <b>размер скидки в процентах</b> (целое число, например: <code>10</code>):",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")]
            ])
        )
    else:
        await state.set_state(PromoStates.waiting_price)
        await message.answer(
            f"Код: <code>{escape(code)}</code>  |  Игра: <b>{escape(game)}</b>\n"
            f"Товар: <b>{escape(product)}</b>\n\n"
            "💰 Введите <b>цену по промокоду</b> (целое число в рублях, например: <code>150</code>):",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")]
            ])
        )


@dp.message(PromoStates.waiting_discount_pct)
async def msg_promo_discount_pct(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    pct = parse_positive_int(message.text)
    if pct is None or pct > 99:
        await message.answer("⚠️ Введите корректное значение от 1 до 99.")
        return
    await state.update_data(promo_discount_pct=pct)
    await state.set_state(PromoStates.waiting_dates)
    await message.answer(
        f"Скидка: <b>{pct}%</b>\n\n"
        "📅 Введите <b>срок действия</b> промокода:\n\n"
        "• Дата окончания: <code>31.12.2026</code>\n"
        "• Диапазон дат: <code>01.07.2026 - 31.12.2026</code>\n"
        "• Или нажмите кнопку ниже, если промокод бессрочный:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="♾️ Бессрочно", callback_data="adm:promo_dates:forever")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")],
        ])
    )


@dp.message(PromoStates.waiting_price)
async def msg_promo_price_input(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    price = parse_positive_int(message.text)
    if price is None:
        await message.answer("⚠️ Введите корректную сумму (целое число больше 0).")
        return
    await state.update_data(promo_price=price)
    await state.set_state(PromoStates.waiting_dates)
    await message.answer(
        "📅 Введите <b>срок действия</b> промокода:\n\n"
        "• Дата окончания: <code>31.12.2026</code>\n"
        "• Диапазон дат: <code>01.07.2026 - 31.12.2026</code>\n"
        "• Или нажмите кнопку ниже, если промокод бессрочный:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="♾️ Бессрочно", callback_data="adm:promo_dates:forever")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")],
        ])
    )


@dp.callback_query(F.data == "adm:promo_dates:forever")
async def cb_adm_promo_dates_forever(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    await state.update_data(promo_starts_at=None, promo_expires_at=None)
    await state.set_state(PromoStates.waiting_max_uses)
    data = await state.get_data()
    code = data.get("promo_code", "?")
    await send_or_edit(
        call,
        f"Код: <code>{escape(code)}</code>  |  Срок: ♾️ Бессрочно\n\n"
        "🔢 Введите <b>лимит активаций</b> (сколько раз можно использовать этот промокод).\n\n"
        "Или нажмите кнопку, если промокод без ограничений:",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="♾️ Без ограничений", callback_data="adm:promo_maxuses:unlimited")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")],
        ]),
    )


@dp.message(PromoStates.waiting_dates)
async def msg_promo_dates(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    text = (message.text or "").strip()
    starts_at = None
    expires_at = None
    separators = [" - ", "–", "—", " — "]
    sep_found = next((s for s in separators if s in text), None)
    if sep_found:
        parts = text.split(sep_found, 1)
        starts_at = _parse_date_iso(parts[0])
        expires_at = _parse_date_iso(parts[1])
        if not starts_at or not expires_at:
            await message.answer(
                "⚠️ Не удалось распознать даты. Используйте формат:\n"
                "<code>01.07.2026 - 31.12.2026</code>",
                parse_mode="HTML"
            )
            return
    else:
        expires_at = _parse_date_iso(text)
        if not expires_at:
            await message.answer(
                "⚠️ Не удалось распознать дату. Введите в формате <code>31.12.2026</code>.",
                parse_mode="HTML"
            )
            return
    data = await state.get_data()
    await state.update_data(promo_starts_at=starts_at, promo_expires_at=expires_at)
    await state.set_state(PromoStates.waiting_max_uses)
    code = data.get("promo_code", "?")
    await message.answer(
        f"Код: <code>{escape(code)}</code>  |  Срок установлен.\n\n"
        "🔢 Введите <b>лимит активаций</b> (сколько раз можно использовать этот промокод).\n\n"
        "Или нажмите кнопку, если промокод без ограничений:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="♾️ Без ограничений", callback_data="adm:promo_maxuses:unlimited")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")],
        ]),
    )


async def _save_promo_and_confirm(
    call: CallbackQuery,
    data: dict,
    starts_at: str | None,
    expires_at: str | None,
    max_uses: int | None = None,
) -> None:
    code = data.get("promo_code")
    game = data.get("promo_game")
    product = data.get("promo_product")
    price = data.get("promo_price")
    discount_pct = data.get("promo_discount_pct", 0)
    is_pct = game == "pct_discount"
    if not all([code, game, product]) or (not is_pct and not price):
        await send_or_edit(call, "❌ Ошибка: данные утеряны. Начните заново.", kb_admin_promos([]))
        return
    promo_id = await db_create_promo(
        code, game, product,
        0 if is_pct else int(price),
        starts_at, expires_at,
        discount_pct=int(discount_pct) if is_pct else 0,
        max_uses=max_uses,
    )
    dates_str = _fmt_promo_dates({"starts_at": starts_at, "expires_at": expires_at})
    uses_str = f"<b>{max_uses}</b>" if max_uses is not None else "♾️ без ограничений"
    if is_pct:
        detail_line = f"💸 Скидка: <b>{discount_pct}%</b>\n"
    else:
        detail_line = f"💰 Цена: <b>{price}₽</b>\n"
    text = (
        "✅ <b>Промокод создан!</b>\n\n"
        f"🎟️ Код: <code>{escape(code)}</code>\n"
        f"🎮 Тип: {escape(game)}\n"
        f"📦 Описание: {escape(product)}\n"
        f"{detail_line}"
        f"📅 Срок: {dates_str}\n"
        f"🔢 Лимит активаций: {uses_str}\n"
        f"🆔 ID: {promo_id}"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🎟️ Все промокоды", callback_data="adm:promos")],
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adm:panel")],
    ])
    await send_or_edit(call, text, kb)


async def _save_promo_and_confirm_msg(
    message: Message,
    data: dict,
    starts_at: str | None,
    expires_at: str | None,
    max_uses: int | None = None,
) -> None:
    code = data.get("promo_code")
    game = data.get("promo_game")
    product = data.get("promo_product")
    price = data.get("promo_price")
    discount_pct = data.get("promo_discount_pct", 0)
    is_pct = game == "pct_discount"
    if not all([code, game, product]) or (not is_pct and not price):
        await message.answer("❌ Ошибка: данные утеряны. Начните заново.")
        return
    promo_id = await db_create_promo(
        code, game, product,
        0 if is_pct else int(price),
        starts_at, expires_at,
        discount_pct=int(discount_pct) if is_pct else 0,
        max_uses=max_uses,
    )
    dates_str = _fmt_promo_dates({"starts_at": starts_at, "expires_at": expires_at})
    uses_str = f"<b>{max_uses}</b>" if max_uses is not None else "♾️ без ограничений"
    if is_pct:
        detail_line = f"💸 Скидка: <b>{discount_pct}%</b>\n"
    else:
        detail_line = f"💰 Цена: <b>{price}₽</b>\n"
    text = (
        "✅ <b>Промокод создан!</b>\n\n"
        f"🎟️ Код: <code>{escape(code)}</code>\n"
        f"🎮 Тип: {escape(game)}\n"
        f"📦 Описание: {escape(product)}\n"
        f"{detail_line}"
        f"📅 Срок: {dates_str}\n"
        f"🔢 Лимит активаций: {uses_str}\n"
        f"🆔 ID: {promo_id}"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🎟️ Все промокоды", callback_data="adm:promos")],
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adm:panel")],
    ])
    await message.answer(text, reply_markup=kb, parse_mode="HTML")


@dp.callback_query(F.data == "adm:promo_maxuses:unlimited")
async def cb_adm_promo_maxuses_unlimited(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    data = await state.get_data()
    await state.clear()
    starts_at = data.get("promo_starts_at")
    expires_at = data.get("promo_expires_at")
    await _save_promo_and_confirm(call, data, starts_at=starts_at, expires_at=expires_at, max_uses=None)


@dp.message(PromoStates.waiting_max_uses)
async def msg_promo_max_uses(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    val = parse_positive_int(message.text)
    if val is None or val <= 0:
        await message.answer(
            "⚠️ Введите положительное целое число (например: <code>100</code>) "
            "или нажмите кнопку «Без ограничений».",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="♾️ Без ограничений", callback_data="adm:promo_maxuses:unlimited")],
                [InlineKeyboardButton(text="❌ Отмена", callback_data="adm:promos")],
            ]),
        )
        return
    data = await state.get_data()
    await state.clear()
    starts_at = data.get("promo_starts_at")
    expires_at = data.get("promo_expires_at")
    await _save_promo_and_confirm_msg(message, data, starts_at=starts_at, expires_at=expires_at, max_uses=val)


@dp.callback_query(F.data.startswith("adm:promo:"))
async def cb_adm_promo_detail(call: CallbackQuery) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    parts = call.data.split(":")
    if len(parts) < 3:
        return
    try:
        promo_id = int(parts[2])
    except ValueError:
        return
    promo = await db_get_promo_by_id(promo_id)
    if not promo:
        await call.answer("Промокод не найден.", show_alert=True)
        return
    usage_cnt = await db_promo_usage_count(promo_id)
    status_text = "✅ Активен" if promo["is_active"] else "🔴 Отключён"
    toggle_text = "🔴 Отключить" if promo["is_active"] else "✅ Включить"
    dates_str = _fmt_promo_dates(promo)
    valid_now = _promo_is_valid_now(promo)
    max_uses = promo.get("max_uses")
    if max_uses is not None:
        uses_line = f"🔢 Использовано: <b>{usage_cnt} / {max_uses}</b>"
    else:
        uses_line = f"🔢 Использовано: <b>{usage_cnt}</b>  (♾️ без лимита)"
    is_pct = promo.get("game") == "pct_discount"
    price_line = (
        f"💸 Скидка: <b>{promo['discount_pct']}%</b>"
        if is_pct and promo.get("discount_pct", 0) > 0
        else f"💰 Цена: <b>{promo['promo_price']}₽</b>"
    )
    text = (
        f"<b>🎟️ Промокод: <code>{escape(promo['code'])}</code></b>\n\n"
        f"🎮 Игра: {escape(promo['game'])}\n"
        f"📦 Товар: {escape(promo['product_title'])}\n"
        f"{price_line}\n"
        f"📅 Срок: {dates_str}\n"
        f"📊 Статус: {status_text}  |  Сейчас: {'🟢 работает' if valid_now else '🔴 не работает'}\n"
        f"{uses_line}"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=toggle_text, callback_data=f"adm:promo_toggle:{promo_id}")],
        [InlineKeyboardButton(text="🗑️ Удалить", callback_data=f"adm:promo_del:{promo_id}")],
        [InlineKeyboardButton(text="⬅️ К промокодам", callback_data="adm:promos")],
    ])
    await send_or_edit(call, text, kb)


@dp.callback_query(F.data.startswith("adm:promo_toggle:"))
async def cb_adm_promo_toggle(call: CallbackQuery) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    try:
        promo_id = int(call.data.split(":")[2])
    except (ValueError, IndexError):
        return
    promo = await db_get_promo_by_id(promo_id)
    if not promo:
        await call.answer("Промокод не найден.", show_alert=True)
        return
    new_state = not bool(promo["is_active"])
    await db_toggle_promo(promo_id, new_state)
    status_word = "включён" if new_state else "отключён"
    await call.answer(f"Промокод {status_word}.", show_alert=True)
    promo["is_active"] = int(new_state)
    usage_cnt = await db_promo_usage_count(promo_id)
    toggle_text = "🔴 Отключить" if new_state else "✅ Включить"
    dates_str = _fmt_promo_dates(promo)
    valid_now = _promo_is_valid_now(promo)
    status_text = "✅ Активен" if new_state else "🔴 Отключён"
    text = (
        f"<b>🎟️ Промокод: <code>{escape(promo['code'])}</code></b>\n\n"
        f"🎮 Игра: {escape(promo['game'])}\n"
        f"📦 Товар: {escape(promo['product_title'])}\n"
        f"💰 Цена: <b>{promo['promo_price']}₽</b>\n"
        f"📅 Срок: {dates_str}\n"
        f"📊 Статус: {status_text}  |  Сейчас: {'🟢 работает' if valid_now else '🔴 не работает'}\n"
        f"🛒 Использовано: <b>{usage_cnt}</b> раз"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=toggle_text, callback_data=f"adm:promo_toggle:{promo_id}")],
        [InlineKeyboardButton(text="🗑️ Удалить", callback_data=f"adm:promo_del:{promo_id}")],
        [InlineKeyboardButton(text="⬅️ К промокодам", callback_data="adm:promos")],
    ])
    await send_or_edit(call, text, kb)


@dp.callback_query(F.data.startswith("adm:promo_del:"))
async def cb_adm_promo_del(call: CallbackQuery) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    try:
        promo_id = int(call.data.split(":")[2])
    except (ValueError, IndexError):
        return
    promo = await db_get_promo_by_id(promo_id)
    if not promo:
        await call.answer("Промокод не найден.", show_alert=True)
        return
    text = (
        f"❓ Удалить промокод <code>{escape(promo['code'])}</code>?\n\n"
        "Все данные об использовании тоже будут удалены."
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🗑️ Да, удалить", callback_data=f"adm:promo_del_yes:{promo_id}")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data=f"adm:promo:{promo_id}")],
    ])
    await send_or_edit(call, text, kb)


@dp.callback_query(F.data.startswith("adm:promo_del_yes:"))
async def cb_adm_promo_del_yes(call: CallbackQuery) -> None:
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    try:
        promo_id = int(call.data.split(":")[2])
    except (ValueError, IndexError):
        return
    await db_delete_promo(promo_id)
    await call.answer("Промокод удалён.", show_alert=True)
    promos = await db_list_regular_promos()
    text = "<b>🎟️ Промокоды</b>\n\n"
    text += f"Всего промокодов: <b>{len(promos)}</b>\n\nНажмите на промокод для управления." if promos else "Промокодов пока нет."
    await send_or_edit(call, text, kb_admin_promos(promos))


# =====================================================================
# Промокоды — Покупатель
# =====================================================================


@dp.callback_query(F.data == "enter_promo")
async def cb_enter_promo(call: CallbackQuery, state: FSMContext) -> None:
    await call.answer()
    await state.set_state(ShopStates.waiting_promo_input)
    await send_or_edit(
        call,
        "✏️ <b>Ввести промокод</b>\n\n"
        "Отправьте код промокода одним сообщением.\n"
        "Промокод появится в разделе «Мои промокоды».",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Профиль", callback_data="profile")]
        ])
    )


@dp.message(ShopStates.waiting_promo_input)
async def msg_promo_input(message: Message, state: FSMContext) -> None:
    code = (message.text or "").strip()
    if not code:
        await message.answer("Введите код промокода.")
        return

    promo = await db_get_promo_by_code(code)
    kb_back = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⬅️ Профиль", callback_data="profile")]
    ])

    if not promo:
        await message.answer(
            "❌ Промокод не найден. Проверьте правильность написания.",
            reply_markup=kb_back
        )
        return

    if not _promo_is_valid_now(promo):
        await message.answer(
            "❌ Этот промокод недействителен или истёк.",
            reply_markup=kb_back
        )
        return

    claimed = await db_claim_promo(message.from_user.id, promo["id"])
    await state.clear()

    if not claimed:
        await message.answer(
            f"ℹ️ Промокод <code>{escape(promo['code'])}</code> уже есть в вашем профиле.",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🎟️ Мои промокоды", callback_data="my_promos")],
                [InlineKeyboardButton(text="⬅️ Профиль", callback_data="profile")],
            ])
        )
        return

    await message.answer(
        f"✅ Промокод <code>{escape(promo['code'])}</code> добавлен!\n\n"
        f"🎮 Игра: {escape(promo['game'])}\n"
        f"📦 Товар: {escape(promo['product_title'])}\n"
        f"💰 Цена по промокоду: <b>{promo['promo_price']}₽</b>\n"
        f"📅 Срок: {_fmt_promo_dates(promo)}\n\n"
        "Найдите его в разделе <b>«Мои промокоды»</b> и воспользуйтесь предложением!",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🎟️ Мои промокоды", callback_data="my_promos")],
            [InlineKeyboardButton(text="⬅️ Профиль", callback_data="profile")],
        ])
    )


@dp.callback_query(F.data == "my_promos")
async def cb_my_promos(call: CallbackQuery) -> None:
    await call.answer()
    all_promos = await db_get_user_promos(call.from_user.id)
    # Реферальные промокоды — только в отдельной вкладке
    user_promos = [up for up in all_promos if up.get("game") != "ref_discount"]

    if not user_promos:
        await send_or_edit(
            call,
            "🎟️ <b>Мои промокоды</b>\n\n"
            "У вас пока нет промокодов.\n"
            "Нажмите «Ввести промокод», чтобы добавить.",
            InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✏️ Ввести промокод", callback_data="enter_promo")],
                [InlineKeyboardButton(text="⬅️ Профиль", callback_data="profile")],
            ])
        )
        return

    rows: list[list[InlineKeyboardButton]] = []
    for up in user_promos:
        used = up.get("used_at") is not None
        valid = _promo_is_valid_now(up) and not used
        if used:
            icon = "✔️"
        elif valid:
            icon = "✅"
        else:
            icon = "🔴"
        if up.get("game") == "pct_discount" and up.get("discount_pct", 0) > 0:
            label = f"{icon} {up['code']} — скидка {up['discount_pct']}% ({up['product_title']})"
        else:
            label = f"{icon} {up['code']} — {up['product_title']} ({up['promo_price']}₽)"
        rows.append([InlineKeyboardButton(text=label, callback_data=f"promo_detail:{up['promo_id']}")])

    rows.append([InlineKeyboardButton(text="🗑️ Удалить использованные", callback_data="del_all_my_promos")])
    rows.append([InlineKeyboardButton(text="✏️ Ввести промокод", callback_data="enter_promo")])
    rows.append([InlineKeyboardButton(text="⬅️ Профиль", callback_data="profile")])

    await send_or_edit(
        call,
        "🎟️ <b>Мои промокоды</b>\n\n"
        "✅ — доступен  |  ✔️ — использован  |  🔴 — недействителен\n\n"
        "Нажмите на промокод, чтобы посмотреть детали.",
        InlineKeyboardMarkup(inline_keyboard=rows)
    )


@dp.callback_query(F.data.startswith("promo_detail:"))
async def cb_promo_detail(call: CallbackQuery) -> None:
    await call.answer()
    try:
        promo_id = int(call.data.split(":")[1])
    except (ValueError, IndexError):
        return

    promo = await db_get_promo_by_id(promo_id)
    if not promo:
        await call.answer("Промокод не найден.", show_alert=True)
        return

    user_promos = await db_get_user_promos(call.from_user.id)
    user_promo = next((up for up in user_promos if up["promo_id"] == promo_id), None)

    if not user_promo:
        await call.answer("Этот промокод не принадлежит вам.", show_alert=True)
        return

    used = user_promo.get("used_at") is not None
    valid_now = _promo_is_valid_now(promo)
    dates_str = _fmt_promo_dates(promo)

    is_pct_type = promo.get("game") == "pct_discount"
    if is_pct_type and promo.get("discount_pct", 0) > 0:
        price_line = f"💸 Скидка: <b>{promo['discount_pct']}%</b>\n"
    else:
        price_line = f"💰 Цена по промокоду: <b>{promo['promo_price']}₽</b>\n"

    text = (
        f"<b>🎟️ Промокод: <code>{escape(promo['code'])}</code></b>\n\n"
        f"🎮 Игра: {escape(promo['game'])}\n"
        f"📦 Товар: {escape(promo['product_title'])}\n"
        f"{price_line}"
        f"📅 Срок: {dates_str}\n\n"
    )

    if used:
        text += "✔️ <b>Промокод уже использован.</b>"
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Мои промокоды", callback_data="my_promos")],
        ])
    elif not valid_now:
        text += "🔴 <b>Промокод недействителен или истёк.</b>"
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Мои промокоды", callback_data="my_promos")],
        ])
    else:
        text += "✅ <b>Промокод активен!</b> Воспользуйтесь предложением ниже."
        if promo.get("game") in ("ref_discount", "pct_discount"):
            kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(
                    text="🎯 Применить скидку к следующей покупке",
                    callback_data=f"activate_ref_promo:{promo_id}",
                )],
                [InlineKeyboardButton(text="⬅️ Мои промокоды", callback_data="my_promos")],
            ])
        else:
            kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🛒 Воспользоваться предложением", callback_data=f"use_promo_confirm:{promo_id}")],
                [InlineKeyboardButton(text="⬅️ Мои промокоды", callback_data="my_promos")],
            ])

    await send_or_edit(call, text, kb)


@dp.callback_query(F.data.startswith("use_promo_confirm:"))
async def cb_use_promo_confirm(call: CallbackQuery) -> None:
    await call.answer()
    try:
        promo_id = int(call.data.split(":")[1])
    except (ValueError, IndexError):
        return

    promo = await db_get_promo_by_id(promo_id)
    if not promo:
        await call.answer("Промокод не найден.", show_alert=True)
        return

    user_promos = await db_get_user_promos(call.from_user.id)
    user_promo = next((up for up in user_promos if up["promo_id"] == promo_id), None)

    if not user_promo or user_promo.get("used_at") is not None:
        await call.answer("Промокод уже использован или недоступен.", show_alert=True)
        return

    if not _promo_is_valid_now(promo):
        await call.answer("Промокод недействителен или истёк.", show_alert=True)
        return

    balance = await db_get_balance(call.from_user.id)
    price = float(promo["promo_price"])

    text = (
        "<b>🛒 Подтверждение покупки по промокоду</b>\n\n"
        f"🎟️ Промокод: <code>{escape(promo['code'])}</code>\n"
        f"🎮 Игра: {escape(promo['game'])}\n"
        f"📦 Товар: {escape(promo['product_title'])}\n"
        f"💰 Сумма: <b>{_fmt_price(price)}₽</b>\n"
        f"💼 Ваш баланс: <b>{_fmt_price(balance)}₽</b>\n\n"
        "Оплата спишется с внутреннего баланса бота.\n"
        "Подтвердите покупку:"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Оплатить", callback_data=f"use_promo_go:{promo_id}")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data=f"promo_detail:{promo_id}")],
    ])
    await send_or_edit(call, text, kb)


@dp.callback_query(F.data.startswith("use_promo_go:"))
async def cb_use_promo_go(call: CallbackQuery) -> None:
    await call.answer()
    try:
        promo_id = int(call.data.split(":")[1])
    except (ValueError, IndexError):
        return

    promo = await db_get_promo_by_id(promo_id)
    if not promo:
        await call.answer("Промокод не найден.", show_alert=True)
        return

    user_promos = await db_get_user_promos(call.from_user.id)
    user_promo = next((up for up in user_promos if up["promo_id"] == promo_id), None)

    if not user_promo or user_promo.get("used_at") is not None:
        await call.answer("Промокод уже использован или недоступен.", show_alert=True)
        return

    if not _promo_is_valid_now(promo):
        await call.answer("Промокод недействителен или истёк.", show_alert=True)
        return

    user = call.from_user
    price = float(promo["promo_price"])
    title = f"[Промокод {promo['code']}] {promo['game']} — {promo['product_title']}"

    if user.id in _processing_payments:
        await call.answer("Платёж уже обрабатывается, подождите...", show_alert=True)
        return
    _processing_payments.add(user.id)
    try:
        ok = await db_try_charge(user.id, price)
        if not ok:
            balance = await db_get_balance(user.id)
            text = (
                "❌ <b>Недостаточно средств на балансе.</b>\n\n"
                f"Сумма заказа: {_fmt_price(price)}₽\n"
                f"Ваш баланс: {_fmt_price(balance)}₽\n\n"
                "Пополните баланс и повторите попытку."
            )
            kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="💳 Пополнить баланс", callback_data="topup")],
                [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
            ])
            await send_or_edit(call, text, kb)
            return

        order_id = await db_create_order(user.id, title, price, status="Оплачен", category="promo")
        await db_add_transaction(user.id, -price, kind="purchase", reason=f"Заказ #{order_id}: {title}")
        await db_use_promo(user.id, promo_id)
    finally:
        _processing_payments.discard(user.id)

    new_balance = await db_get_balance(user.id)

    text = (
        "✅ <b>Оплата прошла успешно. Заказ принят в обработку.</b>\n\n"
        f"🧾 Номер заказа: <code>#{order_id}</code>\n"
        f"🎟️ Промокод: <code>{escape(promo['code'])}</code>\n"
        f"🎮 Игра: {escape(promo['game'])}\n"
        f"📦 Товар: {escape(promo['product_title'])}\n"
        f"💰 Сумма: <b>{_fmt_price(price)}₽</b>\n"
        f"💼 Остаток на балансе: <b>{_fmt_price(new_balance)}₽</b>\n\n"
        "С вами свяжется модератор. Вы также можете написать первым."
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📨 Связаться с модератором", callback_data=f"contact_mod:{order_id}")],
        [InlineKeyboardButton(text="🏠 В меню", callback_data="main")],
    ])
    await send_or_edit(call, text, kb)

    username = f"@{user.username}" if user.username else "—"
    admin_text = (
        f"🎟️ <b>Новый заказ по промокоду!</b>\n\n"
        f"👤 {escape(user.first_name or '?')} ({username})\n"
        f"🆔 TG ID: <code>{user.id}</code>\n"
        f"🧾 Заказ: <code>#{order_id}</code>\n"
        f"🎟️ Промокод: <code>{escape(promo['code'])}</code>\n"
        f"🎮 Игра: {escape(promo['game'])}\n"
        f"📦 Товар: {escape(promo['product_title'])}\n"
        f"💰 Сумма: <b>{_fmt_price(price)}₽</b>"
    )
    admin_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"✅ Выполнен #{order_id}", callback_data=f"mod:done:{order_id}")],
        [InlineKeyboardButton(text="📧 Запросить код из почты", callback_data=f"mod:req_email:{order_id}:{user.id}")],
        [InlineKeyboardButton(text="💸 Возврат", callback_data=f"mod:refund:{order_id}")],
    ])
    await notify_moderator_order(admin_text, reply_markup=admin_kb)


@dp.callback_query(F.data.startswith("mod:done:"))
async def cb_mod_done(call: CallbackQuery) -> None:
    """Модератор отмечает заказ выполненным через кнопку в чате модератора."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    try:
        order_id = int(call.data.split(":")[2])
    except (ValueError, IndexError):
        return

    order = await db_get_order(order_id)
    if not order:
        await call.answer("Заказ не найден.", show_alert=True)
        return
    if order.get("status") == "Выполнен":
        await call.answer("Этот заказ уже отмечен как выполненный.", show_alert=True)
        return

    tg_id = order["tg_id"]
    await db_update_order_status(order_id, "Выполнен")

    referrer_id, level_up, new_level, is_first = await db_process_referral(tg_id)
    if is_first and referrer_id:
        try:
            await bot.send_message(
                referrer_id,
                f"🎉 По вашей реферальной ссылке совершена первая покупка!\n"
                f"{'📈 Ваш уровень повышен до ' + str(new_level) + '!' if level_up else ''}",
            )
        except Exception:
            pass

    try:
        user_kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="✅ Подтвердить получение",
                        callback_data=f"confirm_receipt:{order_id}",
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="❌ Отклонить",
                        callback_data=f"reject_receipt:{order_id}",
                    )
                ],
            ]
        )
        await bot.send_message(
            tg_id,
            f"🎉 <b>Ваш заказ #{order_id} выполнен!</b>\n\n"
            "⚠️ <b>Перед тем как нажать кнопку — проверьте наличие цифрового товара в игре.</b>\n\n"
            "✅ Нажмите <b>«Подтвердить получение»</b>, если товар зачислен.\n"
            "❌ Нажмите <b>«Отклонить»</b>, если товар не поступил. В этом случае потребуется "
            "прислать скриншот из игры в качестве доказательства.\n\n"
            "⚠️ <b>Внимание:</b> ложное отклонение может привести к блокировке в боте.",
            parse_mode="HTML",
            reply_markup=user_kb,
        )
        await db_set_confirm_pending(order_id, True)
        asyncio.create_task(_auto_confirm_receipt(order_id, tg_id))
    except Exception:
        pass

    await call.answer(f"Заказ #{order_id} отмечен выполненным.", show_alert=True)
    try:
        await call.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass

    try:
        await bot.unpin_chat_message(call.message.chat.id, call.message.message_id)
    except Exception:
        pass


@dp.callback_query(F.data.startswith("mod:req_email:"))
async def cb_mod_req_email_code(call: CallbackQuery) -> None:
    """Модератор запрашивает у покупателя код из письма на почте."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    try:
        parts = call.data.split(":")
        order_id = int(parts[2])
        tg_id = int(parts[3])
    except (ValueError, IndexError):
        return

    try:
        await bot.send_message(
            tg_id,
            f"📧 <b>Требуется код из письма</b>\n\n"
            f"По вашему заказу <b>#{order_id}</b> модератор ожидает код подтверждения,\n"
            "который был отправлен на вашу электронную почту.\n\n"
            "Пожалуйста, откройте письмо и пришлите код следующим сообщением.",
            parse_mode="HTML",
        )
        # Ставим покупателю состояние ожидания кода удалённо
        buyer_key = StorageKey(bot_id=bot.id, chat_id=tg_id, user_id=tg_id)
        buyer_fsm = FSMContext(storage=dp.storage, key=buyer_key)
        await buyer_fsm.set_state(ShopStates.waiting_email_code)
        await buyer_fsm.update_data(email_code_order_id=order_id)
        await call.answer("Запрос отправлен покупателю.", show_alert=True)
    except Exception:
        await call.answer("Не удалось отправить уведомление покупателю.", show_alert=True)


@dp.message(ShopStates.waiting_email_code)
async def msg_email_code(message: Message, state: FSMContext) -> None:
    """Покупатель прислал код из почты — пересылаем модератору."""
    data = await state.get_data()
    order_id = data.get("email_code_order_id", "?")
    await state.clear()

    await notify_moderator(
        f"📧 <b>Код из почты по заказу #{order_id}</b>\n\n"
        f"<code>{escape(message.text or '(нет текста)')}</code>"
    )
    await message.answer(
        "✅ Код отправлен модератору. Ожидайте выполнения заказа.",
    )


@dp.callback_query(F.data.startswith("mod:refund:"))
async def cb_mod_refund(call: CallbackQuery) -> None:
    """Модератор делает возврат средств по заказу."""
    await call.answer()
    if not _is_moderator(call.from_user.id):
        return
    try:
        order_id = int(call.data.split(":")[2])
    except (ValueError, IndexError):
        return

    order = await db_get_order(order_id)
    if not order:
        await call.answer("Заказ не найден.", show_alert=True)
        return

    if order.get("status") == "Возврат":
        await call.answer("Возврат по этому заказу уже был выполнен.", show_alert=True)
        return

    tg_id = order["tg_id"]
    refund_amount = float(order["price"])

    await db_update_order_status(order_id, "Возврат")
    await db_add_balance(tg_id, refund_amount)
    await db_add_transaction(
        tg_id, refund_amount, kind="refund",
        reason=f"Возврат по заказу #{order_id}"
    )

    try:
        await bot.send_message(
            tg_id,
            f"💸 <b>Возврат по заказу #{order_id}</b>\n\n"
            f"На ваш баланс возвращено <b>{_fmt_price(refund_amount)}₽</b>.\n\n"
            "Если у вас остались вопросы — обратитесь в поддержку.",
            parse_mode="HTML",
        )
    except Exception:
        pass

    await call.answer(
        f"Возврат {_fmt_price(refund_amount)}₽ по заказу #{order_id} выполнен.",
        show_alert=True,
    )
    try:
        await call.message.edit_text(
            call.message.text + f"\n\n💸 <b>Возврат выполнен</b> модератором.",
            parse_mode="HTML",
            reply_markup=None,
        )
    except Exception:
        try:
            await call.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass

    try:
        await bot.unpin_chat_message(call.message.chat.id, call.message.message_id)
    except Exception:
        pass


# =====================================================================
# Запуск
# =====================================================================


def _acquire_single_instance_lock() -> None:
    """Не даём запустить больше одной копии бота на хосте."""
    import fcntl
    lock_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), ".bot.lock"
    )
    lock_file = open(lock_path, "w")
    try:
        fcntl.flock(lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except OSError:
        logging.error(
            "Другая копия бота уже запущена. Завершаю текущий процесс."
        )
        raise SystemExit(0)
    lock_file.write(str(os.getpid()))
    lock_file.flush()
    globals()["_bot_lock_file"] = lock_file


# =====================================================================
# Ответ покупателю из чата модератора
# =====================================================================


@dp.callback_query(F.data.startswith("reply_buyer:"))
async def cb_reply_buyer(call: CallbackQuery, state: FSMContext) -> None:
    """Модератор нажимает «Ответить покупателю» — бот просит написать текст."""
    if not _is_moderator(call.from_user.id):
        await call.answer("Нет доступа.", show_alert=True)
        return
    buyer_id = int(call.data.split(":")[1])
    await state.set_state(AdminStates.waiting_mod_reply)
    await state.update_data(mod_reply_buyer_id=buyer_id)
    await call.message.answer(
        f"✍️ Напишите сообщение покупателю (поддерживается текст и фото).\n\n"
        f"Для отмены нажмите кнопку ниже.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="❌ Отмена", callback_data="mod_reply_cancel"),
        ]]),
    )
    await call.answer()


@dp.callback_query(F.data == "mod_reply_cancel")
async def cb_mod_reply_cancel(call: CallbackQuery, state: FSMContext) -> None:
    if not _is_moderator(call.from_user.id):
        return
    await state.clear()
    await call.message.edit_text("❌ Ответ покупателю отменён.")
    await call.answer()


@dp.message(AdminStates.waiting_mod_reply)
async def msg_mod_reply(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    data = await state.get_data()
    buyer_id = int(data.get("mod_reply_buyer_id", 0))
    await state.clear()
    try:
        await bot.copy_message(
            chat_id=buyer_id,
            from_chat_id=message.chat.id,
            message_id=message.message_id,
        )
        await message.answer("✅ Сообщение отправлено покупателю.")
    except Exception as e:
        await message.answer(f"❌ Не удалось отправить: {e}")


# =====================================================================
# Реферальная программа — пользовательские callback'и
# =====================================================================


@dp.callback_query(F.data == "referral_info")
async def cb_referral_info(call: CallbackQuery) -> None:
    await call.answer()
    user_row, _ = await db_get_or_create_user(call.from_user)
    level = user_row.get("referral_level", 0)
    ref_count = user_row.get("referral_count", 0)
    ref_code = user_row.get("referral_code") or ""
    discount = _ref_discount_pct(level)
    level_name = _ref_level_name(level)

    # Ссылка вида https://t.me/BOTNAME?start=ref_CODE
    bot_info = await bot.get_me()
    ref_link = f"https://t.me/{bot_info.username}?start=ref_{ref_code}"

    next_level = level + 1
    if next_level <= 5:
        next_refs_needed = next_level * REFERRAL_REFS_PER_LEVEL
        refs_left = next_refs_needed - ref_count
        next_level_name, next_discount = REFERRAL_LEVELS[next_level]
        progress_text = (
            f"\n📈 До уровня <b>{next_level_name}</b>: ещё <b>{refs_left}</b> реферал(ов)"
        )
    else:
        progress_text = "\n🏆 Максимальный уровень достигнут!"

    text = (
        "<b>👥 Реферальная программа</b>\n\n"
        f"Ваш статус: <b>{level_name}</b>\n"
        f"Рефералов засчитано: <b>{ref_count}</b>\n"
        f"Ваша скидка на заказы: <b>{discount}%</b>{' (не действует на Telegram Stars)' if discount else ''}\n"
        f"{progress_text}\n\n"
        f"🔗 Ваша реферальная ссылка:\n<code>{ref_link}</code>\n\n"
        "<b>Как это работает:</b>\n"
        "1. Поделитесь ссылкой с друзьями\n"
        "2. Приглашённый получит 3 промокода на скидку 5%, если перейдёт по вашей ссылке\n"
        "3. Когда приглашённый совершит первую покупку — вам засчитается реферал\n"
        "4. Каждые 2 реферала — новый уровень с бонусными промокодами на скидку\n\n"
        "<b>Уровни:</b>\n"
    )
    for lvl, (name, disc) in REFERRAL_LEVELS.items():
        refs = lvl * REFERRAL_REFS_PER_LEVEL
        current = " ← <b>Вы здесь</b>" if lvl == level else ""
        text += f"• {name} — {refs} реф., скидка {disc}%{current}\n"

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🎟️ Мои реф. промокоды", callback_data="my_ref_promos")],
        [InlineKeyboardButton(text="⬅️ Профиль", callback_data="profile")],
    ])
    await send_or_edit(call, text, kb)


@dp.callback_query(F.data == "my_ref_promos")
async def cb_my_ref_promos(call: CallbackQuery) -> None:
    """Реферальные промокоды пользователя — отдельная вкладка."""
    await call.answer()
    all_promos = await db_get_user_promos(call.from_user.id)
    ref_promos = [up for up in all_promos if up.get("game") == "ref_discount"]

    if not ref_promos:
        await send_or_edit(
            call,
            "🎟️ <b>Мои реферальные промокоды</b>\n\n"
            "У вас пока нет реферальных промокодов.\n"
            "Приглашайте друзей и повышайте уровень!",
            InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="👥 Реф. программа", callback_data="referral_info")],
                [InlineKeyboardButton(text="⬅️ Профиль", callback_data="profile")],
            ])
        )
        return

    rows: list[list[InlineKeyboardButton]] = []
    active_count = 0
    for up in ref_promos:
        used = up.get("used_at") is not None
        valid = _promo_is_valid_now(up) and not used
        disc = up.get("discount_pct", 0)
        if used:
            label = f"✔️ {up['code']} — {disc}% (использован)"
            rows.append([InlineKeyboardButton(text=label, callback_data="referral_info")])
        elif valid:
            active_count += 1
            label = f"✅ {up['code']} — скидка {disc}%"
            rows.append([InlineKeyboardButton(text=label, callback_data=f"activate_ref_promo:{up['promo_id']}")])
        else:
            label = f"🔴 {up['code']} — {disc}% (недействителен)"
            rows.append([InlineKeyboardButton(text=label, callback_data="referral_info")])

    if active_count > 0:
        rows.append([InlineKeyboardButton(
            text=f"🗑️ Удалить все активные ({active_count})",
            callback_data="del_all_ref_promos",
        )])
    rows.append([InlineKeyboardButton(text="👥 Реф. программа", callback_data="referral_info")])
    rows.append([InlineKeyboardButton(text="⬅️ Профиль", callback_data="profile")])

    await send_or_edit(
        call,
        f"🎟️ <b>Мои реферальные промокоды</b>\n\n"
        f"✅ — активен (нажмите, чтобы применить к покупке)\n"
        f"✔️ — использован  |  🔴 — недействителен\n\n"
        f"Всего: <b>{len(ref_promos)}</b>, активных: <b>{active_count}</b>",
        InlineKeyboardMarkup(inline_keyboard=rows)
    )


@dp.callback_query(F.data == "del_all_ref_promos")
async def cb_del_all_ref_promos(call: CallbackQuery) -> None:
    """Пользователь удаляет все свои активные реферальные промокоды."""
    deleted = await db_delete_all_user_ref_promos(call.from_user.id)
    await call.answer(f"Удалено промокодов: {deleted}", show_alert=True)
    await send_or_edit(
        call,
        "🎟️ <b>Мои реферальные промокоды</b>\n\n"
        "Активных реферальных промокодов нет.\n"
        "Приглашайте друзей и повышайте уровень!",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="👥 Реф. программа", callback_data="referral_info")],
            [InlineKeyboardButton(text="⬅️ Профиль", callback_data="profile")],
        ])
    )


@dp.callback_query(F.data == "del_all_my_promos")
async def cb_del_all_my_promos(call: CallbackQuery) -> None:
    """Пользователь удаляет все использованные промокоды из своего инвентаря."""
    deleted = await db_delete_all_used_user_promos(call.from_user.id)
    await call.answer(f"Удалено использованных промокодов: {deleted}", show_alert=True)
    await send_or_edit(
        call,
        "🎟️ <b>Мои промокоды</b>\n\n"
        "У вас нет промокодов.\n"
        "Нажмите «Ввести промокод», чтобы добавить.",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✏️ Ввести промокод", callback_data="enter_promo")],
            [InlineKeyboardButton(text="⬅️ Профиль", callback_data="profile")],
        ])
    )


@dp.callback_query(F.data.startswith("activate_ref_promo:"))
async def cb_activate_ref_promo(call: CallbackQuery) -> None:
    await call.answer()
    try:
        promo_id = int(call.data.split(":")[1])
    except (ValueError, IndexError):
        return

    promo = await db_get_promo_by_id(promo_id)
    if not promo or promo.get("game") not in ("ref_discount", "pct_discount"):
        await call.answer("Промокод не найден.", show_alert=True)
        return

    user_promos = await db_get_user_promos(call.from_user.id)
    user_promo = next((up for up in user_promos if up["promo_id"] == promo_id), None)
    if not user_promo or user_promo.get("used_at") is not None:
        await call.answer("Промокод уже использован или недоступен.", show_alert=True)
        return

    await db_set_active_ref_promo(call.from_user.id, promo_id)

    discount = promo.get("discount_pct", 0)
    text = (
        f"✅ <b>Скидка {discount}% активирована!</b>\n\n"
        f"При следующей покупке (кроме Telegram Stars) скидка применится автоматически.\n\n"
        f"Промокод: <code>{promo['code']}</code>\n"
        f"Скидка действует на один заказ.\n\n"
        f"Перейдите в меню и выберите товар."
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🛒 В магазин", callback_data="shop")],
        [InlineKeyboardButton(text="🏠 Главное меню", callback_data="main")],
    ])
    await send_or_edit(call, text, kb)


# =====================================================================
# Рассылка всем пользователям (только модератор)
# =====================================================================


@dp.message(Command("broadcast"))
async def cmd_broadcast(message: Message, state: FSMContext) -> None:
    """Команда /broadcast — начать рассылку всем пользователям."""
    if not _is_moderator(message.from_user.id):
        return
    await state.set_state(AdminStates.waiting_broadcast)
    user_count = await db_users_count()
    await message.answer(
        f"📢 <b>Рассылка</b>\n\n"
        f"Всего получателей: <b>{user_count}</b> чел.\n\n"
        f"Пришлите сообщение (текст, фото, видео — любой формат).\n"
        f"Оно будет отправлено каждому пользователю бота.\n\n"
        f"Для отмены: /cancel",
        parse_mode="HTML",
    )


@dp.message(Command("cancel"), AdminStates.waiting_broadcast)
async def cmd_broadcast_cancel(message: Message, state: FSMContext) -> None:
    if not _is_moderator(message.from_user.id):
        return
    await state.clear()
    await message.answer("❌ Рассылка отменена.")


@dp.message(AdminStates.waiting_broadcast)
async def msg_broadcast(message: Message, state: FSMContext) -> None:
    """Получает сообщение от модератора и рассылает всем пользователям."""
    if not _is_moderator(message.from_user.id):
        return
    await state.clear()
    user_ids = await db_get_all_user_ids()
    sent = 0
    failed = 0
    status_msg = await message.answer(
        f"⏳ Рассылка запущена — {len(user_ids)} получателей..."
    )
    for uid in user_ids:
        try:
            await bot.copy_message(
                chat_id=uid,
                from_chat_id=message.chat.id,
                message_id=message.message_id,
            )
            sent += 1
        except Exception:
            failed += 1
        await asyncio.sleep(0.05)  # не превышаем лимиты Telegram (20 msg/s)
    await status_msg.edit_text(
        f"✅ <b>Рассылка завершена</b>\n\n"
        f"📨 Отправлено: <b>{sent}</b>\n"
        f"❌ Ошибок: <b>{failed}</b>",
        parse_mode="HTML",
    )


# =====================================================================
# Экспорт данных (только модератор)
# =====================================================================


def _xlsx_style_sheet(ws, headers: list[str], rows: list[list], header_hex: str) -> None:
    """Оформляет лист Excel: цветная шапка, автоширина колонок, зебра-подсветка, рамки, freeze panes."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra_fill = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(rows, start=2):
        ws.append(row)
        for cell in ws[r_idx]:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if r_idx % 2 == 0:
                cell.fill = zebra_fill

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row in rows:
            val = row[col_idx - 1]
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


@dp.message(Command("export"))
async def cmd_export(message: Message) -> None:
    """Команда /export — красивая выгрузка всех данных магазина в Excel (только модератор)."""
    if not _is_moderator(message.from_user.id):
        return

    wait_msg = await message.answer("⏳ Собираю данные и формирую отчёт, подождите...")

    pool = await get_pool()

    # Тестовый аккаунт модератора исключаем из всей статистики —
    # его транзакции/заказы не отражают реальную активность магазина.
    users = await pool.fetch(
        "SELECT tg_id, username, first_name, balance, is_blacklisted, "
        "referral_level, referral_count, created_at "
        "FROM users WHERE tg_id != $1 ORDER BY tg_id",
        MODERATOR_CHAT_ID,
    )
    orders = await pool.fetch(
        "SELECT id, tg_id, title, price, status, category, contact, login_data, created_at "
        "FROM orders WHERE tg_id != $1 ORDER BY id",
        MODERATOR_CHAT_ID,
    )
    transactions = await pool.fetch(
        "SELECT id, tg_id, amount, kind, reason, created_at "
        "FROM transactions WHERE tg_id != $1 ORDER BY id",
        MODERATOR_CHAT_ID,
    )

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---------- Лист "Сводка" ----------
    ws_sum = wb.active
    ws_sum.title = "📊 Сводка"

    total_balance = sum(float(u["balance"] or 0) for u in users)
    blacklisted_count = sum(1 for u in users if u["is_blacklisted"])
    total_orders = len(orders)
    completed_orders = sum(1 for o in orders if o["status"] == "Выполнен")
    paid_orders = sum(1 for o in orders if o["status"] == "Оплачен")
    refunded_orders = sum(1 for o in orders if "озврат" in (o["status"] or ""))
    revenue = sum(float(o["price"] or 0) for o in orders if o["status"] in ("Выполнен", "Оплачен"))
    topups = sum(float(t["amount"] or 0) for t in transactions if t["kind"] == "topup")
    purchases_sum = sum(-float(t["amount"] or 0) for t in transactions if t["kind"] == "purchase")
    avg_check = revenue / completed_orders if completed_orders else 0.0
    success_rate = (completed_orders / total_orders * 100) if total_orders else 0.0

    product_sales: dict[str, int] = {}
    for o in orders:
        if o["status"] in ("Выполнен", "Оплачен"):
            product_sales[o["title"] or "—"] = product_sales.get(o["title"] or "—", 0) + 1
    top_products = sorted(product_sales.items(), key=lambda kv: kv[1], reverse=True)[:10]

    now_str = datetime.now(MSK_TZ).strftime("%Y-%m-%d_%H-%M")

    ws_sum["A1"] = "🛍️ Отчёт по магазину"
    ws_sum["A1"].font = Font(size=20, bold=True, color="2F5597")
    ws_sum.merge_cells("A1:D1")
    ws_sum["A2"] = f"Сформировано: {now_str} МСК"
    ws_sum["A2"].font = Font(size=10, italic=True, color="808080")
    ws_sum.merge_cells("A2:D2")

    kpi_rows = [
        ("👥 Всего пользователей", len(users)),
        ("🚫 В чёрном списке", blacklisted_count),
        ("💰 Суммарный баланс пользователей (₽)", round(total_balance, 2)),
        ("🛒 Всего заказов", total_orders),
        ("✅ Выполнено заказов", completed_orders),
        ("🕓 Ожидают выполнения (Оплачен)", paid_orders),
        ("💸 Возвраты", refunded_orders),
        ("📊 Успешность заказов (%)", round(success_rate, 1)),
        ("🧮 Средний чек (₽)", round(avg_check, 2)),
        ("📈 Выручка по заказам (₽)", round(revenue, 2)),
        ("➕ Пополнений баланса (₽)", round(topups, 2)),
        ("➖ Списано на покупки (₽)", round(purchases_sum, 2)),
        ("🧾 Всего транзакций", len(transactions)),
    ]
    start_row = 4
    ws_sum[f"A{start_row}"] = "Показатель"
    ws_sum[f"B{start_row}"] = "Значение"
    for col in ("A", "B"):
        c = ws_sum[f"{col}{start_row}"]
        c.font = Font(color="FFFFFF", bold=True)
        c.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
    for i, (label, value) in enumerate(kpi_rows, start=start_row + 1):
        ws_sum[f"A{i}"] = label
        ws_sum[f"B{i}"] = value
        if i % 2 == 0:
            for col in ("A", "B"):
                ws_sum[f"{col}{i}"].fill = PatternFill(
                    start_color="F2F6FC", end_color="F2F6FC", fill_type="solid"
                )
    ws_sum.column_dimensions["A"].width = 42
    ws_sum.column_dimensions["B"].width = 20

    # Круговая диаграмма по статусам заказов
    status_counts: dict[str, int] = {}
    for o in orders:
        status_counts[o["status"] or "—"] = status_counts.get(o["status"] or "—", 0) + 1
    if status_counts:
        chart_start = start_row + len(kpi_rows) + 3
        ws_sum[f"A{chart_start}"] = "Статус заказа"
        ws_sum[f"B{chart_start}"] = "Количество"
        for i, (status, count) in enumerate(status_counts.items(), start=chart_start + 1):
            ws_sum[f"A{i}"] = status
            ws_sum[f"B{i}"] = count
        pie = PieChart()
        pie.title = "Заказы по статусам"
        data = Reference(ws_sum, min_col=2, min_row=chart_start, max_row=chart_start + len(status_counts))
        cats = Reference(ws_sum, min_col=1, min_row=chart_start + 1, max_row=chart_start + len(status_counts))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        pie.height = 8
        pie.width = 12
        ws_sum.add_chart(pie, f"D{start_row}")

    # Таблица "ТОП товаров"
    if top_products:
        top_start = chart_start + len(status_counts) + 3 if status_counts else start_row + len(kpi_rows) + 3
        ws_sum[f"A{top_start}"] = "🏆 ТОП товаров"
        ws_sum[f"A{top_start}"].font = Font(size=13, bold=True, color="2F5597")
        ws_sum.merge_cells(f"A{top_start}:B{top_start}")

        header_row = top_start + 1
        ws_sum[f"A{header_row}"] = "Товар"
        ws_sum[f"B{header_row}"] = "Продаж"
        for col in ("A", "B"):
            c = ws_sum[f"{col}{header_row}"]
            c.font = Font(color="FFFFFF", bold=True)
            c.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            c.alignment = Alignment(horizontal="center")

        for i, (product, count) in enumerate(top_products, start=header_row + 1):
            ws_sum[f"A{i}"] = product
            ws_sum[f"B{i}"] = count
            if i % 2 == 0:
                for col in ("A", "B"):
                    ws_sum[f"{col}{i}"].fill = PatternFill(
                        start_color="FCE9DA", end_color="FCE9DA", fill_type="solid"
                    )

    # ---------- Лист "Пользователи" ----------
    ws_u = wb.create_sheet("👥 Пользователи")
    u_headers = ["TG ID", "Username", "Имя", "Баланс (₽)", "Реф. уровень", "Реф. кол-во", "Чёрный список", "Дата регистрации"]
    u_rows = [
        [
            u["tg_id"],
            f"@{u['username']}" if u["username"] else "—",
            u["first_name"] or "—",
            round(float(u["balance"] or 0), 2),
            u["referral_level"] or 0,
            u["referral_count"] or 0,
            "🚫 Да" if u["is_blacklisted"] else "Нет",
            _fmt_msk(u["created_at"]),
        ]
        for u in users
    ]
    _xlsx_style_sheet(ws_u, u_headers, u_rows, "2F5597")

    # ---------- Лист "Заказы" ----------
    ws_o = wb.create_sheet("🛒 Заказы")
    o_headers = ["ID", "TG ID", "Товар", "Цена (₽)", "Статус", "Категория", "Контакт", "Данные входа", "Дата"]
    o_rows = [
        [
            o["id"],
            o["tg_id"],
            o["title"],
            round(float(o["price"] or 0), 2),
            o["status"],
            o["category"] or "—",
            o["contact"] or "—",
            o["login_data"] or "—",
            _fmt_msk(o["created_at"]),
        ]
        for o in orders
    ]
    _xlsx_style_sheet(ws_o, o_headers, o_rows, "1F7A46")
    status_col = o_headers.index("Статус") + 1
    status_colors = {
        "Выполнен": "C6EFCE",
        "Оплачен": "FFEB9C",
    }
    from openpyxl.styles import PatternFill as _PF
    for r_idx, o in enumerate(orders, start=2):
        color = status_colors.get(o["status"])
        if "озврат" in (o["status"] or ""):
            color = "FFC7CE"
        if color:
            ws_o.cell(row=r_idx, column=status_col).fill = _PF(
                start_color=color, end_color=color, fill_type="solid"
            )

    # ---------- Лист "Транзакции" ----------
    ws_t = wb.create_sheet("💳 Транзакции")
    t_headers = ["ID", "TG ID", "Сумма (₽)", "Тип", "Причина", "Дата"]
    t_rows = [
        [
            t["id"],
            t["tg_id"],
            round(float(t["amount"] or 0), 2),
            t["kind"],
            t["reason"] or "—",
            _fmt_msk(t["created_at"]),
        ]
        for t in transactions
    ]
    _xlsx_style_sheet(ws_t, t_headers, t_rows, "7B3FA0")
    amount_col = t_headers.index("Сумма (₽)") + 1
    for r_idx, t in enumerate(transactions, start=2):
        cell = ws_t.cell(row=r_idx, column=amount_col)
        cell.font = Font(color="1F7A46" if float(t["amount"] or 0) >= 0 else "C00000", bold=True)

    if orders:
        bar = BarChart()
        bar.title = "Заказы по категориям"
        bar.type = "col"
        cat_counts: dict[str, int] = {}
        for o in orders:
            cat_counts[o["category"] or "—"] = cat_counts.get(o["category"] or "—", 0) + 1
        helper_row = len(o_rows) + 4
        ws_o.cell(row=helper_row, column=1, value="Категория")
        ws_o.cell(row=helper_row, column=2, value="Кол-во")
        for i, (cat, count) in enumerate(cat_counts.items(), start=helper_row + 1):
            ws_o.cell(row=i, column=1, value=cat)
            ws_o.cell(row=i, column=2, value=count)
        data = Reference(ws_o, min_col=2, min_row=helper_row, max_row=helper_row + len(cat_counts))
        cats = Reference(ws_o, min_col=1, min_row=helper_row + 1, max_row=helper_row + len(cat_counts))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 8
        bar.width = 14
        ws_o.add_chart(bar, f"K2")

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    await wait_msg.delete()
    await message.answer_document(
        BufferedInputFile(xlsx_bytes, filename=f"shop_export_{now_str}.xlsx"),
        caption=(
            f"📊 <b>Полный отчёт по магазину</b>\n\n"
            f"👥 Пользователей: <b>{len(users)}</b>\n"
            f"🛒 Заказов: <b>{len(orders)}</b> (✅ {completed_orders} / 🕓 {paid_orders} / 💸 {refunded_orders})\n"
            f"📊 Успешность заказов: <b>{success_rate:.1f}%</b>\n"
            f"🧮 Средний чек: <b>{_fmt_price(avg_check)}₽</b>\n"
            f"💳 Транзакций: <b>{len(transactions)}</b>\n"
            f"📈 Выручка: <b>{_fmt_price(revenue)}₽</b>\n\n"
            f"🕒 Сформировано: {now_str} МСК"
        ),
        parse_mode="HTML",
    )


async def _health_server() -> None:
    """Маленький HTTP-сервер для health-check на Render Web Service."""
    from aiohttp import web

    async def handle(request: web.Request) -> web.Response:
        return web.Response(text="ok")

    port = int(os.environ.get("PORT", 8000))
    app = web.Application()
    app.router.add_get("/", handle)
    app.router.add_get("/health", handle)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", port)
    await site.start()
    logging.info("Health-check сервер запущен на порту %s", port)
    await asyncio.Event().wait()


@dp.errors()
async def global_error_handler(event: ErrorEvent) -> bool:
    """Глобальный обработчик ошибок: логирует падение хендлера,
    уведомляет модератора и мягко отвечает пользователю, чтобы
    сбои не оставались незамеченными молча."""
    update = event.update
    exception = event.exception
    logging.exception(
        "Необработанная ошибка при обработке апдейта %s: %s", update, exception
    )

    try:
        error_text = (
            f"⚠️ <b>Ошибка в боте</b>\n\n"
            f"<code>{escape(str(exception))[:500]}</code>\n\n"
            f"Тип апдейта: <code>{escape(update.event_type or '?')}</code>"
        )
        await notify_moderator(error_text)
    except Exception as e:
        logging.warning(f"Не удалось уведомить модератора об ошибке: {e}")

    try:
        call = update.callback_query
        if call is not None:
            try:
                await call.answer(
                    "⚠️ Произошла ошибка. Попробуйте ещё раз или обратитесь в поддержку.",
                    show_alert=True,
                )
            except Exception:
                pass
        else:
            message = update.message
            chat_id = message.chat.id if message else None
            if chat_id is not None:
                await bot.send_message(
                    chat_id,
                    "⚠️ Произошла ошибка при обработке запроса. "
                    "Попробуйте ещё раз или обратитесь в поддержку.",
                )
    except Exception as e:
        logging.warning(f"Не удалось уведомить пользователя об ошибке: {e}")

    return True


async def main() -> None:
    _acquire_single_instance_lock()
    await db_init()
    logging.info("Бот запускается...")
    await bot.delete_webhook(drop_pending_updates=True)
    await asyncio.gather(
        dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types()),
        _health_server(),
    )


if __name__ == "__main__":
    asyncio.run(main())
